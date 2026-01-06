from django.shortcuts import render, redirect
from main.models import Produit, Mark, Category, Supplier, Stockin, Itemsbysupplier, Client, Represent, Order, Orderitem, Clientprices, Bonlivraison, Facture, Outfacture, Livraisonitem, PaymentClientbl, PaymentClientfc,  PaymentSupplier, Bonsregle, Returnedsupplier, Avoirclient, Returned, Avoirsupplier, Orderitem, Carlogos, Ordersnotif, Connectedusers, Promotion, UserSession, Refstats, Notavailable, Cart, Wich, Wishlist, Notification, Modifierstock, Cartitems, Notesrepresentant, Achathistory, Excelecheances, Tva, Etude, EtudeItem, Setting
from django.contrib.auth import logout
from django.http import JsonResponse, HttpResponse
import openpyxl
# import Count
from django.contrib.auth.models import User
from django.db.models import Count, F, Sum, Q, ExpressionWrapper, Func, fields, IntegerField
from django.db.models.functions import Cast
from django.contrib.sessions.models import Session
from functools import wraps
from django.contrib.auth.decorators import user_passes_test
import json
from django.contrib.auth.models import Group
from django.db import transaction
from datetime import datetime, date, timedelta
from django.utils import timezone
import pandas as pd
from itertools import chain, groupby
from django.core.serializers import serialize
import re
from django.contrib.auth import authenticate, login, logout
import requests as req
from collections import defaultdict
import calendar
from django.db.models.functions import TruncDay
import uuid
today = timezone.now().date()
thisyear=timezone.now().year


def isadmin(user):
    if not user.groups.filter(name='admin').exists():
        redirect('main:logoutuser')

# call isadmin here so that it executes before executing all the views

def getproductsbycategory(request):
    # category = Category.objects.get(pk=request.POST.get('category'))
    # products = category.product.filter(category=category)[:10]
    # get ten products from the category
    products = Produit.objects.filter(category__pk=request.GET.get('category')).order_by('code')
    facture=request.GET.get('facture')=='1'
    print('>> facture', facture)
    # get marks of the products filtered
    marks = Mark.objects.filter(produit__in=products).distinct().annotate(count=Count('produit'))
    ctx={
        'products':products,
        'home':False,
        'marks':marks,
        'facture':facture
    }

    return JsonResponse({
        'data':render(request, 'product_search.html', ctx).content.decode('utf-8'),
        'stocktotal':products.aggregate(Sum('stocktotal'))['stocktotal__sum'] or 0,
        'stockfacture':products.aggregate(Sum('stockfacture'))['stockfacture__sum'] or 0,
    })

def adminaddproductpage(request):
    categories=Category.objects.all()
    marques=Mark.objects.all()
    ctx={'categories':categories,
         'marques':marques,
         'commercials':Represent.objects.all(),
         'carlogos':Carlogos.objects.all(),
        }
    return render(request, 'addproduct.html', ctx)



def categoriespage(request):
    ctx={
        'categories':Category.objects.all().order_by('code'),
        'commercials':Represent.objects.all(),
        'title':'Categories'
    }
    return render(request, 'categories.html', ctx)

def createcategory(request):
    name=request.POST.get('categoryname')
    code=request.POST.get('categorycode')
    affichage=request.POST.get('categoryaffichage')
    hideclient=request.POST.get('hideclient')=='True'
    commercialexcluded=request.POST.getlist('commercialexcluded')
    reps=Represent.objects.filter(pk__in=commercialexcluded)
    # get image file
    image=request.FILES.get('categoryimage')
    files={'image':image}
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    if serverip:
        try:
            res=req.post(f'http://{serverip}/products/createcategory', {
                'name':name,
                'code':code,
                'affichage':affichage,
                'hideclient':hideclient,
                'commercialexcluded':commercialexcluded,
                # get image file
            }, files=files)
            res.raise_for_status()
        except req.exceptions.RequestException as e:
            return JsonResponse({
                'success':False,
                'error':'Error in request to the server'
            })
        
    category=Category.objects.create(name=name, image=image, code=code, affichage=affichage, masqueclients=hideclient)
    # create category
    # try:
    #     res=req.get(f'http://{serverip}/products/createcategory', {

    #         'name':name,
    #         'code':code,
    #         'affichage':affichage,
    #         'hideclient':hideclient,
    #         'commercialexcluded':commercialexcluded,
    #         # get image file
    #         #'image':category.image.url.replace('/media/', '') if category.image else ''
    #     })
    #     res.raise_for_status()
    # except req.exceptions.RequestException as e:
    #     print('Error in request:', e)
    #     return JsonResponse({
    #         'success':False,
    #         'error':'Error in request to the server'
    #     })
    
    if len(commercialexcluded) > 0:
        category.excludedrep.set(reps)
    ctx={
        'categories':Category.objects.all().order_by('code'),
        'title':'Categories'
    }
    # print({

    #     'name':name,
    #     'code':code,
    #     'affichage':affichage,
    #     'hideclient':hideclient,
    #     'commercialexcluded':commercialexcluded,
    #     # get image file
    #     'image':category.image.url.replace('/media/', '') if category.image else ''
    # })
    # req.get(f'http://{serverip}/products/createcategory', {

    #     'name':name,
    #     'code':code,
    #     'affichage':affichage,
    #     'hideclient':hideclient,
    #     'commercialexcluded':commercialexcluded,
    #     # get image file
    #     'image':category.image.url.replace('/media/', '') if category.image else ''
    # })
    return JsonResponse({
        'html':render(request, 'categories.html', ctx).content.decode('utf-8')
    })

def updatecategory(request):
    print(request.POST.get('updatecategoryaffichage'))
    image=request.FILES.get('updatecategoryimage') or None
    id=request.POST.get('id')
    hideclient=request.POST.get('hideclient')=='True'
    print('>>>>>>>in updtae',request.POST.get('hideclient'))
    commercialexcluded=request.POST.getlist('commercialexcluded')
    reps=Represent.objects.filter(pk__in=commercialexcluded)
    category=Category.objects.get(pk=id)
    files={'image':image}
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    if serverip:
        try:
            res=req.post(f'http://{serverip}/products/updatecategory', {
                'id':id,
                'hideclient':hideclient,
                'commercialexcluded':commercialexcluded,
                'name':request.POST.get('updatecategoryname'),
                'code':request.POST.get('updatecategorycode'),
                'affichage':request.POST.get('updatecategoryaffichage')
            }, files=files)
            res.raise_for_status()
        except req.exceptions.RequestException as e:
            print('>>> error', e)
            return JsonResponse({
                'success':False,
                'error':'Error in request to the server'
            })
    category.masqueclients=hideclient
    category.excludedrep.clear()
    category.excludedrep.set(reps)
    category.name=request.POST.get('updatecategoryname')
    category.code=request.POST.get('updatecategorycode')
    category.affichage=request.POST.get('updatecategoryaffichage')
    if image:
        category.image=image

    category.save()
    ctx={
        'categories':Category.objects.all().order_by('code'),
        'title':'Categories'
    }
    # req.get(f'http://{serverip}/products/updatecategory', {
    #     'id':id,
    #     'image':category.image.url.replace('/media/', '') if category.image else '',
    #     'hideclient':hideclient,
    #     'commercialexcluded':commercialexcluded,
    #     'name':request.POST.get('updatecategoryname'),
    #     'code':request.POST.get('updatecategorycode'),
    #     'affichage':request.POST.get('updatecategoryaffichage'),
    # })
    return JsonResponse({
        'html':render(request, 'categories.html', ctx).content.decode('utf-8')
    })


def marquespage(request):
    ctx={
        'marques':Mark.objects.all(),
        'title':'List des marques',
        'commercials':Represent.objects.all()
    }
    return render(request, 'marquesadmin.html', ctx)

def createmarque(request):
    name=request.POST.get('marquename')
    # get image file
    image=request.FILES.get('marqueimage')
    # create category
    hideclient=request.POST.get('hideclientmrk')=='True'
    commercialexcluded=request.POST.getlist('commercialexcludedmrk')
    reps=Represent.objects.filter(pk__in=commercialexcluded)
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    mrq=Mark.objects.create(name=name, image=image, masqueclients=hideclient)
    if serverip:
        req.get(f'http://{serverip}/products/createmarque', {
            'name':name,
            'hideclient':hideclient,
            'commercialexcluded':commercialexcluded,
            # get image file
            'image':mrq.image.url.replace('/media/', '') if mrq.image else ''
        })
    if len(commercialexcluded) > 0:
        mrq.excludedrep.set(reps)
    return JsonResponse({
        'success':True
    })

def updatemarque(request):
    image=request.FILES.get('image') or None
    id=request.POST.get('id')
    hideclient=request.POST.get('hideclientmrk')=='True'
    commercialexcluded=request.POST.getlist('commercialexcludedmrk')
    reps=Represent.objects.filter(pk__in=commercialexcluded)
    mark=Mark.objects.get(pk=id)
    files={'image':image}
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    if serverip:
        try:
            res=req.post(f'http://{serverip}/products/updatemarque', {
                'id':id,
                'hideclient':hideclient,
                'commercialexcluded':commercialexcluded,
                'name':request.POST.get('name')
            }, files=files)
            res.raise_for_status()
        except req.exceptions.RequestException as e:
            return JsonResponse({
                'success':False,
                'error':'Error in request to the server'
            })
    mark.name=request.POST.get('name')
    mark.masqueclients=hideclient
    mark.excludedrep.set(reps)
    if image:
        mark.image=image
    mark.save()
    req.get(f'http://{serverip}/products/updatemarque', {
        'id':id,
        'name':request.POST.get('name'),
        'hideclient':hideclient,
        'commercialexcluded':commercialexcluded,
        # get image file
        'image':mark.image.url.replace('/media/', '') if mark.image else ''
    })
    ctx={
        'marques':Mark.objects.all(),
        'title':'List des marques'
    }
    return JsonResponse({
        'html':render(request, 'marques.html', ctx).content.decode('utf-8')
    })

def checkref(request):
    ref=request.POST.get('ref').lower().strip()
    product=Produit.objects.filter(ref=ref)
    print(ref)
    if product:
        return JsonResponse({
            'exist':True
        })
    return JsonResponse({
        'exist':False
    })

    # print(ref, category, product)
    # if product:
    # else:
    #     return JsonResponse({
    #         'exist':False
    #     })

def supplierspage(request):
    ctx={
        'suppliers':Supplier.objects.all(),
        'title':'List des fournisseurs'
    }
    return render(request, 'suppliers.html', ctx)

def addsupplier(request):
    name=request.POST.get('suppnameinp')
    phone=request.POST.get('suppphone')
    address=request.POST.get('address')
    Supplier.objects.create(
        name=name,
        phone=phone,
        address=address
    )
    ctx={
        'suppliers':Supplier.objects.all(),
        'title':'List des fournisseurs'
    }
    return JsonResponse({
        'html':render(request, 'suppliers.html', ctx).content.decode('utf-8')
    })

def getsupplierdata(request):
    id=request.POST.get('id')
    supplier=Supplier.objects.get(pk=id)
    return JsonResponse({
        'name':supplier.name,
        'phone':supplier.phone,
        'address':supplier.address,
        'id':supplier.id
    })


def updatesupplier(request):
    id=request.POST.get('pid')
    name=request.POST.get('pname')
    phone=request.POST.get('pphone')
    address=request.POST.get('paddress')
    supplier=Supplier.objects.get(pk=id)
    supplier.name=name
    supplier.phone=phone
    supplier.address=address
    supplier.save()
    ctx={
        'suppliers':Supplier.objects.all(),
        'title':'List des fournisseurs'
    }
    return JsonResponse({
        'html':render(request, 'suppliers.html', ctx).content.decode('utf-8')
    })
# @user_passes_test(isadmin, login_url='main:loginpage')
# @login_required(login_url='main:loginpage')
def system(request):

    ctx={
        'title':'Dashboard',




    }
    return render(request, 'admindashboard.html', ctx)
def addoneproduct(request):
    lastid=0 # zero cause we will add one
    if Produit.objects.last():
        lastid=Produit.objects.last().id
    uniqcode=f'pdct{lastid+1}'
    ref=request.POST.get('refinadd').lower().strip().replace('§', '-').replace("'", '')
    name=request.POST.get('nameinadd').strip()
    category=request.POST.get('categoryinadd')
    commercialsprix=request.POST.get('commercialsprix') or "[]"
    mark=request.POST.get('marqueinadd') or None
    logo=request.POST.get('logoinadd', None)
    image=request.FILES.get('imageinadd', None)
    sellprice=request.POST.get('sellpriceinadd') or 0
    supplier=request.POST.get('supplier') or None
    minstock=request.POST.get('minstock') or 0
    buyprice=request.POST.get('buyprice') or 0
    remise=request.POST.get('remiseinadd') or 0
    diametre=request.POST.get('diametreinadd') or ''
    representprice=request.POST.get('repprice') or None
    code=request.POST.get('codeinadd') or ''
    block=request.POST.get('blockinadd') or ''
    equivalent=request.POST.get('equivinadd') or ''
    cars=request.POST.get('carsinadd') or ''
    netprice=round(float(sellprice)-(float(sellprice)*float(remise)/100), 2)
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    if serverip:
        try:
            res=req.get(f'http://{serverip}/products/addoneproduct', {
                'ref':ref,
                'name':name,
                'buyprice':buyprice,
                'diametre':diametre,
                'sellprice':sellprice,
                'remise':remise,
                'prixnet':netprice,
                'representprice':representprice,
                'minstock':minstock,
                'equivalent':equivalent,
                'cars':cars,
                'category':category,
                'supplier':supplier,
                'mark':mark,
                'image':'',
                'code':code,
                'repsprice':commercialsprix,
                'block':block,
                'carlogos_id':logo,
                'stocktotal':0,
                'stockfacture':0,
                'uniqcode':uniqcode
            })
            res.raise_for_status()

        except req.exceptions.RequestException as e:
            print('>>> error', e)
            return JsonResponse({
                'success':False,
                'error':f'error {e}'
            })
        # create product
    product=Produit.objects.create(
        ref=ref,
        uniqcode=uniqcode,
        name=name,
        buyprice=0,
        diametre=diametre,
        sellprice=sellprice,
        remise=remise,
        prixnet=netprice,
        representprice=representprice,
        minstock=minstock,
        equivalent=equivalent,
        cars=cars,
        category_id=category,
        supplier_id=supplier,
        mark_id=mark,
        image=image,
        code=code,
        repsprice=commercialsprix,
        block=block,
        #carlogos_id=logo,
        stocktotal=0,
        stockfacture=0,
        isactive=True
    )
    # if image:
    #     image=product.image.url.replace('/media/', '')
    # the product is created in the server without the image, you can send another request to assign the image if it's selected by the user


    return JsonResponse({
        'success':True,

    })
    


def viewoneproduct(request, id):
    product=Produit.objects.get(pk=id)
    commercial_prices = product.getcommercialsprice()  # Note the parentheses ()
    # stockin will exclude bons with ismanuel is True because they dont add to stock, only add to stock facture
    # nbon__ismanual=False this filters the unmanual bons #here
    stockin=Stockin.objects.filter(product=product, nbon__ismanual=False)|Stockin.objects.filter(product=product, isinventaire=True).order_by('-date')
    # outs in bon livraisons
    outbl=Livraisonitem.objects.filter(product=product, isfacture=False).aggregate(Sum('qty'))['qty__sum'] or 0
    # outs in factures, exclude the factures created manually
    outfacture=Outfacture.objects.filter(product=product, facture__bon__isnull=False).aggregate(Sum('qty'))['qty__sum'] or 0
    revbl=Livraisonitem.objects.filter(product=product, isfacture=False).aggregate(Sum('total'))['total__sum'] or 0
    revfacture=Outfacture.objects.filter(product=product, facture__bon__isnull=False).aggregate(Sum('total'))['total__sum'] or 0
    # avoir supplier is needed to be with out items
    avoirsupp=Returnedsupplier.objects.filter(product=product)
    qtyavoirsupp=avoirsupp.aggregate(Sum('qty'))['qty__sum'] or 0
    totalout=outbl+outfacture+qtyavoirsupp
    totalrev=round(revbl+revfacture, 2)
    print('>>>', revbl, revfacture, Outfacture.objects.filter(product=product))
    print(totalrev)
    stockininventaire=Stockin.objects.filter(product=product, isinventaire=True)
    stockout=Livraisonitem.objects.filter(product=product, isfacture=False).order_by('-id')
    stockoutfc=Outfacture.objects.filter(product=product, facture__bon__isnull=False).order_by('-id')
    avoirs=Returned.objects.filter(product=product)
    qtyin=stockin.aggregate(Sum('quantity'))['quantity__sum'] or 0
    qtyavoir=avoirs.aggregate(Sum('qty'))['qty__sum'] or 0
    print('>>> qtyavoir', qtyavoir)
    releve = chain(*[
    ((outbl, 'outbl') for outbl in stockout),
    ((outfc, 'outfc') for outfc in stockoutfc),
    ])
    thisproductreliquat=Orderitem.objects.filter(order__note__icontains='Reliquat', product=product, islivraison= False)


    # Sort the items by date
    outs = sorted(releve, key=lambda item: item[0].date, reverse=True)
    ctx={
        'stockininventaire':stockininventaire,
        'thisproductreliquat':thisproductreliquat,
        'outs':outs,
        'title':'Detail de '+product.ref,
        'product':product,
        'cars':product.getcars(),
        'carlogos':Carlogos.objects.all(),
        'categories':Category.objects.all(),
        'marques':Mark.objects.all(),
        'suppliers':Supplier.objects.all(),
        'entries':stockin,
        'sorties':stockout,
        'totalqtyin':qtyin+product.stockinitial+qtyavoir,
        'totalcout':stockin.aggregate(Sum('total'))['total__sum'] or 0,
        'totalqtyout':totalout,
        'totalcoutout':totalrev,
        'avoirs':avoirs,
        'reps':Represent.objects.all(),
        'repswithprice':commercial_prices,
        'today':timezone.now().date(),
        'avoirsupp':avoirsupp
    }
    return render(request, 'viewoneproduct.html', ctx)

def updateproduct(request):
    ref=request.POST.get('ref').lower().strip().replace('§', '-').replace("'", '')
    productid=request.POST.get('productid')
    product=Produit.objects.filter(ref=ref).exclude(pk=productid).first()
    if product:
        return JsonResponse({
            'success':False,
            'error':'Ref exist deja'
            })
    image=request.FILES.get('image') or None
    new=request.POST.get('switch')
    near=True if  request.POST.get('nearswitch') == 'on' else False
    selected_reps = request.POST.getlist('updatereps')
    remise=request.POST.get('remise')
    sellprice=request.POST.get('sellprice')
    minstock=request.POST.get('updateminstock')
    equivalent=request.POST.get('equivalent')
    logos = request.POST.getlist('updatepdctlogo')
    netprice=round(float(sellprice)-(float(sellprice)*float(remise)/100), 2)
    product=Produit.objects.get(pk=productid)
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    files = {}
    if image:
        files['image'] = image    
    if serverip:
        print('>> serverip', serverip)
        data={
            #'image':product.image.url.replace('/media/', '') if product.image else '/media/default.png',
            'new':True if request.POST.get('switch')=='on' else False,
            'logos':logos,
            'productid':request.POST.get('productid'),
            'remise':request.POST.get('remise'),
            'sellprice':request.POST.get('sellprice'),
            'netprice':round(float(sellprice)-(float(sellprice)*float(remise)/100), 2),
            'equivalent':equivalent,
            'near':near,
            'minstock':minstock,
            'code':request.POST.get('updatecode'),
            'refeq1':request.POST.get('refeq1'),
            'refeq2':request.POST.get('refeq2'),
            'refeq3':request.POST.get('refeq3'),
            'refeq4':request.POST.get('refeq4'),
            'repprice':request.POST.get('updaterepprice') or 0,
            # 'coderef':request.POST.get('updatecoderef'),
            'name':request.POST.get('name'),
            'cars':request.POST.get('cars'),
            'ref':request.POST.get('ref').lower().strip(),
            'category_id':request.POST.get('category'),
            'mark_id':request.POST.get('marque'),
            'diametre':request.POST.get('diametre'),
            'stock':product.stocktotal,
            'uniqcode':product.uniqcode
        }
        print('>> data', data)
        # if image:
        #     data['image']=product.image.url.replace('/media/', '') if product.image else '/media/default.png',
        
        try:
            headers = {
                'X-SYNC-TOKEN': '8a7f5b2c9d3e4f1g0h6j'
            }
            res=req.post(f'http://{serverip}/products/updateproduct', data=data, files=files, headers=headers)
            res.raise_for_status()
        except req.exceptions.RequestException as e:
            print('>>> error', e)
            return JsonResponse({
                'success':False,
                'error':f'error {e}'
            })
    #if price changed itshould be changed in reliquat and panier of clients
    # if float(sellprice) != float(product.sellprice):
    #     print('price changed')
    #     reliquas=Wishlist.objects.filter(product=product)
    #     for i in reliquas:
    #         i.total=round(float(netprice)*float(i.qty), 2)
    #         i.save()
    #     cartitems=Cartitems.objects.filter(product=product)
    #     for i in cartitems:
    #         newtotal=round(float(netprice)*float(i.qty), 2)
    #         newcarttotal=i.cart.total-i.total+newtotal
    #         i.total=newtotal
    #         i.save()
    #         i.cart.total=newtotal
    #         i.cart.save()
    equivalent=request.POST.get('equivalent')
    logos = request.POST.getlist('updatepdctlogo')  # because it's multiple
    product.carlogos.set(logos)
    if new=='on':
      product.isnew=True
    else:
      product.isnew=False
    product.near=near
    product.minstock=minstock
    product.repsprice=json.dumps(selected_reps)
    product.equivalent=equivalent
    product.code=request.POST.get('updatecode')
    product.refeq1=request.POST.get('refeq1').strip().upper()
    product.refeq2=request.POST.get('refeq2').strip().upper()
    product.refeq3=request.POST.get('refeq3').strip().upper()
    product.refeq4=request.POST.get('refeq4').strip().upper()
    product.coderef=request.POST.get('updatecoderef')
    product.representprice=request.POST.get('updaterepprice') or 0
    product.representremise=request.POST.get('updaterepremise') or 0
    product.sellprice=sellprice
    product.remise=remise
    product.prixnet=netprice
    product.name=request.POST.get('name').replace('§', '-')
    product.cars=request.POST.get('cars')
    product.ref=ref
    product.category_id=request.POST.get('category')
    product.mark_id=request.POST.get('marque')
    product.diametre=request.POST.get('diametre')
    product.block=request.POST.get('block')
    if image:
        product.image=image
    product.save()
    # res=req.get(f'http://{serverip}/products/updateproduct', data)
    # print('>>>>>>', res)
    # if not res.status_code == 200:
    #         print('Error message:', res.text)
    
        # req.get(f'http://{serverip}/products/updateproduct', {
        #     'password':'gadwad123',
        #     'id':request.POST.get('productid'),
        #     'ref':request.POST.get('ref').lower().strip(),
        #     'stocktotal':product.stocktotal,
        #     'cars':json.dumps(request.POST.getlist('cars')),
        # })
    return JsonResponse({
        'success':True
    })

def alertstock(request):
    targets = Category.objects.filter(produit__stocktotal__lte=F('produit__minstock')).annotate(
    total_products=Count('produit')
    )
    return render(request, 'alertstock.html', {'title':'Alert Stock', 'categories':targets,
    'suppliers':Supplier.objects.all()})

def getlowbycategory(request):
    category=request.POST.get('category')
    products=Produit.objects.filter(category__id=category, stocktotal__lte=F('minstock'))
    return JsonResponse({
        'data':render(request, 'lowstockproducts.html', {'products':products}).content.decode('utf-8')
    })

def commandsupplier(request):
    productid=request.POST.get('productid')
    supplierid=request.POST.get('supplierid')
    qty=request.POST.get('qty')
    product=Produit.objects.get(pk=productid)
    product.qtycommand=qty
    product.suppliercommand_id=supplierid
    product.iscommanded=True
    product.save()
    return JsonResponse({
        'success':True
    })

def cacelcommand(request):
    productid=request.POST.get('productid')
    product=Produit.objects.get(pk=productid)
    product.qtycommand=0
    product.suppliercommand_id=None
    product.iscommanded=False
    product.save()
    return JsonResponse({
        'success':True
    })

def recevoir(request):
    return render(request, 'recevoir.html', {'title':"Bon d'achat", 'suppliers':Supplier.objects.all(), 'today':timezone.now().date()})

def bonlivraison(request):
    # get the last order_no
    # if there is no order_no then set it to this format 'ym0001'
    # else increment it by one

    # increment it
    year = timezone.now().strftime("%y")
    latest_receipt = Bonlivraison.objects.filter(
        bon_no__startswith=f'BL{year}'
    ).order_by("-bon_no").first()
    if latest_receipt:
        latest_receipt_no = int(latest_receipt.bon_no[-5:])
        receipt_no = f"BL{year}{latest_receipt_no + 1:05}"
    else:
        receipt_no = f"BL{year}00001"
    print('>>>>>>', receipt_no)
    return render(request, 'bonlivraison.html', {
        'title':'Bon de livraison',
        # 'clients':Client.objects.all(),
        # 'products':Produit.objects.all(),
        'commercials':Represent.objects.all(),
    })

def facture(request):
    # get the last order_no
    # if there is no order_no then set it to this format 'ym0001'
    # else increment it by one

    # increment it
    return render(request, 'facture.html', {
        'title':'Facture',
        # 'clients':Client.objects.all(),
        # 'products':Produit.objects.all(),
        'commercials':Represent.objects.all(),
        #'order_no':receipt_no
    })


def suppliercommanproducts(request):
    supplierid=request.POST.get('supplierid')
    products=Produit.objects.filter(suppliercommand_id=supplierid)
    return JsonResponse({
        'data':render(request, 'suppliercommandproducts.html', {'products':products}).content.decode('utf-8')
    })


def searchref(request):
    ref=request.POST.get('ref')
    products=Produit.objects.filter(ref__istartswith=ref)
    return JsonResponse({
        'data':render(request, 'productsbon.html', {'products':products}).content.decode('utf-8')
    })

def addsupply(request):
    supplierid=request.POST.get('supplierid')
    products=request.POST.get('products')
    datebon=datetime.strptime(request.POST.get('datebon'), '%Y-%m-%d')
    datefacture=datetime.strptime(request.POST.get('datefacture'), '%Y-%m-%d')
    nbon=request.POST.get('nbon')
    isfacture= True if request.POST.get('mode')=='facture' else False
    totalbon=request.POST.get('totalbon')
    tva=0
    if isfacture:
        tva=round(float(totalbon)-(float(totalbon)/1.2), 2)

    supplier=Supplier.objects.get(pk=supplierid)
    # print(f'''
    #     bon achat N°: {nbon}
    #     date facture: {datefacture}
    #     date reception: {datebon}
    #     fournisseur: {supplier.name}
    #     sold fournisseur: {float(supplier.rest)+float(totalbon)}
    #     isfacure: {isfacture}
    #     tva: {tva}
    # ''')
    # for i in json.loads(products):
    #     qty=int(i['qty'])
    #     devise=0 if i['devise']=='' else i['devise']
    #     product=Produit.objects.get(pk=i['productid'])
    #     remise=0 if i['remise']=='' else int(i['remise'])
    #     buyprice=0 if i['price']=='' else i['price']
    #     netprice=round(float(buyprice)-(float(buyprice)*float(remise)/100), 2)

    #     print(f'''
    #         produit: {product.ref}
    #         nouvau prix achat: {netprice}
    #         devise: {devise}
    #         'qty': {qty}
    #     ''')
    #     if isfacture:
    #         print('add stock facture: ', int(product.stockfacture)+int(i['qty']))
    #     print('addstocktotal ', int(product.stocktotal)+int(i['qty']))
    # adding bontotal to sold
    supplier.rest=float(supplier.rest)+float(totalbon)
    supplier.save()
    bon=Itemsbysupplier.objects.create(
        supplier_id=supplierid,
        total=totalbon,
        date=datebon,
        nbon=nbon,
        isfacture=isfacture,
        tva=tva,
        dateentree=datefacture
    )
    for i in json.loads(products):
        devise=0 if i['devise']=='' else i['devise']
        product=Produit.objects.get(pk=i['productid'])
        remise=0 if i['remise']=='' else float(i['remise'])
        buyprice=0 if i['price']=='' else i['price']
        netprice=round(float(buyprice)-(float(buyprice)*float(remise)/100), 2)
        if isfacture:
            product.stockfacture=int(product.stockfacture)+int(i['qty'])
        # calculater cout moyen
        if product.stocktotal>0:
            totalqtys=int(product.stocktotal)+int(i['qty'])
            actualtotal=round(float(product.buyprice)*float(product.stocktotal), 2)
            newtotal=round(float(i['qty'])*netprice)
            totalprices=round(actualtotal+newtotal, 2)
            coutmoyen=round(totalprices/totalqtys, 2)
            product.coutmoyen=coutmoyen
            product.save()
        product.buyprice=netprice
        product.stocktotal=int(product.stocktotal)+int(i['qty'])
        product.devise= devise

        #product.isnew=True
        Stockin.objects.create(
            date=datebon,
            product=product,
            devise=devise,
            quantity=i['qty'],
            price=i['price'],
            ref=i['ref'],
            name=i['name'],
            remise=remise,
            qtyofprice=i['qty'],
            total=i['total'],
            supplier_id=supplierid,
            nbon=bon,
            facture=isfacture
        )
    # # update cout moyen, it will be calculated by deviding total prices by total qty

        # totalprices=Stockin.objects.filter(product=product).aggregate(Sum('total'))['total__sum'] or 0
        # totalqty=Stockin.objects.filter(product=product).aggregate(Sum('quantity'))['quantity__sum'] or 0
        # product.coutmoyen=round(totalprices/totalqty, 2)
        product.qtycommande=0
        product.supplier=supplier
        product.save()
    return JsonResponse({
        'html': render(request, 'recevoir.html', {'title':'Recevoir Les produits', 'suppliers':Supplier.objects.all(), #'products':Produit.objects.all()
        }).content.decode('utf-8')
    })


def addbonlivraison(request):

    #current_time = datetime.now().strftime('%H:%M:%S')
    clientid=request.POST.get('clientid')
    repid=request.POST.get('repid')
    products=request.POST.get('products')
    totalbon=request.POST.get('totalbon')
    orderid=request.POST.get('orderid', None)
    # orderno
    transport=request.POST.get('transport')
    note=request.POST.get('note')
    paymenttype=request.POST.get('iscontre')
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(f'{datebon}', '%Y-%m-%d')
    client=Client.objects.get(pk=clientid)
    client.soldtotal=round(float(client.soldtotal)+float(totalbon), 2)
    client.soldbl=round(float(client.soldbl)+float(totalbon), 2)
    client.save()
    if orderid is not None:
        cmnd=Order.objects.get(pk=orderid)
        cmnd.isdelivered=True
        cmnd.save()
    # get the last bon no
    year = timezone.now().strftime("%y")
    latest_receipt = Bonlivraison.objects.filter(
        bon_no__startswith=f'BL{year}'
    ).order_by("-bon_no").first()
    if latest_receipt:
        latest_receipt_no = int(latest_receipt.bon_no[-5:])
        receipt_no = f"BL{year}{latest_receipt_no + 1:05}"
    else:
        receipt_no = f"BL{year}00001"
    order=Bonlivraison.objects.create(
        commande_id=orderid,
        client_id=clientid,
        salseman_id=repid,
        total=totalbon,
        date=datebon,
        modlvrsn=transport,
        bon_no=receipt_no,
        note=note,
        # iscontre=iscontre
        paymenttype=paymenttype
    )
    print('>>>>>>', len(json.loads(products))>0)
    if len(json.loads(products))>0:
        with transaction.atomic():
            for i in json.loads(products):
                product=Produit.objects.get(pk=i['productid'])
                product.stocktotal=int(product.stocktotal)-int(i['qty'])
                product.save()
                Livraisonitem.objects.create(
                    bon=order,
                    remise=i['remise'],
                    name=i['name'],
                    ref=i['ref'],
                    product=product,
                    qty=i['qty'],
                    price=i['price'],
                    total=i['total'],
                    client_id=clientid,
                    date=datebon
                )


    # increment it
    return JsonResponse({
        "success":True
    })

# add facture not generer
def addfacture(request):
    clientid=request.POST.get('clientid')
    repid=request.POST.get('repid')
    products=request.POST.get('products')
    totalbon=request.POST.get('totalbon')
    # orderno
    transport=request.POST.get('transport', '')
    note=request.POST.get('note', '')
    #orderno=request.POST.get('orderno')
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    client=Client.objects.get(pk=clientid)
    client.soldtotal=round(float(client.soldtotal)+float(totalbon), 2)
    client.soldfacture=round(float(client.soldfacture)+float(totalbon), 2)
    client.save()
    tva=round(float(totalbon)-(float(totalbon)/1.2), 2)
    year = timezone.now().strftime("%y")
    latest_receipt = Facture.objects.filter(facture_no__startswith=f'FC{year}').order_by('-facture_no').first()
    if latest_receipt:
        latest_receipt_no = int(latest_receipt.facture_no[-5:])
        receipt_no = f"FC{year}{latest_receipt_no + 1:05}"
    else:
        receipt_no = f"FC{year}00001"
    print('>>>>>>>',latest_receipt)
    facture=Facture.objects.create(
        facture_no=receipt_no,
        total=totalbon,
        tva=tva,
        date=datebon,
        client_id=clientid,
        salseman_id=repid,
        transport=transport,
        note=note,
        hascopy=True,
        copynumber=receipt_no.replace('FC', 'BL')
    )
    if len(json.loads(products))>0:
        with transaction.atomic():
            for i in json.loads(products):
                product=Produit.objects.get(pk=i['productid'])
                product.stockfacture=int(product.stockfacture)-int(i['qty'])
                product.save()
                Outfacture.objects.create(
                    facture=facture,
                    remise=i['remise'],
                    name=i['name'],
                    ref=i['ref'],
                    product=product,
                    qty=i['qty'],
                    price=i['price'],
                    total=i['total'],
                    client_id=clientid,
                    date=datebon,
                )

    # year = timezone.now().strftime("%y")
    # latest_receipt = Facture.objects.filter(
    #     facture_no__startswith=f'FC{year}'
    # ).order_by("-facture_no").first()
    # latest_receipt_no = int(latest_receipt.facture_no[-5:])
    # receipt_no = f"FC{year}{latest_receipt_no + 1:05}"

    # increment it
    return JsonResponse({
        'success':True
    })



def supplierinfo(request, id):
    supplier=Supplier.objects.get(pk=id)
    ctx={
        'title':f'Info fournisseur {supplier.name.upper}',
        'supplier':supplier,
        'totalavoirs':Avoirsupplier.objects.filter(supplier=supplier).aggregate(Sum('total'))['total__sum'] or 0,
        'totalpayments':PaymentSupplier.objects.filter(supplier=supplier).aggregate(Sum('amount'))['amount__sum'] or 0,
        'totaltr':Itemsbysupplier.objects.filter(supplier=supplier).aggregate(Sum('total'))['total__sum'] or 0,
        'bons':Itemsbysupplier.objects.filter(supplier=supplier),
        'payments':PaymentSupplier.objects.filter(supplier=supplier)
    }
    return render(request, 'supplierinfo.html', ctx)

def clientinfo(request, id):
    client=Client.objects.get(pk=id)
    ctx={
        'client':client,
        'totalavoirs':Avoirclient.objects.filter(client=client).aggregate(Sum('total'))['total__sum'] or 0,
        'totalpayments':PaymentClientbl.objects.filter(client=client).aggregate(Sum('amount'))['amount__sum'] or 0,
        'totaltr':Bonlivraison.objects.filter(client=client).aggregate(Sum('total'))['total__sum'] or 0,
        'bons':Bonlivraison.objects.filter(client=client, total__gt=0)[:30],
        'payments':PaymentClientbl.objects.filter(client=client)[:30],
        'avoirs':Avoirclient.objects.filter(client=client)[:30]
    }
    return render(request, 'clientinfo.html', ctx)

def addpaymentssupp(request):
    supplierid=request.POST.get('supplierid')
    pass


def dashboard(request):

    ctx={
        'title':'Dashboard',
        'orders':Order.objects.filter(date__date=date.today()).count(),
        'products':Produit.objects.all().count(),
        'productthismonth':Orderitem.objects.filter(order__date__month=date.today().month).order_by('-qty')[:20],
        'alerts':Produit.objects.filter(stocktotal__lte=F('minstock')).count(),
        'blnotpaid':Bonlivraison.objects.filter(ispaid=False).count(),
        'boncommand':Order.objects.filter(isdelivered=False).count(),
        'soldtotal':round(Client.objects.aggregate(Sum('soldtotal'))['soldtotal__sum'] or 0, 2),



    }
    return render(request, 'pdashboard.html', ctx)

def clientspage(request):
    try:
        lastcode = Client.objects.order_by('code').last()
        print('lastcode', lastcode.code)
        if lastcode:

            codecl = f"{int(lastcode.code) + 1:06}"
        else:
            codecl = f"000001"
    except:
        codecl="000001"
    facture=request.GET.get('facture')=='1'
    if facture:
        clients=Client.objects.all().order_by('-soldfacture')[:50]
    else:
        clients=Client.objects.all().order_by('-soldtotal')[:50]
    ctx={
        'facturesection':facture,
        'clients':clients,
        'title':'List des clients',
        'commerciaux':Represent.objects.all(),
        'lastcode':codecl,
        'soldtotal':round(Client.objects.aggregate(Sum('soldtotal'))['soldtotal__sum'] or 0, 2),
        'soldbl':round(Client.objects.aggregate(Sum('soldbl'))['soldbl__sum'] or 0, 2),
        'soldfacture':round(Client.objects.aggregate(Sum('soldfacture'))['soldfacture__sum'] or 0, 2),
    }
    return render(request, 'clients.html', ctx)



def checkusername(request):
    username=request.POST.get('username')
    if User.objects.filter(username=username).exists():
        return JsonResponse({
            'exist':True
        })
    else:
        return JsonResponse({
            'exist':False
        })

def commercialspage(request):
    ctx={
        'commercials':Represent.objects.all(),
        'title':'List des Commeciaux'
    }
    return render(request, 'commerciaux.html', ctx)

def addcommercial(request):
    repusername=request.POST.get('repusername')
    reppassword=request.POST.get('reppassword')
    repname=request.POST.get('repname')
    repphone=request.POST.get('repphone')
    repregion=request.POST.get('repregion')
    repinfo=request.POST.get('repinfo')
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    if serverip:
        try:
            request=req.get(f'http://{serverip}/products/addcommercial',{
                'repusername':repusername,
                'reppassword':reppassword,
                'repname':repname,
                'repphone':repphone,
                'repregion':repregion,
                'repinfo':repinfo
            })
            request.raise_for_status()
        except Exception as e:
            print('>>> error', e)
            return JsonResponse({
                'success':False,
                'message':f'ERROR CONEXION AU SERVEUR DISTANT: {e}'
            })
        #user=User.objects.create_user(username=repusername, password=reppassword)
        # Get or create the group
        #group, created = Group.objects.get_or_create(name="salsemen")

        # Add the user to the group
        #user.groups.add(group)

        # Save the user
        #user.save()
    Represent.objects.create(
        #user=user,
        name=repname,
        phone=repphone,
        region=repregion,
        info=repinfo
    )
        # old code 04/07/2024
        # ctx={
        #     'commercials':Represent.objects.all(),
        #     'title':'List des Commeciaux'
        # }
        # return JsonResponse({
        #     'html':render(request, 'commerciaux.html', ctx).content.decode('utf-8')
        # })
    return JsonResponse({
        'success':True
    })
    

def checkcodeclient(request):
    code=request.POST.get('code')
    name=request.POST.get('name')
    print(Client.objects.filter(Q(name=name) | Q(code=code)))
    if Client.objects.filter(Q(name=name) | Q(code=code)).exists():
        return JsonResponse({
            'exist':True
        })
    return JsonResponse({
        'exist':False
    })
#this to add clients that are divers
def addclientdivers(request):
    name=request.GET.get('name')
    code=request.GET.get('code')
    city=request.GET.get('ville')
    client=Client.objects.create(
        city=city,
        code=code,
        name=name,
        soldtotal=0.00,
        soldfacture=0.00,
        soldbl=0.00,
        diver=True
    )
    return JsonResponse({
        'succes':True
    })



def addclient(request):
    name=request.POST.get('clientnameinp')
    phone=request.POST.get('clientphone')
    phone2=request.POST.get('clientphone2')
    address=request.POST.get('clientaddress')
    code=request.POST.get('clientcode')
    city=request.POST.get('clientcity')
    ice=request.POST.get('clientice')
    region=request.POST.get('clientregion').lower().strip()
    representant=request.POST.get('clientrep')
    if Client.objects.filter(Q(name=name) | Q(code=code)).exists():
        return JsonResponse({
            'success':False,
            'error':'Code ou Nom exist deja'
        })
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    if serverip:
        try:
            response=req.get(f'http://{serverip}/products/addclient', {
                'city':city,
                'ice':ice,
                'region':region,
                'represent_id':representant,
                'code':code,
                'name':name,
                'phone':phone,
                'address':address,
            })
            response.raise_for_status()
        except Exception as e:
            return JsonResponse({
                'success':False,
                'error':f'Error connexion au serveur distant: {e}'
            })
    client=Client.objects.create(
        city=city,
        ice=ice,
        region=region,
        represent_id=representant,
        code=code,
        name=name,
        phone=phone,
        phone2=phone2,
        address=address,
        soldtotal=0.00,
        soldfacture=0.00,
        soldbl=0.00,
        diver=False
    )
    return JsonResponse({
        'success':True
    })
    
def getclientdata(request):
    id=request.POST.get('id')
    client=Client.objects.get(pk=id)
    return JsonResponse({
        'personalname':client.clientname,
        'name':client.name,
        'phone':client.phone,
        'phone2':client.phone2,
        'address':client.address,
        'id':client.id,
        'code':client.code,
        'moderegl':client.moderegl,
        'city':client.city,
        'location':client.location,
        'region':client.region,
        'ice':client.ice,
        'rep':client.represent_id,
    })


def updateclient(request):
    id=request.POST.get('updateclientid')
    code=request.POST.get('updateclientcode')
    name=request.POST.get('updateclientname')


    client=Client.objects.get(pk=id)
    if Client.objects.filter(Q(name=name) | Q(code=code)).exclude(pk=id).exists():
         return JsonResponse({
             'success':False,
             'error':'Code ou Nom exist deja'
         })
    oldcode=client.code
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    if serverip:
        try:
            res=req.get(f'http://{serverip}/products/updateclient', {
                'clientcode':oldcode,
                'name':request.POST.get('updateclientname'),
                'phone':request.POST.get('updateclientphone'),
                'address':request.POST.get('updateclientaddress'),
                'ice':request.POST.get('updateclientice'),
                'code':request.POST.get('updateclientcode'),
                'city':request.POST.get('updateclientcity'),
                'address':request.POST.get('updateclientaddress'),
                'region':request.POST.get('updateclientregion'),
                'rep':request.POST.get('updateclientrep'),
            })
            res.raise_for_status

        
        except req.exceptions.RequestException as e:
            print('>>> ', e)
            return JsonResponse({
                'success':False
            })
    client.name=request.POST.get('updateclientname')
    client.phone=request.POST.get('updateclientphone')
    client.clientname=request.POST.get('updateclientpersonalname')
    client.phone2=request.POST.get('updateclientphone2')
    client.address=request.POST.get('updateclientaddress')
    client.ice=request.POST.get('updateclientice')
    client.code=request.POST.get('updateclientcode')
    client.city=request.POST.get('updateclientcity')
    client.location=request.POST.get('updateclientlocation')
    client.address=request.POST.get('updateclientaddress')
    client.region=request.POST.get('updateclientregion').lower().strip()
    client.moderegl=request.POST.get('updateclientmoderegl').strip()
    client.represent_id=request.POST.get('updateclientrep')
    client.save()
    return JsonResponse({
        'success':True
    })

def getscommercialdata(request):
    id=request.POST.get('id')
    rep=Represent.objects.get(pk=id)
    return JsonResponse({
        'name':rep.name,
        'phone':rep.phone,
        'phone2':rep.phone2,
        'region':rep.region,
        'info':rep.info,
        'id':rep.id
    })

def updatecommercial(request):
    id=request.POST.get('updaterepid')
    name=request.POST.get('updaterepname')
    phone=request.POST.get('updaterepphone')
    phone2=request.POST.get('updaterepphone2')
    region=request.POST.get('updaterepregion')
    region=request.POST.get('updaterepregion')
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    if serverip:
        try:
            res=req.get(f'http://{serverip}/products/updatecommercial', {
                'id':id,
                'name':name,
                'phone':phone,
                'phone2':phone2,
                'region':region
            })
            res.raise_for_status()
        except req.exceptions.RequestException as e:
            print('>>> ', e)
            return JsonResponse({
                'success':False,
                'error':f'Error connexion au serveur distant: {e}'
            })
    rep=Represent.objects.get(pk=id)
    rep.name=name
    rep.phone=phone
    rep.phone2=phone2
    rep.region=region
    rep.save()
    ctx={
        'commercials':Represent.objects.all(),
        'title':'List des Commeciaux'
    }
    return JsonResponse({
        'html':render(request, 'commerciaux.html', ctx).content.decode('utf-8')
    })


def onereppage(request, id):
    rep=Represent.objects.get(pk=id)
    note=Notesrepresentant.objects.filter(represent=rep).first()
    print('>>>>>>>>', note)
    ctx={
        'title':f'Page de {rep.name}',
        'rep':rep,
        'notes':note.note if note else '',
        'today':timezone.now().date()
    }
    return render(request, 'onereppage.html', ctx)

def adminpage(request):
    return render(request, 'adminpage.html')

def bonlivraisondetails(request, id):
    order=Bonlivraison.objects.get(pk=id)
    orderitems=Livraisonitem.objects.filter(bon=order, isfacture=False).order_by('product__name')
    print('orderitems', orderitems)
    reglements=PaymentClientbl.objects.filter(bons__in=[order])
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+34] for i in range(0, len(orderitems), 34)]
    ctx={
        'title':f'Bon de livraison {order.bon_no}',
        'order':order,
        'bon_no':order.bon_no.replace('BL', ''),
        'orderitems':orderitems,
        'reglements':reglements,
        'reps':Represent.objects.all()
    }
    return render(request, 'bonlivraisondetails.html', ctx)


def facturedetails(request, id):
    order=Facture.objects.get(pk=id)
    orderitems=Outfacture.objects.filter(facture=order).order_by('product__name')
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+30] for i in range(0, len(orderitems), 30)]
    reglements=order.reglementsfc.all()
    print('>>>>> reglements', reglements)
    ctx={
        'title':f'Facture {order.facture_no}',
        'facture':order,
        'orderitems':orderitems,
        'tva':order.tva,
        'ttc':order.total,
        'ht':round(order.total-order.tva, 2),
        'reps':Represent.objects.all(),
        'reglements':reglements
    }
    return render(request, 'facturedetails.html', ctx)

def facturedetailscopy(request, id):
    order=Facture.objects.get(pk=id)
    orderitems=Outfacture.objects.filter(facture=order).order_by('product__name')
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+30] for i in range(0, len(orderitems), 30)]
    reglements=order.reglementsfc.all()
    print('>>>>> reglements', reglements)
    ctx={
        'title':f'Facture {order.facture_no}',
        'facture':order,
        'orderitems':orderitems,
        'tva':order.tva,
        'ttc':order.total,
        'ht':round(order.total-order.tva, 2),
        'reps':Represent.objects.all(),
        'reglements':reglements
    }
    return render(request, 'facturedetailscopy.html', ctx)


def avoirdetails(request, id):
    order=Avoirclient.objects.get(pk=id)
    orderitems=Returned.objects.filter(avoir=order)
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+36] for i in range(0, len(orderitems), 36)]
    ht=round(order.total/1.2, 2)
    tva=order.total-ht
    ctx={
        'title':f'avoir {order.no}',
        'avoir':order,
        'orderitems':orderitems,
        'ht':ht,
        'tva':tva,
    }
    return render(request, 'avoirdetails.html', ctx)

def avoirsuppdetails(request, id):
    order=Avoirsupplier.objects.get(pk=id)
    orderitems=Returnedsupplier.objects.filter(avoir=order)
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+36] for i in range(0, len(orderitems), 36)]
    ctx={
        'title':f'avoir {order.no}',
        'avoir':order,
        'orderitems':orderitems,
    }
    return render(request, 'avoirsuppdetails.html', ctx)



def getrepswithprice(request):
    id=request.POST.get('id')
    product=Produit.objects.get(pk=id)
    repsprices=[]
    if product.repsprice:
        repsprices=json.loads(product.repsprice)
        print('get reps with price',json.loads(product.repsprice))
    return JsonResponse({
        'repswithprice': repsprices,
        'representremise':product.representremise
    })


def getclientprice(request):
    pdctid=request.POST.get('id')
    clientid=request.POST.get('clientid')
    target=request.POST.get('target')
    print('>> target', target, clientid, pdctid)
    print('>> target', target, clientid, pdctid)
    price=0
    remise=0
    #where we got that pdct, which bpn/facture
    source=''
    # try:
    #         clientprice=Livraisonitem.objects.filter(bon__client_id=clientid, product_id=pdctid).last()
    #         price=clientprice.price
    #         remise=clientprice.remise
    #         return JsonResponse({
    #             'price':price,
    #             'remise':remise
    #         })
    # except:
    #     return JsonResponse({
    #         'price':0
    #     })
    if target=='bl':
        print('target>> bl')
        print('target>> bl', pdctid, clientid)
        clientprice=Livraisonitem.objects.filter(bon__client_id=clientid, product_id=pdctid).last()
        print('clientprice', clientprice)
        if clientprice:
            print('>> getting client price')
            price=clientprice.price
            remise=clientprice.remise

            return JsonResponse({
            'price':price,
            'remise':remise,
            'source':source
            })
    else:
        print('target>> fc')

        clientprice=Outfacture.objects.filter(facture__client_id=clientid, product_id=pdctid).last()
        print('clientprice', clientprice)

        if clientprice:
            print('>> getting client price in fc')
            print('>>>', clientprice.facture.facture_no)
            price=clientprice.price
            remise=clientprice.remise
            source=clientprice.facture.facture_no
            return JsonResponse({
            'price':price,
            'remise':remise,
            'source':source
            })
    return JsonResponse({
    'price':price,
    'remise':remise,
    'source':source
    })

def listbonlivraison(request):
    facture=request.GET.get('facture')=='1'
    today = timezone.now().date()
    thisyear=timezone.now().year
    current_time = datetime.now().strftime('%H:%M:%S')
    three_months_ago = timezone.now() - timedelta(days=90)  # Assuming 30 days per month on average

    # Query for Bonlivraison objects that have a 'date' field earlier than three months ago
    depasser = Bonlivraison.objects.filter(date__lt=three_months_ago, ispaid=False, total__gt=0).count()
    # get only the last 100 orders of the current year
    bons= Bonlivraison.objects.filter(date__year=timezone.now().year).order_by('-bon_no')[:50]
    if facture:
        specific_date = datetime(2024, 11, 7).date()
        bons = Bonlivraison.objects.filter(
            Q(date__date=specific_date) | Q(date__date__lt=specific_date, isfacture=True) & ~Q(total__gt=0),
            date__year=timezone.now().year
        ).order_by('-bon_no')[:50]

    total=Bonlivraison.objects.filter(date__year=timezone.now().year).aggregate(Sum('total')).get('total__sum')
    ctx={
        'title':'Bons de livraison',
        'bons':bons,
        'total':total,
        'boncommand':Order.objects.filter(isdelivered=False).exclude(note__icontains='Reliquat').count(),
        'depasser':depasser,
        'reps':Represent.objects.all(),
        'today':timezone.now().date(),
        'facturesection':facture
    }
    return render(request, 'listbonlivraison.html', ctx)

def exportbl(request):
    rep=request.GET.get('rep')
    datefrom=request.GET.get('startdate')
    dateto=request.GET.get('enddate')
    region=request.GET.get('region').lower().strip()
    print('>>>>>>', rep, datefrom, dateto)
    if rep and region:
        print('rep and region')
        bons=Bonlivraison.objects.filter(salseman_id=rep,client__region=region, date__range=[datefrom, dateto])
    if rep and not region:
        print('rep and not region')
        bons=bons=Bonlivraison.objects.filter(salseman_id=rep, date__range=[datefrom, dateto])
    if not rep and region:
        print('not rep and region')
        bons=bons=Bonlivraison.objects.filter(client__region=region, date__range=[datefrom, dateto])
    if not region and not rep:
        print('nothing')
        bons=bons=Bonlivraison.objects.filter(date__range=[datefrom, dateto])

    # if rep and datefrom and dateto:
    #     print('date and rep')
    #     # convert date to datetime
    #     datefrom=datetime.strptime(datefrom, '%Y-%m-%d')
    #     dateto=datetime.strptime(dateto, '%Y-%m-%d')
    #     bons=Bonlivraison.objects.filter(salseman_id=rep, date__range=[datefrom, dateto])
    # if rep and not datefrom and not dateto:
    #     print('only rep')
    #     bons=Bonlivraison.objects.filter(salseman_id=rep, date__year=timezone.now().year)
    # if not rep and datefrom and dateto:
    #     print('only date')
    #     datefrom=datetime.strptime(datefrom, '%Y-%m-%d')
    #     dateto=datetime.strptime(dateto, '%Y-%m-%d')
    #     bons=Bonlivraison.objects.filter(date__range=[datefrom, dateto])
    # if not rep and not datefrom and not dateto:
    #     print('nothing')
    #     bons=Bonlivraison.objects.filter(date__year=timezone.now().year)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


    # Create a new Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write column headers
    ws.append(['N° Bon', 'Date', 'Client', 'Code cl.', 'total', 'region', 'ville', 'soldbl', 'rep.', 'status', 'facture', 'transport'])

    # Write product data
    for product in bons:
        ws.append([
            product.bon_no, product.date.strftime("%d/%m/%Y"), product.client.name, product.client.code, product.total, product.client.region, product.client.city, product.client.soldbl, product.salseman.name, 'R0' if product.ispaid else 'N1', 'OUI' if product.isfacture else 'NON', product.modlvrsn])

    response['Content-Disposition'] = f'attachment; filename="bonlivraison.xlsx"'
    # Save the workbook to the response
    wb.save(response)
    return response

def exceljvc(request):
    year=request.GET.get('year')
    items=Outfacture.objects.filter(date__year=year).order_by('date')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


    # Create a new Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write column headers
    ws.append(['N° facture', 'Date', 'Ref', 'Designation', 'Client', 'Code cl.', 'quantité', 'prix', 'total',])

    # Write product data
    for product in items:
        ws.append([
            product.facture.facture_no, product.date.strftime("%d/%m/%Y"), product.ref, product.product.name, product.client.name, product.client.code, product.qty, product.price, product.total])

    response['Content-Disposition'] = f'attachment; filename="factureproduit"'+year+'".xlsx"'
    # Save the workbook to the response
    wb.save(response)
    return response


def exportfc(request):
    rep=request.GET.get('rep')
    region=request.GET.get('region')
    datefrom=request.GET.get('startdate')
    dateto=request.GET.get('enddate')
    print('>>>>>>', rep, datefrom, dateto)
    if rep and region:
        print('rep and region')
        bons=Facture.objects.filter(salseman_id=rep,client__region=region, date__range=[datefrom, dateto]).order_by('-date')
    if rep and not region:
        print('rep and not region')
        bons=Facture.objects.filter(salseman_id=rep, date__range=[datefrom, dateto]).order_by('-date')
    if not rep and region:
        print('not rep and region')
        bons=Facture.objects.filter(client__region=region, date__range=[datefrom, dateto]).order_by('-date')
    if not region and not rep:
        print('nothing')
        bons=Facture.objects.filter(date__range=[datefrom, dateto]).order_by('-date')

    # if rep and datefrom and dateto:
    #     print('date and rep')
    #     # convert date to datetime
    #     datefrom=datetime.strptime(datefrom, '%Y-%m-%d')
    #     dateto=datetime.strptime(dateto, '%Y-%m-%d')
    #     bons=Facture.objects.filter(salseman_id=rep, date__range=[datefrom, dateto])
    # if rep and not datefrom and not dateto:
    #     print('only rep')
    #     bons=Facture.objects.filter(salseman_id=rep, date__year=timezone.now().year)
    # if not rep and datefrom and dateto:
    #     print('only date')
    #     datefrom=datetime.strptime(datefrom, '%Y-%m-%d')
    #     dateto=datetime.strptime(dateto, '%Y-%m-%d')
    #     bons=Facture.objects.filter(date__range=[datefrom, dateto])
    # if not rep and not datefrom and not dateto:
    #     print('nothing')
    #     bons=Facture.objects.filter(date__year=timezone.now().year)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


    # Create a new Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write column headers
    ws.append(['N° facture', 'Date', 'Client', 'Code cl.', 'total', 'region', 'ville', 'soldfc', 'rep.', 'status', ])

    # Write product data
    for bon in bons:
        ws.append([
            bon.facture_no, bon.date.strftime("%d/%m/%Y"), bon.client.name if bon.client else '--', bon.client.code, bon.total, bon.client.region, bon.client.city, bon.client.soldfacture, bon.salseman.name if bon.salseman else '--', 'R0' if bon.ispaid else 'N1'])

    response['Content-Disposition'] = f'attachment; filename="facture.xlsx"'
    # Save the workbook to the response
    wb.save(response)
    return response





def listavoirclient(request):
    thisyear=timezone.now().year
    facture=request.GET.get('facture')=='1'
    if facture:
        bons= Avoirclient.objects.filter(date__year=thisyear, avoirfacture=True).order_by('-date')
    else:
        bons= Avoirclient.objects.filter(date__year=thisyear).order_by('-date')
    total=bons.aggregate(Sum('total')).get('total__sum')
    ctx={
        'title':'Avoir Client',
        'bons':bons,
        'total':total,
        'facturesection':facture
    }
    return render(request, 'listavoirclient.html', ctx)

def yeardataavcl(request):
    year=request.GET.get('year')

    bons= Avoirclient.objects.filter(date__year=year).order_by('-date')
    total=bons.aggregate(Sum('total')).get('total__sum')
    ctx={
        'title':'Avoir Client',
        'bons':bons,
        'total':total,
    }
    return render(request, 'avoircltrs.html', ctx)


def listavoirsupplier(request):
    thisyear=timezone.now().year
    print('>>>>>>',)
    facture=request.GET.get('facture')
    if facture:
        bons= Avoirsupplier.objects.filter(date__year=thisyear, avoirfacture=True).order_by('-date')
    else:
        bons= Avoirsupplier.objects.filter(date__year=thisyear).order_by('-date')
    total=bons.aggregate(Sum('total')).get('total__sum')
    ctx={
        'title':'AVOIR FOURNISSEUR',#last edit
        'bons':bons,
        'total':total,
        'facturesection':facture
    }
    return render(request, 'listavoirsupplier.html', ctx)

def yeardataavsupp(request):
    facture=request.GET.get('facture')=='1'
    print('>> facture', facture)
    year=request.GET.get('year')
    if facture:
        bons= Avoirsupplier.objects.filter(date__year=year, avoirfacture=True).order_by('-date')
    else:
        bons= Avoirsupplier.objects.filter(date__year=year).order_by('-date')
    return JsonResponse({
        'trs':render(request, 'avsupptrs.html', {'bons':bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total'))['total__sum'] or 0, 2)
    })
def listfactures(request):
    year=request.GET.get('year', timezone.now().year)
    print('>>, ', year)
    facture=request.GET.get('facture')=='1'
    three_months_ago = timezone.now() - timedelta(days=90)
    depasser = Facture.objects.filter(date__lt=three_months_ago, ispaid=False).count()
    # get only the last 100 orders of the current year
    if facture:
        bons= Facture.objects.filter(date__year=timezone.now().year).exclude(client_id=3731).order_by('-facture_no')[:50]
    else:
        bons= Facture.objects.filter(date__year=year).order_by('-facture_no')[:50]
    ctx={
        'title':'List des factures',
        'bons':bons,
        'reps':Represent.objects.all(),
        'depasserfc':depasser,
        'today':timezone.now().date(),
        'facturesection':facture
    }
    if bons:
        ctx['total']=round(Facture.objects.filter(date__year=timezone.now().year).aggregate(Sum('total'))['total__sum'] or 0, 2)
        ctx['totaltva']=round(Facture.objects.filter(date__year=timezone.now().year).aggregate(Sum('tva'))['tva__sum'] or 0, 2)
    return render(request, 'listfactures.html', ctx)

def listfacturescopy(request):
    year=request.GET.get('year', timezone.now().year)
    print('>>, ', year)
    facture=request.GET.get('facture')=='1'
    three_months_ago = timezone.now() - timedelta(days=90)
    depasser = Facture.objects.filter(date__lt=three_months_ago, ispaid=False).count()
    # get only the last 100 orders of the current year
    if facture:
        bons= Facture.objects.filter(hascopy=True, date__year=timezone.now().year).exclude(client_id=3731).order_by('-facture_no')[:50]
    else:
        bons= Facture.objects.filter(hascopy=True, date__year=year).order_by('-facture_no')[:50]
    ctx={
        'title':'List bl',
        'bons':bons,
        'reps':Represent.objects.all(),
        'depasserfc':depasser,
        'today':timezone.now().date(),
        'facturesection':facture
    }
    if bons:
        ctx['total']=round(Facture.objects.filter(hascopy=True, date__year=timezone.now().year).aggregate(Sum('total'))['total__sum'] or 0, 2)
        ctx['totaltva']=round(Facture.objects.filter(hascopy=True, date__year=timezone.now().year).aggregate(Sum('tva'))['tva__sum'] or 0, 2)
    return render(request, 'listcopyfactures.html', ctx)


def activerproduct(request):
    id=request.POST.get('id')
    product=Produit.objects.get(pk=id)
    uniqcode=product.uniqcode
    serverip= Setting.objects.only('serverip').first()
    serverip=serverip.serverip if serverip else None
    if serverip:
        res=req.get(f'http://{serverip}/products/activerproduct', {
            'uniqcode':uniqcode
        })
        if json.loads(res.text)['success']:
            product.isactive=True
            product.save()
            return JsonResponse({
                'success':True
            })
        else:
            return JsonResponse({
                'success':False,
                'error':f"{json.loads(res.text)['error']}"
            })
    # user=User.objects.filter(username=username).first()
    # if user:
    product.isactive=True
    product.save()
    return JsonResponse({
        'success':False,
        'error':'No server'
    })

            
    product.isactive=True
    product.save()
    ctx={
        'title':'Detail de '+product.ref,
        'product':product,
        'cars':product.getcars(),
        'categories':Category.objects.all(),
        'marques':Mark.objects.all(),
        'suppliers':Supplier.objects.all(),
        'entries':Stockin.objects.filter(product=product),
        'sorties':Orderitem.objects.filter(product=product),
    }
    
    return JsonResponse({
        'html':render(request, 'viewoneproduct.html', ctx).content.decode('utf-8')
    })

def desactiverproduct(request):
    id=request.POST.get('id')
    product=Produit.objects.get(pk=id)
    uniqcode=product.uniqcode
    serverip= Setting.objects.only('serverip').first()
    serverip=serverip.serverip if serverip else None
    if serverip:
        res=req.get(f'http://{serverip}/products/desactiverproduct', {
            'uniqcode':uniqcode
        })
        if json.loads(res.text)['success']:
            product.isactive=False
            product.save()
            return JsonResponse({
                'success':True
            })
        else:
            return JsonResponse({
                'success':False,
                'error':f"{json.loads(res.text)['error']}"
            })
    product.isactive=False
    product.save()
    return JsonResponse({
        'success':False,
        'error':'No server'
    })
    # product.isactive=False
    # product.save()
    # ctx={
    #     'title':'Detail de '+product.ref,
    #     'product':product,
    #     'cars':product.getcars(),
    #     'categories':Category.objects.all(),
    #     'marques':Mark.objects.all(),
    #     'suppliers':Supplier.objects.all(),
    #     'entries':Stockin.objects.filter(product=product),
    #     'sorties':Orderitem.objects.filter(product=product),
    # }
    # req.get(f'http://{serverip}/products/desactiverproduct', {
    #     'id':request.POST.get('id')
    # })
    # return JsonResponse({
    #     'html':render(request, 'viewoneproduct.html', ctx).content.decode('utf-8')
    # })

def generatefacture(request, id):
    livraison=Bonlivraison.objects.get(pk=id)
    items=Livraisonitem.objects.filter(bon=livraison)
    lastdate=Facture.objects.last()
    if lastdate:
        lastdate=Facture.objects.last().date
    year = timezone.now().strftime("%y")
    latest_receipt = Facture.objects.filter(
        facture_no__startswith=f'FC{year}'
    ).order_by("-facture_no").first()
    if latest_receipt:
        latest_receipt_no = int(latest_receipt.facture_no[-5:])
        receipt_no = f"FC{year}{latest_receipt_no + 1:05}"
    else:
        receipt_no = f"FC{year}00001"
    ctx={
        'today':timezone.now().date(),
        'livraison':livraison,
        'items':items,
        'receipt_no':receipt_no,
        'lastdate':lastdate
    }
    return render(request, 'generatefacture.html', ctx)

# genereate facture from bl
def createfacture(request):
    bon=request.POST.get('bon')
    total=request.POST.get('totalbon')
    datefacture=request.POST.get('datefacture')
    datefacture=datetime.strptime(datefacture, '%Y-%m-%d')
    orderno=request.POST.get('orderno')
    products=json.loads(request.POST.get('products'))
    livraison=Bonlivraison.objects.get(pk=bon)
    livraison.isfacture=True
    livraison.statusfc='f1'

    # watch out for negative total
    livraison.total=round(livraison.total-float(total), 2)
    livraison.save()
    thisclient=Client.objects.get(pk=livraison.client_id)
    # we substract sold bl because we generate from bin livraisso
    thisclient.soldbl=round(thisclient.soldbl-float(total), 2)
    thisclient.soldfacture=round(livraison.client.soldfacture+float(total), 2)
    thisclient.save()
    # list of ids in oredritems
    tva=round(float(total)-(float(total)/1.2), 2)
    facture=Facture.objects.create(
        bon_id=bon,
        facture_no=orderno,
        total=total,
        tva=tva,
        date=datefacture,
        client=livraison.client,
        salseman=livraison.salseman,
        printed=False,
        hascopy=True,
        iscontre=livraison.iscontre,
        copynumber=orderno.replace('FC', 'BL')
    )


    product_ids_to_remove = [i['productid'] for i in products]

    # Delete the matching Orderitem objects in a single transaction
    Livraisonitem.objects.filter(bon=livraison, product_id__in=product_ids_to_remove).update(isfacture=True)
    with transaction.atomic():

        for i in products:
            product=Produit.objects.get(pk=i['productid'])
            product.stockfacture=int(product.stockfacture)-int(i['qty'])
            product.save()
            Outfacture.objects.create(
                facture=facture,
                product=product,
                qty=i['qty'],
                price=i['price'],
                total=i['total'],
                remise=i['remise'],
                ref=i['ref'],
                name=i['name'],
                client=livraison.client,
                date=datefacture
            )
    # if livraison.ispaid:
    #     reglement=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison, amount__gte=float(total)).first()
    #     if reglement is not None:
    #         # give reglement bl to regl fact
    #         reglement.amount=round(float(reglement.amount)-float(total), 2)
    #         reglement.usedinfacture=True
    #         reglement.save()

    #         regfac=PaymentClientfc.objects.create(
    #             client=livraison.client,
    #             amount=float(total),
    #             mode=reglement.mode,
    #             factures=facture,
    #             amountofeachbon=reglement.echance,
    #             npiece=reglement.npiece
    #         )
    #     else:
    #         reglements=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison)
    #         reglfacture=0
    #         for i in reglements:
    #             if reglfacture==float(total):
    #                 break
    #             if i.amount<=round(float(total)-reglfacture, 2):
    #                 regfac=PaymentClientfc.objects.create(
    #                     client=livraison.client,
    #                     amount=i.amount,
    #                     mode=i.mode,
    #                     factures=facture,
    #                     echance=i.echance,
    #                     npiece=i.npiece
    #                 )
    #                 reglfacture+=i.amount
    #                 i.amount=0
    #                 i.usedinfacture=True
    #                 i.save()
    #             else:
    #                 wanted=round(float(total)-reglfacture, 2)
    #                 regfac=PaymentClientfc.objects.create(
    #                     client=livraison.client,
    #                     amount=wanted,
    #                     mode=i.mode,
    #                     factures=facture,
    #                     echance=i.echance,
    #                     npiece=i.npiece
    #                 )
    #                 reglfacture+=wanted
    #                 i.amount=round(float(i.amount)-wanted)
    #                 i.usedinfacture=True
    #                 i.save()

    #     # Facturesregle.objects.create(
    #     #     payment=regfac,
    #     #     bon=facture,
    #     #     amount=float(total)

    #     # )
    #     # finish substraction form reglement bon
    #     facture.ispaid=True
    #     facture.save()
    # elif livraison.rest > 0:
    #     if float(total)==float(livraison.total):
    #         reglements=PaymentClientbl.objects.filter
    #     reglement=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison, amount__gte=float(total)).first()
    #     if reglement is not None:
    #         reglement.amount=round(float(reglement.amount)-float(total), 2)
    #         reglement.usedinfacture=True
    #         reglement.save()
    #         regfac=PaymentClientfc.objects.create(
    #             client=livraison.client,
    #             amount=float(total),
    #             mode=reglement.mode,
    #             factures=facture,
    #             echance=reglement.echance,
    #             npiece=reglement.npiece
    #         )
    #         facture.ispaid=True
    #         facture.save()
    #         # new bon rest
    #         #total reglements

    #     else:
    #         # if onre of the reglements is not gte total of facture
    #         # we need to iterate over regl and sum up the regl
    #         reglements=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison)
    #         reglfacture=0
    #         for i in reglements:
    #             if reglfacture==float(total):
    #                 break
    #             reglfacture+=i.amount
    #             i.amount=0
    #             i.usedinfacture=True
    #             i.save()
    #     totalreglements=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison).aggregate(Sum('amount'))['amount__sum']
    #     livraison.rest=round(livraison.total-totalreglements, 2)
    #     livraison.save()
    #     livraison.client.soldbl=round(livraison.client.soldbl-float(totalreglements), 2)
    # else:
    #     livraison.client.soldfacture=round(livraison.client.soldfacture+float(total), 2)
    #     livraison.client.save()
    return JsonResponse({
        'success':True
    })


def degenerer(request):
    bonid=request.POST.get('bonid')
    livraison=Bonlivraison.objects.get(pk=bonid)
    Livraisonitem.objects.filter(bon=livraison).update(isfacture=False)
    # delete facture and outfacture
    facture=Facture.objects.get(bon=livraison)
    outfactures=Outfacture.objects.filter(facture=facture)
    print(livraison, facture, outfactures)

    # products=Produit.objects.get(pk=i.product_id) for i in outfactures

    for i in outfactures:

        product=Produit.objects.get(pk=i.product_id)
        product.stockfacture=int(product.stockfacture)+int(i.qty)
        product.save()
        i.delete()
    livraison.isfacture=False
    livraison.statusfc='b1'
    livraison.total=round(livraison.total+float(facture.total), 2)
    livraison.save()
    livraison.client.soldbl=round(livraison.client.soldbl+float(facture.total), 2)
    livraison.client.soldfacture=round(livraison.client.soldfacture-float(facture.total), 2)
    livraison.client.save()
    facture.delete()
    return JsonResponse({
        'html':render(request, 'bonlivraisonbody.html', {'order':livraison}).content.decode('utf-8')
    })

def modifierlivraison(request, id):
    livraison=Bonlivraison.objects.get(pk=id)
    items=Livraisonitem.objects.filter(bon=livraison, isfacture=False)
    ctx={
        'title':'Modifier '+livraison.bon_no,
        'livraison':livraison,
        'items':items,
        'commercials':Represent.objects.all(),
        # 'products':Produit.objects.all(),
        # 'clients':Client.objects.all(),
    }
    return render(request, 'modifierlivraison.html', ctx)

def modifieravoir(request, id):
    avoir=Avoirclient.objects.get(pk=id)
    items=Returned.objects.filter(avoir=avoir)
    ctx={
        'avoir':avoir,
        'items':items,
        'commercials':Represent.objects.all(),
    }
    return render(request, 'modifieravoir.html', ctx)

def modifieravoirsupp(request, id):
    avoir=Avoirsupplier.objects.get(pk=id)
    items=Returnedsupplier.objects.filter(avoir=avoir)
    ctx={
        'avoir':avoir,
        'items':items,
    }
    return render(request, 'modifieravoirsupp.html', ctx)


def modifierfacture(request, id):
    facture=Facture.objects.get(pk=id)
    items=Outfacture.objects.filter(facture=facture)
    try:
        lastcode = Client.objects.order_by('code').last()
        print('lastcode', lastcode.code)
        if lastcode:

            codecl = f"{int(lastcode.code) + 1:06}"
        else:
            codecl = f"000001"
    except:
        codecl="000001"
    ctx={
        'lastcode':codecl,
        'facture':facture,
        'items':items,
        'products':Produit.objects.all(),
        'commercials':Represent.objects.all(),
        'clients':Client.objects.all(),
    }
    return render(request, 'modifierfacture.html', ctx)


def updatebonlivraison(request):
    id=request.POST.get('bonid')
    livraison=Bonlivraison.objects.get(pk=id)
    client=Client.objects.get(pk=request.POST.get('clientid'))
    totalbon=request.POST.get('totalbon')
    transport=request.POST.get('transport')
    note=request.POST.get('note')
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(f'{datebon}', '%Y-%m-%d')

    thisclient=livraison.client
    if livraison.client==client:
        print('same client', float(thisclient.soldtotal), float(livraison.total), float(totalbon))
        thisclient.soldtotal=round(float(thisclient.soldtotal)-float(livraison.total)+float(totalbon), 2)
        thisclient.soldbl=round(float(thisclient.soldbl)-float(livraison.total)+float(totalbon), 2)
        thisclient.save()
    else:
        print('not same')
        thisclient.soldtotal=round(float(thisclient.soldtotal)-float(livraison.total), 2)
        thisclient.soldbl=round(float(thisclient.soldbl)-float(livraison.total), 2)
        thisclient.save()
        print('old', thisclient.soldtotal)
        client.soldtotal+=float(totalbon)
        client.soldbl+=float(totalbon)
        client.save()
        print('new', client.soldtotal)
    livraison.modlvrsn=transport
    livraison.client=client
    livraison.note=note
    livraison.salseman_id=request.POST.get('repid')
    livraison.total=totalbon
    livraison.date=datebon
    livraison.bon_no=request.POST.get('orderno')
    livraison.save()
    items=Livraisonitem.objects.filter(bon=livraison)
    # update this items
    for i in items:
        product=Produit.objects.get(pk=i.product_id)
        product.stocktotal=int(product.stocktotal)+int(i.qty)
        product.save()
        i.delete()

    print('client:', livraison.client.id)
    for i in json.loads(request.POST.get('products')):
        # ABORTER FOR NOW
        # clientpricehistory=Clientprices.objects.filter(client_id=livraison.client.id, product_id=i['productid']) or None
        # if clientpricehistory:
        #     print('clientpricehistory exist')
        #     clientpricehistory.update(price=i['price'])
        # else:
        #     print('clientpricehistory not exist')
        #     Clientprices.objects.create(client_id=livraison.client.id, product_id=i['productid'], price=i['price'])


        qty=int(i['qty'])
        product=Produit.objects.get(pk=i['productid'])
        product.stocktotal=int(product.stocktotal)-qty

        product.save()

        # create new livraison items
        Livraisonitem.objects.create(
            bon=livraison,
            remise=i['remise'],
            name=i['name'],
            ref=i['ref'],
            product=product,
            qty=qty,
            price=i['price'],
            total=i['total'],
            date=datebon,
            client=client
        )

    return JsonResponse({
        'success':True
    })


def updatebonfacture(request):
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    client=Client.objects.get(pk=request.POST.get('clientid'))
    facture=Facture.objects.get(pk=request.POST.get('factureid'))
    totalbon=request.POST.get('totalbon')
    thisclient=facture.client
    if facture.client==client:
        print('same client', float(thisclient.soldtotal), float(facture.total), float(totalbon))
        thisclient.soldtotal=round(float(thisclient.soldtotal)-float(facture.total)+float(totalbon), 2)
        thisclient.soldfacture=round(float(thisclient.soldfacture)-float(facture.total)+float(totalbon), 2)
        thisclient.save()
    else:
        print('not same')
        thisclient.soldtotal=round(float(thisclient.soldtotal)-float(facture.total), 2)
        thisclient.soldfacture=round(float(thisclient.soldfacture)-float(facture.total), 2)
        thisclient.save()
        print('old', thisclient.soldtotal)
        client.soldtotal+=float(totalbon)
        client.soldfacture+=float(totalbon)
        client.save()
        print('new', client.soldtotal)
    tva=round(float(totalbon)-(float(totalbon)/1.2), 2)
    facture.tva=tva
    facture.client=client
    facture.salseman_id=request.POST.get('repid')
    facture.total=totalbon
    facture.facture_no=request.POST.get('orderno')
    facture.note=request.POST.get('note')
    facture.date=datebon
    facture.save()
    items=Outfacture.objects.filter(facture=facture)
    # update this items
    for i in items:
        product=Produit.objects.get(pk=i.product_id)
        product.stockfacture=int(product.stockfacture)+int(i.qty)
        product.save()
        i.delete()

    print('client:', facture.client.id)
    with transaction.atomic():
        for i in json.loads(request.POST.get('products')):

            # update price in facture
            # clientpricehistory=Clientprices.objects.filter(client_id=facture.client.id, product_id=i['productid']) or None
            # if clientpricehistory:
            #     print('clientpricehistory exist')
            #     clientpricehistory.update(price=i['price'])
            # else:
            #     print('clientpricehistory not exist')
            #     Clientprices.objects.create(client_id=facture.client.id, product_id=i['productid'], price=i['price'])


            qty=int(i['qty'])
            product=Produit.objects.get(pk=i['productid'])
            product.stockfacture=int(product.stockfacture)-qty

            product.save()

            # create new livraison items
            Outfacture.objects.create(
                facture=facture,
                remise=i['remise'],
                name=i['name'],
                ref=i['ref'],
                product=product,
                qty=qty,
                price=i['price'],
                total=i['total'],
                client=client,
                date=datebon,
            )

    return JsonResponse({
        'success':True
    })


def listreglementbl(request):
    thisyear=timezone.now().year
    reglements=PaymentClientbl.objects.filter(date__year=thisyear).order_by('-echance')[:50]
    print(date.today())
    ctx={
        'title':'List des reglements CL BL',
        'reglements':reglements,
        'today':timezone.now().date()
    }
    if reglements:
        ctx['total']=round(PaymentClientbl.objects.filter(date__year=thisyear).aggregate(Sum('amount'))['amount__sum'], 2)

    return render(request, 'listreglementbl.html', ctx)


def listreglementsupp(request):
    thisyear=timezone.now().year
    reglements=PaymentSupplier.objects.filter(date__year=thisyear).order_by('-id')
    # reglements=PaymentSupplier.objects.filter(date__year=thisyear).order_by('-id')[:50]
    ctx={
        'title':'List des reglements Fournisseur',
        'reglements':reglements,
        'suppliers':Supplier.objects.all(),
        'today':timezone.now().date()
    }
    if reglements:
        ctx['total']=round(PaymentSupplier.objects.filter(date__year=thisyear).aggregate(Sum('amount'))['amount__sum'], 2)
    return render(request, 'listreglementsupp.html', ctx)

def yeardatareglsupp(request):
    year=request.GET.get('year')
    print('>>>', year)
    reglements=PaymentSupplier.objects.filter(date__year=year).order_by('-id')
    ctx={
        'title':'List des reglements Fournisseur',
        'reglements':reglements,
        'suppliers':Supplier.objects.all(),
        'today':timezone.now().date()
    }
    if reglements:
        ctx['total']=round(PaymentSupplier.objects.filter(date__year=year).aggregate(Sum('amount'))['amount__sum'], 2)
    return render(request, 'reglsupptrs.html', ctx)

def laodblreglsupp(request):
    page = int(request.GET.get('page', 1))
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    regls=PaymentSupplier.objects.filter(date__year=thisyear).order_by('-id')[start:end]
    return JsonResponse({
        'trs':render(request, 'reglsupptrs.html', {'reglements':regls}).content.decode('utf-8'),
        'has_more': len(regls) == per_page
    })


def listreglementfc(request):
    thisyear=timezone.now().year
    # add :50 to this
    reglements=PaymentClientfc.objects.filter(date__year=thisyear).order_by('-echance')[:50]
    print(round(PaymentClientfc.objects.filter(date__year=thisyear).aggregate(Sum('amount'))['amount__sum'] or 0, 2))
    ctx={
        'title':'List des reglements CL fc',
        'reglements':reglements,
        'today':timezone.now().date(),


    }
    if reglements:
        ctx['total']=round(PaymentClientfc.objects.filter(date__year=thisyear).aggregate(Sum('amount'))['amount__sum'] or 0, 2)


    return render(request, 'listreglementfc.html', ctx)


def getclientbons(request):
    clientid=request.POST.get('clientid')
    print('>>>', clientid)
    bons=Bonlivraison.objects.filter(client_id=clientid).order_by('date')[:50]
    total=round(Bonlivraison.objects.filter(client_id=clientid).aggregate(Sum('total')).get('total__sum')or 0,  2)
    trs=''
    for i in bons:
        # old code, if reglement is paid it's checked from here
        # trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)" {"checked" if i.reglements.exists() else ""}></td></tr>'
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)"></td></tr>'


    return JsonResponse({
        'trs':trs,
        'total':total,
        'soldbl':round(Client.objects.get(pk=clientid).soldbl, 2)
    })

def filternonreglr(request):
    clientid=request.GET.get('clientid')
    bons=Bonlivraison.objects.filter(client_id=clientid, ispaid=False).order_by('date')[:50]
    total=round(Bonlivraison.objects.filter(client_id=clientid, ispaid=False).aggregate(Sum('total')).get('total__sum') or 0, 2)
    trs=''
    for i in bons:
        trs+=f'<tr class="blreglrow nr" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">NR</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)" {"checked" if i.reglements.exists() else ""}></td></tr>'

    return JsonResponse({
        'trs':trs,
        'total':total,
        'soldbl':Client.objects.get(pk=clientid).soldbl
    })

def filternonreglrfc(request):
    clientid=request.GET.get('clientid')
    print(clientid)
    bons=Facture.objects.filter(client_id=clientid, ispaid=False).order_by('date')[:50]
    total=round(Facture.objects.filter(client_id=clientid, ispaid=False).aggregate(Sum('total')).get('total__sum')or 0, 2)
    trs=''
    for i in bons:
        trs+=f'<tr class="fcreglrow nr" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">NR</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'

    return JsonResponse({
        'trs':trs,
        'total':total,
        'soldfc':Client.objects.get(pk=clientid).soldfacture
    })

def getclientfactures(request):
    clientid=request.POST.get('clientid')
    bons=Facture.objects.filter(client_id=clientid).order_by('date')[:50]
    trs=''
    for i in bons:
        trs+=f'<tr  class="fcreglrow" style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td>{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'trs':trs,
        'totalfactures':round(Facture.objects.filter(client_id=clientid).aggregate(Sum('total'))['total__sum']or 0, 2),
        'soldfactureregl':round(Client.objects.get(pk=clientid).soldfacture, 2)
    })

def reglefactures(request):
    clientid=request.POST.get('clientid')
    client=Client.objects.get(pk=clientid)
    print('>>> ', clientid)
    bons=json.loads(request.POST.get('bons'))
    mantant=json.loads(request.POST.get('mantant'))
    mode=json.loads(request.POST.get('mode'))
    npiece=json.loads(request.POST.get('npiece'))
    date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=json.loads(request.POST.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]
    livraisons=Facture.objects.filter(pk__in=bons)
    livraisons.update(ispaid=True)
    livraisons.update(statusreg='f1')
    # livraisonstotals=0
    totalmantant=sum(mantant)
    # for i in livraisons:
    #     if i.rest>0:
    #         livraisonstotals=round(livraisonstotals+i.rest, 2)
    #     else:
    #         livraisonstotals=round(livraisonstotals+i.total, 2)

    # print(totalmantant, livraisonstotals, livraisons)
    # # update client sold
    # # case1: 5000==5000:
    # if totalmantant==livraisonstotals:
    #     print('case1')
    #     livraisons.update(ispaid=True)
    #     livraisons.update(rest=0)
    # # case2: 5000>4500:
    # if totalmantant>livraisonstotals:
    #     print('case2')
    #     diff=totalmantant-livraisonstotals
    #     livraisons.update(ispaid=True)
    #     livraisons.update(rest=0)

    #     # else:
    #     #     livraisons.update(ispaid=True)
    #     # sub diff from client ref
    # # case3: 5000<6000:
    # if totalmantant<livraisonstotals:
        # print('case3')
        # amount=totalmantant
        # for i in livraisons:
        #     if amount<=0:
        #         break
        #     else:
        #         # if facture has rest
        #         if i.rest>0:
        #             if amount>=i.rest:
        #                 i.rest=0
        #                 i.ispaid=True
        #                 print(f'facture {i.facture_no} has rest is paid here')
        #                 i.save()
        #                 amount-=i.rest
        #             else:
        #                 print(f'facture {i.facture_no} has rest has rest here')
        #                 i.rest=round(i.rest-amount, 2)
        #                 i.save()
        #                 break
        #         else:
        #             if amount>=i.total:
        #                 amount-=i.total
        #                 print(f'facture {i.facture_no} is paid here')
        #                 i.ispaid=True
        #                 i.save()
        #             else:
        #                 print(f'facture {i.facture_no} hasrestof {round(i.total-amount, 2)}')
        #                 i.rest=round(i.total-amount, 2)
        #                 i.save()
        #                 break
        # print('rest of amount', amount)
    for m, mod, np, ech in zip(mantant, mode, npiece, echeance):
        tva=round(m-(m/1.2), 2)
        regl=PaymentClientfc.objects.create(
            client_id=clientid,
            amount=m,
            date=date,
            tva=tva,
            echance=ech,
            mode=mod,
            npiece=np.lower(),
        )
        regl.factures.set(livraisons)
        # storing factures in facturesregle
        # for i in livraisons:
        #     Facturesregle.objects.create(
        #         payment=regl,
        #         bon=i,
        #         amount=m
        #     )


    client.soldtotal=round(float(client.soldtotal)-float(totalmantant), 2)
    client.soldfacture=round(float(client.soldfacture)-float(totalmantant), 2)
    client.save()
    return JsonResponse({
        "success":True
    })


def reglebons(request):
    clientid=request.POST.get('clientid')
    client=Client.objects.get(pk=clientid)
    bons=json.loads(request.POST.get('bons'))
    mantant=json.loads(request.POST.get('mantant'))
    mode=json.loads(request.POST.get('mode'))
    npiece=json.loads(request.POST.get('npiece'))
    date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=json.loads(request.POST.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]
    livraisons=Bonlivraison.objects.filter(pk__in=bons)
    livraisons.update(ispaid=True)
    livraisons.update(statusreg='r0')
    totalmantant=sum(mantant)
    # for i in livraisons:
    #     if i.rest>0:
    #         livraisonstotals=round(livraisonstotals+i.rest, 2)
    #     else:
    #         livraisonstotals=round(livraisonstotals+i.total, 2)

    # print(totalmantant, livraisonstotals, mantant, mode, npiece, echeance)

    # # update client sold
    # # case1: 5000==5000:
    # amountofeachbon=[]
    # if totalmantant==livraisonstotals:
    #     print('case1')
    #     livraisons.update(ispaid=True)
    #     livraisons.update(rest=0)
    #     for i in livraisons:
    #         amountofeachbon.append(i.total)
    # # case2: 5000>4500:
    # if totalmantant>livraisonstotals:
    #     print('case2')
    #     diff=totalmantant-livraisonstotals
    #     livraisons.update(ispaid=True)
    #     livraisons.update(rest=0)
    #     for i in livraisons:
    #         amountofeachbon.append(i.total)
    #     # else:
    #     #     livraisons.update(ispaid=True)
    #     # sub diff from client ref
    # # case3: 5000<6000:
    # if totalmantant<livraisonstotals:
    #     print('case3')
    #     amount=totalmantant
    #     for i in livraisons:
    #         if amount<=0:
    #             break
    #         else:
    #             if i.rest>0:
    #                 if amount>=i.rest:
    #                     i.rest=0
    #                     i.ispaid=True
    #                     i.save()
    #                     amount-=i.rest
    #                     amountofeachbon.append(amount)
    #                 else:

    #                     i.rest=round(i.rest-amount, 2)
    #                     i.save()
    #                     amountofeachbon.append(i.rest)
    #                     break
    #             else:
    #                 if amount>=i.total:
    #                     amount-=i.total
    #                     i.ispaid=True
    #                     i.save()
    #                     amountofeachbon.append(amount)
    #                 else:
    #                     i.rest=round(i.total-amount, 2)
    #                     amountofeachbon.append(amount)
    #                     i.save()
    #                     break
    #     print('rest of amount', amount)


    for m, mod, np, ech in zip(mantant, mode, npiece, echeance):
        regl=PaymentClientbl.objects.create(
            client_id=clientid,
            amount=m,
            date=date,
            echance=ech,
            mode=mod,
            npiece=np.lower()
        )
        regl.bons.set(livraisons)
        # for i in livraisons:
        #     Bonsregle.objects.create(
        #         payment=regl,
        #         bon=i,
        #         amount=m
        #     )

    client.soldtotal=round(float(client.soldtotal)-float(totalmantant), 2)
    client.soldbl=round(float(client.soldbl)-float(totalmantant), 2)
    client.save()
    return JsonResponse({
        "success":True
    })



def checknpiece(request):
    npiece=request.GET.get('npiece')
    if PaymentClientbl.objects.filter(npiece=npiece).exists() or PaymentClientfc.objects.filter(npiece=npiece).exists():
        return JsonResponse({
            'exist':True
        })
    return JsonResponse({
        'exist':False
    })


def viewreglement(request, id):
    reglement=PaymentClientbl.objects.get(pk=id)
    reglements=Bonsregle.objects.filter(payment=reglement)
    ctx={
        'title':'Reglement',
        'reglement':reglement,
        'reglements':reglements
    }
    return render(request, 'viewreglement.html', ctx)

def viewreglementfc(request, id):
    reglement=PaymentClientfc.objects.get(pk=id)
    ctx={
        'title':'Reglement',
        'reglement':reglement,
        'reglfc':True
    }
    return render(request, 'viewreglement.html', ctx)



def situationcl(request):
    ctx={
        'title':'Situation des clients',
        'today':timezone.now().date()
    }
    return render(request, 'situationcl.html', ctx)

def situationclblfc(request):
    # when clicking on download in situationclbl/fc get the client id from function arguments
    clientname=request.GET.get('clientname')
    monthtostart=request.GET.get('monthtostart')
    monthtoend=request.GET.get('monthtoend')
    print('>>', monthtostart, monthtoend)
    clientid=request.GET.get('clientid')
    ctx={
        'monthtostart':monthtostart,
        'monthtoend':monthtoend,
        'title':'Situation des clients',
        'clientname':clientname,
        'clientid':clientid,
        'today':timezone.now().date()
    }
    return render(request, 'situationclblfc.html', ctx)

def situationsupplier(request):
    ctx={
        'title':'Situation des Fouurnisseurs',
        'suppliers':Supplier.objects.all()
    }
    return render(request, 'situationsupplier.html', ctx)

def situationclfc(request):
    ctx={
        'title':'Situation des clients (Factures)',
        'today':timezone.now().date()
    }
    return render(request, 'situationclfc.html', ctx)

def recevoirexcel(request):
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    df = df.fillna(0)
    print('>>>>>>>>>> exceel')
    trs=''
    totalbon=0
    refnotexist=[]
    for d in df.itertuples():
        try:
            ref = d.ref.lower().strip()
        except:
            ref=d.ref
        #name=d.name
        qty=d.qty
        prixachat=float(d.prixachat)
        # prixachat 12,0 make it 12,00
        remise=0 if pd.isna(d.remise) else int(d.remise)
        devise=0 if pd.isna(d.devise) else d.devise
        prixnet=round(prixachat-(prixachat*float(remise)/100), 2)

        # prixnet=round(prixachat-(prixachat*float(remise)/100), 2)
        # if remise==0:
        #     total=round(prixachat*qty, 2)
        # else:
        #     total=round(prixnet*qty, 2)
        # totalbon+=total
        # make total 2 digits after point
        if remise==0:
            total=round(prixachat*qty, 2)
        else:
            total=round(prixnet*qty, 2)
        product=Produit.objects.filter(ref=ref).first()
        if product:
            trs+=f"""<tr>
            <td class="ref">{ref}</td>
            <td class="name">{product.name}</td>
            <td>{product.buyprice if product.buyprice else 0}</td>
            <td>{product.stocktotal if product.stocktotal else 0}</td>
            <td>
            <input style="width:65px;" type="number" class="devise" value="{devise:.2f}">
            </td>
            <td><input style="width:65px;" type="number" class="calculatebonachat qty" onkeyup="calculate(event)" name="qtybonachat" value="{qty}"></td>
            <td><input style="width:65px;" type="number" class="calculatebonachat price" onkeyup="calculate(event)" name="pricebonachat" value="{prixachat:.2f}"></td>
            <td><input style="width:65px;" type="number" class="calculatebonachat remise" onkeyup="calculate(event)" name="remise" value="{remise}"></td>
            <td class="total">{total:.2f}</td>
            <input type="hidden" name="productid" value="{product.id}">
            </tr>"""

            totalbon+=total
        else:
            print('W>>>>not exist', ref)
            refnotexist.append(ref)
        # print('product exists')
        # trs+=f"""<tr>
        #     <td class="ref">{ref}</td>
        #     <td class="name">{product.name}</td>
        #     <td>{product.buyprice if product.buyprice else 0}</td>
        #     <td>{product.stocktotal if product.stocktotal else 0}</td>
        #     <td>
        #     <input style="width:65px;" type="number" class="devise" value="{devise:.2f}">
        #     </td>
        #     <td><input style="width:65px;" type="number" class="calculatebonachat qty" onkeyup="calculate(event)" name="qtybonachat" value="{qty}"></td>
        #     <td><input style="width:65px;" type="number" class="calculatebonachat price" onkeyup="calculate(event)" name="pricebonachat" value="{prixachat:.2f}"></td>
        #     <td><input style="width:65px;" type="number" class="calculatebonachat remise" onkeyup="calculate(event)" name="remise" value="{remise}"></td>
        #     <td class="total">{total:.2f}</td>
        #     <input type="hidden" name="productid" value="{product.id}">
        # </tr>"""
    return JsonResponse({
        'trs':trs,
        'totalbon':totalbon,
        'refnotexist':refnotexist
    })


def avoirclient(request):

    ctx={
            'title':'Avoir client',
            'commercials':Represent.objects.all(),
            #'receipt_no':receipt_no
        }
    return render(request, 'avoirclient.html', ctx)



def addavoirclient(request):
    clientid=request.POST.get('clientid')
    repid=request.POST.get('repid')
    products=request.POST.get('products')
    totalbon=request.POST.get('totalbon')
    mode=request.POST.get('mode')
    isfacture=True if mode=='facture' else False
    # orderno
    #orderno=request.POST.get('orderno')
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    client=Client.objects.get(pk=clientid)

    year = timezone.now().strftime("%y")
    if isfacture:
        prefix = f'AVF{year}'
    else:
        prefix = f'AV{year}'
    try:
        avoirclients = Avoirclient.objects.filter(no__startswith=prefix).last()
        latest_receipt_no = int(avoirclients.no.split('/')[1])
        receipt_no = f"{prefix}/{latest_receipt_no + 1}"
    except:
        receipt_no = prefix+"/1"
    print(receipt_no, clientid, repid, totalbon, datebon, isfacture)
    try:

        avoir=Avoirclient.objects.create(
            no=receipt_no,
            client_id=clientid,
            representant_id=repid,
            total=totalbon,
            date=datebon,
            avoirfacture=isfacture
        )

        for i in json.loads(products):
            product=Produit.objects.get(pk=i['productid'])
            if isfacture:
                product.stockfacture=int(product.stockfacture)+int(i['qty'])
            product.stocktotal=int(product.stocktotal)+int(i['qty'])
            product.save()
            Returned.objects.create(
                avoir=avoir,
                product=product,
                qty=i['qty'],
                source=i['source'],
                remise=0 if i['remise']=="" else i['remise'],
                price=0 if i['price']=="" else i['price'],
                total=i['total'],
            )
        client.soldtotal=round(float(client.soldtotal)-float(totalbon), 2)
        if isfacture:
            client.soldfacture=round(float(client.soldfacture)-float(totalbon), 2)
        else:
            client.soldbl=round(float(client.soldbl)-float(totalbon), 2)

        client.save()
    except Exception as e:
        print('>>error av cl:', e)

    # increment it
    return JsonResponse({
        'html':render(request, 'avoirclient.html', {
            'title':'Bon de livraison',
            #'clients':Client.objects.all(),
            #'products':Produit.objects.all(),
            'commercials':Represent.objects.all(),
            #'receipt_no':receipt_no
        }).content.decode('utf-8')
    })

def avoirsupplier(request):
    year = timezone.now().strftime("%y")
    prefix = f'FAV{year}'
    try:
        avoirsuppliers = Avoirsupplier.objects.filter(no__startswith=prefix).last()
        print('>>>>', avoirsuppliers.no)
        latest_receipt_no = int(avoirsuppliers.no.split('/')[1])
        receipt_no = f"FAV{year}/{latest_receipt_no + 1}"

    except:
        receipt_no = f"FAV{year}/1"
    print('>>>>>>rec av supp', receipt_no)
    ctx={
            'title':'Avoir Fournisseur',
        }
    return render(request, 'avoirsupplier.html', ctx)



def addavoirsupp(request):
    supplierid=request.POST.get('supplierid')
    products=request.POST.get('products')
    totalbon=request.POST.get('totalbon')
    # orderno
    avoirfacture=True if request.POST.get('mode')=='facture' else False
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    supplier=Supplier.objects.get(pk=supplierid)
    year = timezone.now().strftime("%y")

    prefix = f'FAV{year}'
    try:
        avoirsuppliers = Avoirsupplier.objects.filter(no__startswith=prefix).last()
        latest_receipt_no = int(avoirsuppliers.no.split('/')[1])
        receipt_no = f"FAV{year}/{latest_receipt_no + 1}"
    except:
        receipt_no = f"FAV{year}/1"
    print(receipt_no)
    avoir=Avoirsupplier.objects.create(
        no=receipt_no,
        supplier_id=supplierid,
        total=totalbon,
        date=datebon,
        avoirfacture=avoirfacture
    )
    supplier.rest-=float(totalbon)
    supplier.save()
    with transaction.atomic():
        for i in json.loads(products):
            product=Produit.objects.get(pk=i['productid'])
            product.stocktotal=int(product.stocktotal)-int(i['qty'])
            if avoirfacture:
                product.stockfacture=int(product.stockfacture)-int(i['qty'])
            product.save()
            Returnedsupplier.objects.create(
                avoir=avoir,
                product=product,
                qty=i['qty'],
                remise=0 if i['remise']=="" else i['remise'],
                price=i['price'],
                total=i['total'],
            )


    # increment it
    return JsonResponse({
        'success':True
    })


def checkbl(request, id):
    client=Client.objects.get(pk=id)
    print(Bonlivraison.objects.filter(client=client).first())
    if Bonlivraison.objects.filter(client=client).first() is not None:
        return JsonResponse({
            'hasbl':True
        })
    else:
        return JsonResponse({
            'hasbl':False
        })

def checkfacture(request, id):
    client=Client.objects.get(pk=id)
    print(client.name)
    print(Facture.objects.filter(client=client).first())
    if Facture.objects.filter(client=client).first() is not None:
        # get products
        return JsonResponse({
            'hasfacture':True
        })
    else:
        return JsonResponse({
            'hasfacture':False
        })

def sendrelevclient(request):
    clientcode=request.GET.get('clientcode')
    client=Client.objects.get(code=clientcode)
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    avoirs=Avoirclient.objects.filter(client=client, avoirfacture=False, date__range=[startdate, enddate])
    reglementsbl=PaymentClientbl.objects.filter(client=client, date__range=[startdate, enddate])
    bons=Bonlivraison.objects.filter(client=client, date__range=[startdate, enddate], total__gt=0)
    # totalcredit=round(avoirs.aggregate(Sum('total'))['total__sum'], 2)+round(reglementsbl.aggregate(Sum('amount'))['amount__sum'], 2)
    # totaldebit=round(bons.aggregate(Sum('total'))['total__sum'], 2)
    # sold=round(totaldebit-totalcredit, 2)

    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'Bonlivraison') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return JsonResponse({
        'html':render(request, 'relevecl.html', {
            'releve':[sorted_releve[i:i+34] for i in range(0, len(sorted_releve), 34)],
            'client':client,

            'startdate':startdate,
            'enddate':enddate,

        }).content.decode('utf-8'),
        'soldfc':client.soldfacture,
    })

def relevclient(request):
    clientid=request.POST.get('clientid')
    client=Client.objects.get(pk=clientid)
    startdate=request.POST.get('datefrom')
    enddate=request.POST.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate])
    reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate])
    bons=Bonlivraison.objects.filter(client_id=clientid, date__range=[startdate, enddate], total__gt=0)
    #here we calculate the new sold and assign it, inorder to avoid previous calculation problems
    # avoirsfc=round(Avoirclient.objects.filter(client_id=clientid, avoirfacture=True, date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'] or 0, 2)
    #
    # reglementsfc=round(PaymentClientfc.objects.filter(client_id=clientid, date__range=[startdate, enddate]).aggregate(Sum('amount'))['amount__sum'] or 0, 2)
    #
    # factures=round(Facture.objects.filter(client_id=clientid, date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'] or 0, 2)

    # soldfc with no range
    norangeavoirsfc=round(Avoirclient.objects.filter(client_id=clientid, avoirfacture=True).aggregate(Sum('total'))['total__sum'] or 0, 2)

    norangereglementsfc=round(PaymentClientfc.objects.filter(client_id=clientid).aggregate(Sum('amount'))['amount__sum'] or 0, 2)

    norangefactures=round(Facture.objects.filter(client_id=clientid).aggregate(Sum('total'))['total__sum'] or 0, 2)



    norangeavoirsbl=round(Avoirclient.objects.filter(client_id=clientid, avoirfacture=False).aggregate(Sum('total'))['total__sum'] or 0, 2)
    norangereglementsbl=round(PaymentClientbl.objects.filter(client_id=clientid).aggregate(Sum('amount'))['amount__sum'] or 0, 2)
    norangebons=round(Bonlivraison.objects.filter(client_id=clientid, total__gt=0).aggregate(Sum('total'))['total__sum'] or 0, 2)

    soldfcnorange=round(norangefactures-norangereglementsfc-norangeavoirsfc, 2)

    soldblnorange=round(norangebons-norangereglementsbl-norangeavoirsbl, 2)
    print('>> soldbl', soldblnorange)
    print('>> soldfc', soldfcnorange)
    client.soldfacture=soldfcnorange
    client.soldbl=soldblnorange
    client.soldtotal=round(soldfcnorange+soldblnorange, 2)
    client.save()
    # totalcredit=round(avoirs.aggregate(Sum('total'))['total__sum'], 2)+round(reglementsbl.aggregate(Sum('amount'))['amount__sum'], 2)
    # totaldebit=round(bons.aggregate(Sum('total'))['total__sum'], 2)
    # sold=round(totaldebit-totalcredit, 2)

    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'Bonlivraison') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return JsonResponse({
        'html':render(request, 'relevecl.html', {
            'releve':[sorted_releve[i:i+32] for i in range(0, len(sorted_releve), 32)],
            'client':client,

            'startdate':startdate,
            'enddate':enddate,

        }).content.decode('utf-8'),
        'soldfc':soldfcnorange,
    })


def relevsupplier(request):
    supplierid=request.POST.get('supplierid')
    supplier=Supplier.objects.get(pk=supplierid)
    startdate=request.POST.get('datefrom')
    enddate=request.POST.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    avoirs=Avoirsupplier.objects.filter(supplier_id=supplierid, avoirfacture=False, date__range=[startdate, enddate])
    reglementsbl = PaymentSupplier.objects.filter(
        supplier_id=supplierid,
        date__range=[startdate, enddate]
    ).filter(
        Q(bons__isfacture=False) | Q(bons__isnull=True)
    ).distinct()

    bons=Itemsbysupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfacture=False)
    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'Itemsbysupplier') for bon in bons),
    ((avoir, 'Avoirsupplier') for avoir in avoirs),
    ((reglementbl, 'Paymentsupplier') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return JsonResponse({
        'html':render(request, 'relevesupp.html', {
            'releve':sorted_releve,
            'supplier':supplier,
            'startdate':startdate,
            'enddate':enddate,

        }).content.decode('utf-8')
    })
def relevsupplierfc(request):
    supplierid=request.POST.get('supplierid')
    supplier=Supplier.objects.get(pk=supplierid)
    startdate=request.POST.get('datefrom')
    enddate=request.POST.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    avoirs=Avoirsupplier.objects.filter(supplier_id=supplierid, avoirfacture=True, date__range=[startdate, enddate])
    reglementsbl = PaymentSupplier.objects.filter(
        supplier_id=supplierid,
        date__range=[startdate, enddate],
        bons__isfacture=True
    ).distinct()
    bons=Itemsbysupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfacture=True)
    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'Itemsbysupplier') for bon in bons),
    ((avoir, 'Avoirsupplier') for avoir in avoirs),
    ((reglementbl, 'Paymentsupplier') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return JsonResponse({
        'html':render(request, 'relevesupp.html', {
            'releve':sorted_releve,
            'supplier':supplier,
            'startdate':startdate,
            'enddate':enddate,
            # to distanguish, sice we will use the same file
            'relvfc':True
        }).content.decode('utf-8')
    })


def sendrelevclientfc(request):
    clientcode=request.GET.get('clientcode')
    client=Client.objects.get(code=clientcode)
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    avoirs=Avoirclient.objects.filter(client=client, avoirfacture=True, date__range=[startdate, enddate])
    reglementsfc=PaymentClientfc.objects.filter(client=client, date__range=[startdate, enddate])

    bons=Facture.objects.filter(client=client, date__range=[startdate, enddate])

    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'Facture') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((reglementfc, 'PaymentClientfc') for reglementfc in reglementsfc),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)

    return JsonResponse({
        'html':render(request, 'releveclfc.html', {
            'releve':[sorted_releve[i:i+26] for i in range(0, len(sorted_releve), 26)],
            'client':client,
            'startdate':startdate,
            'enddate':enddate,

        }).content.decode('utf-8')
    })
def relevclientfc(request):
    clientid=request.POST.get('clientid')
    client=Client.objects.get(pk=clientid)
    startdate=request.POST.get('datefrom')
    enddate=request.POST.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=True, date__range=[startdate, enddate])
    reglementsfc=PaymentClientfc.objects.filter(client_id=clientid, date__range=[startdate, enddate])

    bons=Facture.objects.filter(client_id=clientid, date__range=[startdate, enddate])

    norangeavoirsfc=round(Avoirclient.objects.filter(client_id=clientid, avoirfacture=True).aggregate(Sum('total'))['total__sum'] or 0, 2)
    norangereglementsfc=round(PaymentClientfc.objects.filter(client_id=clientid).aggregate(Sum('amount'))['amount__sum'] or 0, 2)
    norangefactures=round(Facture.objects.filter(client_id=clientid).aggregate(Sum('total'))['total__sum'] or 0, 2)
    soldfacture=round(float(norangefactures)-float(norangeavoirsfc)-float(norangereglementsfc), 2)
    client.soldfacture=soldfacture
    client.save()
    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'Facture') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((reglementfc, 'PaymentClientfc') for reglementfc in reglementsfc),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)

    return JsonResponse({
        'html':render(request, 'releveclfc.html', {
            'releve':[sorted_releve[i:i+26] for i in range(0, len(sorted_releve), 26)],
            'client':client,
            'startdate':startdate,
            'enddate':enddate,

        }).content.decode('utf-8')
    })

def getclientrep(request, id):
    client=Client.objects.get(pk=id)
    print(client.represent_id)
    return JsonResponse({
        'id':client.represent_id,
        'name':client.represent.name
    })

def listbonachat(request):
    thisyear=timezone.now().year
    facture=request.GET.get('facture')=='1'
    print('>> facture', facture)
    thisyear=timezone.now().year
    if facture:
        bons=Itemsbysupplier.objects.filter(date__year=thisyear, isfacture=True).order_by('-date')
    else:
        bons=Itemsbysupplier.objects.filter(date__year=thisyear).order_by('-date')
    ctx={
        'title':'List des bon achat',
        'bons':bons[:50],
        'loadmore':True,
        'facturesection':facture
    }
    if bons:
        ctx['total']=round(bons.aggregate(Sum('total'))['total__sum'], 2)
        ctx['totaltva']=round(bons.aggregate(Sum('tva'))['tva__sum'], 2)
    # load more is true by default, but when facture, set it to be false
    if facture:
        ctx['loadmore']=False
    return render(request, 'listbonachat.html', ctx)

def listboncommnd(request):
    today = timezone.now().date()
    thisyear=timezone.now().year
    current_time = datetime.now().strftime('%H:%M:%S')
    # serverip = Setting.objects.only('serverip').first()
    # serverip = serverip.serverip if serverip else None
    ordersnotif=Ordersnotif.objects.first()
    if ordersnotif:
        # means ther is order
        orders=ordersnotif.orders
        # items=data['items']
        #print('orders', orders, 'items', items)
        for o in orders:
            # 1. Create order
            order = Order.objects.create(
                date=o['date'],
                total=o['total'],
                note=o['note'],
                client=Client.objects.get(code=o['clientcode']),
                salseman=Represent.objects.get(pk=o['salsemanid']),
                order_no=o['order_no'],
                isclientcommnd=o['isclientcommnd'],
            )

            items = o['items']

            # 2. Collect uniqcodes
            uniqcodes = [item['uniqcode'] for item in items]

            # 3. Fetch products in ONE query
            produits = Produit.objects.filter(uniqcode__in=uniqcodes)
            produit_map = {p.uniqcode: p for p in produits}

            # 4. Prepare Orderitem objects
            order_items = [
                Orderitem(
                    order=order,
                    product=produit_map[item['uniqcode']],
                    qty=item['qty'],
                    price=item['price'],
                    total=item['total'],
                )
                for item in items
            ]

            # 5. Bulk insert
            Orderitem.objects.bulk_create(order_items)
        Ordersnotif.objects.all().delete()
    orders=Order.objects.filter(date__year=thisyear).order_by('-id')[:50]
    ctx={
        'title':'List des bon commnd',
        'bons':orders,
        'today':timezone.now().date()
    }
    if orders:
        ctx['total']=round(Order.objects.all().aggregate(Sum('total'))['total__sum'], 2)
    return render(request, 'listboncommnd.html', ctx)

def bonachatdetails(request, id):
    bon=Itemsbysupplier.objects.get(pk=id)
    items=Stockin.objects.filter(nbon=bon)
    payments=PaymentSupplier.objects.filter(bons__in=[bon])
    orderitems=[items[i:i+36] for i in range(0, len(items), 36)]

    ctx={
        'title':f'Bon achat {bon.nbon}',
        'bon':bon,
        'items':items,
        'payments':payments,
        'orderitems':orderitems
    }
    return render(request, 'bonachatdetails.html', ctx)

def modifierbonachat(request, id):
    bon=Itemsbysupplier.objects.get(pk=id)
    items=Stockin.objects.filter(nbon=bon)
    ctx={
        'title':'Modifier bon achat',
        'bon':bon,
        'items':items,
        'suppliers':Supplier.objects.all()
    }
    return render(request, 'modifierbonachat.html', ctx)


def updatebonachat(request):
    id=request.POST.get('bonid')
    bon=Itemsbysupplier.objects.get(pk=id)
    bon.date=datetime.strptime(request.POST.get('datebon'), '%Y-%m-%d')
    bon.nbon=request.POST.get('orderno')
    isfacture= True if request.POST.get('mode')=='facture' else False
    totalbon=request.POST.get('totalbon')
    tva=round(float(totalbon)-(float(totalbon)/1.2), 2)
    supplier=Supplier.objects.get(pk=request.POST.get('supplierid'))
    thissupplier=bon.supplier
    if bon.supplier==supplier:
        print('same supplier', float(thissupplier.rest), float(bon.total), float(totalbon))
        thissupplier.rest=round(float(thissupplier.rest)-float(bon.total)+float(totalbon), 2)
        thissupplier.save()
    else:
        print('not same')
        thissupplier.rest=round(float(thissupplier.rest)-float(bon.total), 2)
        thissupplier.save()
        print('old', thissupplier.rest)
        supplier.rest=round(float(supplier.rest)+float(totalbon), 2)
        supplier.save()
        print('new', supplier.rest)
    # bon.supplier.rest=float(bon.supplier.rest)-float(bon.total)
    # bon.supplier.save()
    items=Stockin.objects.filter(nbon=bon)
    # update this items
    for i in items:
        product=i.product
        print('removing from total')
        product.stocktotal=int(product.stocktotal)-int(i.quantity)
        if bon.isfacture:
            print('removing from facture')
            product.stockfacture=int(product.stockfacture)-int(i.quantity)
        product.save()
        i.delete()

    bon.supplier=supplier
    bon.total=totalbon
    bon.nbon=request.POST.get('orderno')
    bon.isfacture=isfacture
    if isfacture:
        bon.tva=tva
    bon.save()

    with transaction.atomic():
        for i in json.loads(request.POST.get('products')):


            qty=0 if i['qty']=="" else int(i['qty'])
            product=Produit.objects.get(pk=i['productid'])
            print('>>>>>>>adding total')
            product.stocktotal=int(product.stocktotal)+qty
            if isfacture:
                print('>>>>>>>adding fc')
                product.stockfacture=int(product.stockfacture)+qty
            #product.save()
            # create new livraison items
            Stockin.objects.create(
                nbon=bon,
                supplier=supplier,
                remise=0 if i['remise']=="" else i['remise'],
                devise=0 if i['devise']=="" else i['devise'],
                product=product,
                ref=i['ref'],
                name=i['name'],
                date=datetime.strptime(request.POST.get('datebon'), '%Y-%m-%d'),
                quantity=qty,
                price=0 if i['price']=="" else i['price'],
                total=0 if i['total']=="" else i['total'],
                facture=isfacture
            )
            totalprices=Stockin.objects.filter(product=product).aggregate(Sum('total'))['total__sum'] or 0
            totalqty=Stockin.objects.filter(product=product).aggregate(Sum('quantity'))['quantity__sum'] or 0
            print(totalprices, totalqty)
            product.coutmoyen=round(totalprices/totalqty, 2)
            product.buyprice=0 if i['price']=="" else i['price']
            product.save()

    return JsonResponse({
        'success':True
    })

def getsuppbons(request):
    supplierid=request.POST.get('supplierid')
    bons=Itemsbysupplier.objects.filter(supplier_id=supplierid).order_by('-date')
    trs=''
    for i in bons:
        trs+=f'<tr style="background:{"aliceblue" if i.ispaid else "none"}; color:{"red" if i.ismanual and i.isfacture else "none"};"><td>{i.date.strftime("%d/%m/%Y")} ({"FC" if i.isfacture else "BL"})</td><td>{i.nbon}</td><td>{i.supplier.name}</td><td>{i.total}</td><td class="text-danger">{i.rest if i.rest>0 else "---"}</td> <td><input type="checkbox" value="{i.id}" name="bonsachattopay" onchange="checkreglementbox(event)"></td></tr>'

    return JsonResponse({
        'trs':trs
    })

def reglebonsachat(request):
    supplierid=request.POST.get('supplierid')
    supplier=Supplier.objects.get(pk=supplierid)
    bons=json.loads(request.POST.get('bons'))
    mantant=json.loads(request.POST.get('mantant'))
    mode=json.loads(request.POST.get('mode'))
    npiece=json.loads(request.POST.get('npiece'))
    date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=json.loads(request.POST.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]

    livraisons=Itemsbysupplier.objects.filter(pk__in=bons)
    livraisons.update(ispaid=True)
    totalmantant=sum(mantant)
    # livraisonstotals=0
    # for i in livraisons:
    #     if i.rest>0:
    #         livraisonstotals=round(livraisonstotals+i.rest, 2)
    #     else:
    #         livraisonstotals=round(livraisonstotals+i.total, 2)

    # # update client sold
    # # case1: 5000==5000:
    # if totalmantant==livraisonstotals:
    #     print('case1')
    #     livraisons.update(ispaid=True)
    #     livraisons.update(rest=0)
    # # case2: 5000>4500:
    # if totalmantant>livraisonstotals:
    #     print('case2')
    #     diff=totalmantant-livraisonstotals
    #     livraisons.update(ispaid=True)
    #     livraisons.update(rest=0)
    #     # else:
    #     #     livraisons.update(ispaid=True)
    #     # sub diff from client ref
    # # case3: 5000<6000:
    # if totalmantant<livraisonstotals:
    #     print('case3')
    #     amount=totalmantant
    #     for i in livraisons:
    #         if amount<=0:
    #             break
    #         else:
    #             if i.rest>0:
    #                 if amount>=i.rest:
    #                     i.rest=0
    #                     i.ispaid=True
    #                     i.save()
    #                     amount-=i.rest
    #                 else:

    #                     i.rest=round(i.rest-amount, 2)
    #                     i.save()
    #                     break
    #             else:
    #                 if amount>=i.total:
    #                     amount-=i.total
    #                     i.ispaid=True
    #                     i.save()
    #                 else:
    #                     i.rest=round(i.total-amount, 2)
    #                     i.save()
    #                     break
    #     print('rest of amount', amount)

    for m, mod, np, ech in zip(mantant, mode, npiece, echeance):
        regl=PaymentSupplier.objects.create(
            supplier_id=supplierid,
            amount=m,
            date=date,
            echeance=ech,
            mode=mod,
            npiece=np.lower(),
        )
        regl.bons.set(livraisons)
        # for i in livraisons:
        #     Bonsregle.objects.create(
        #         payment=regl,
        #         bon=i,
        #         amount=m
        #     )

    supplier.rest=round(float(supplier.rest)-float(totalmantant), 2)
    supplier.save()
    return JsonResponse({
        "success":True
    })


def journalachat(request):
    thisyear=timezone.now().year
    items=Stockin.objects.filter(isinventaire=False, date__year=thisyear).order_by('-id')[:50]
    ctx={
        'title':'Journal Achat',
        'items':items,
        'today':timezone.now().date(),
        'totaljach':round(Stockin.objects.filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalqtyjach':round(Stockin.objects.filter(date__year=thisyear).aggregate(Sum('quantity'))['quantity__sum'] or 0, 2),
    }
    return render(request, 'journalachat.html', ctx)

def laodjournalachat(request):
    page = int(request.GET.get('page', 1))

    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    products = Stockin.objects.all()[start:end]
    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalacha-row">
            <td>{i.date}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page
    })

def journalachatfc(request):
    thisyear=timezone.now().year
    items=Stockin.objects.filter(facture=True, date__year=thisyear)[:50]
    print('>>>>>>>>>>', round(Stockin.objects.filter(facture=True, date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2),)
    ctx={
        'title':'Journal Achat Facture',
        'items':items,
        'today':timezone.now().date(),
        'total':round(Stockin.objects.filter(facture=True, date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalqty':round(Stockin.objects.filter(facture=True, date__year=thisyear).aggregate(Sum('quantity'))['quantity__sum'] or 0, 2),
    }
    return render(request, 'journalachatfc.html', ctx)

def loadjournalachatfc(request):
    page = int(request.GET.get('page', 1))
    print(">>page ", page)
    year = request.GET.get('year')
    print(">>>>> journal achat")

    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    products = Stockin.objects.filter(facture=True, date__year=year).order_by('-id')[start:end]
    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalachafc-row">
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total:.2f}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page
    })

def loadjournalachat(request):
    page = int(request.GET.get('page', 1))

    per_page = 100  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    products = Stockin.objects.order_by('-id')[start:end]
    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalacha-row">
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total:.2f}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page
    })


def journalvente(request):
    thisyear=timezone.now().year
    items=Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).order_by('-date')[:50]
    bons=Bonlivraison.objects.filter(date__year=thisyear)
    totaltotal=round(bons.aggregate(Sum('total'))['total__sum'] or 0, 2)
    total2=round(Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
    print('>>>>>', totaltotal, total2)
    ctx={
        'title':'Journal vente',
        'items':items,
        'totalqty':Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).aggregate(Sum('qty'))['qty__sum'] or 0,
        'totaltotal':round(Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'today':timezone.now().date()
    }
    return render(request, 'journalvente.html', ctx)

def yeardatajournalv(request):
    year=request.GET.get('year')
    print(year)
    items=Livraisonitem.objects.filter(isfacture=False, date__year=year, isinventaire=False).order_by('-date')[:50]
    trs=''
    totalmarge=0
    for i in items:
        try:
            marge_value = round((i.product.prixnet - (i.product.coutmoyen if i.product.coutmoyen else 0)) * i.qty, 2)
        except:
            marge_value = 0
        totalmarge+=marge_value
        trs+=f'''
        <tr class="journalvente-row" year="{year}">
            <td>{i.date.strftime('%d/%m/%Y') if i.date else ''}</td>
            <td>{i.bon.bon_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjv">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjv">{i.qty}</td>
            <td class="totaljv">{i.total}</td>
            <td>{i.bon.client.name}</td>
            <td>{i.bon.salseman.name}</td>
            <td class="text-success margejv">
                {marge_value}
            </td>
        </tr>
        '''

    return JsonResponse({
        'trs':trs,
        'totalqty':Livraisonitem.objects.filter(isfacture=False, date__year=year).aggregate(Sum('qty'))['qty__sum'] or 0,
        'total':round(Livraisonitem.objects.filter(isfacture=False, date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalmarge':round(totalmarge, 2)
    })

def yeardatajournalvfc(request):
    year=request.GET.get('year')
    print(year)
    items=Outfacture.objects.filter(date__year=year).order_by('-date')[:50]
    trs=''
    totalmarge=0
    for i in items:
        try:
            marge_value = round((i.product.prixnet - (i.product.coutmoyen if i.product.coutmoyen else 0)) * i.qty, 2)
        except:
            marge_value = 0
        totalmarge+=marge_value
        trs+=f'''
        <tr class="journalventefc-row" year="{year}">
            <td>{i.date.strftime('%d/%m/%Y') if i.date else ''}</td>
            <td>{i.facture.facture_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjv">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjv">{i.qty}</td>
            <td class="totaljv">{i.total}</td>
            <td class="totaljv"></td>
            <td>{i.facture.client.name}</td>
            <td>{i.facture.salseman.name}</td>
            <td class="text-success margejv">
                {marge_value}
            </td>
        </tr>
        '''

    return JsonResponse({
        'trs':trs,
        'totalqty':Outfacture.objects.filter(date__year=year).aggregate(Sum('qty'))['qty__sum'] or 0,
        'total':round(Outfacture.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalmarge':round(totalmarge, 2)
    })

def journalventefc(request):
    thisyear=timezone.now().year
    items=Outfacture.objects.filter(date__year=thisyear).order_by('-id')[0:50]
    ctx={
        'title':'Journal vente Facture',
        'items':items,
        'totalqty':Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('qty'))['qty__sum'] or 0,
        'total':round(Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0,2),
        'today':timezone.now().date()
    }
    return render(request, 'journalventefc.html', ctx)

def loadjournalvente(request):
    page = int(request.GET.get('page', 1))
    year = request.GET.get('year')
    term = request.GET.get('term')
    startdate = request.GET.get('startdate')
    enddate = request.GET.get('enddate')
    per_page = 50  # Adjust as needed
    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    if term != '0':

        # Split the term into individual words separated by '*'
        search_terms = term.split('+')
        # Create a list of Q objects for each search term and combine them with &

        q_objects = Q()
        for term in search_terms:
            if term:
                q_objects &= (Q(client__name__iregex=term)|Q(ref__iregex=term)|Q(name__iregex=term)|Q(total__iregex=term)|Q(bon__bon_no__iregex=term)|Q(bon__salseman__name__iregex=term))
        if year=='0':
            # means the year is not selected, so the records of the current year
            products = Livraisonitem.objects.filter(isfacture=False).filter(q_objects).order_by('-date')[start:end]
            total=round(Livraisonitem.objects.filter(isfacture=False).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
            totalqty=Livraisonitem.objects.filter(isfacture=False).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
        else:
            products = Livraisonitem.objects.filter(isfacture=False, date__year=year, isinventaire=False).filter(q_objects).order_by('-date')[start:end]
            total=round(Livraisonitem.objects.filter(isfacture=False, date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
            totalqty=Livraisonitem.objects.filter(isfacture=False, date__year=year).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
        for i in products:
            trs+=f'''
            <tr class="journalvente-row" year={year} term={term}>
                <td>{i.date.strftime('%d/%m/%Y')}</td>
                <td>{i.bon.bon_no}</td>
                <td>{i.product.ref.upper()}</td>
                <td>{i.product.name}</td>
                <td>{i.price}</td>
                <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
                <td style="color:blue" class="">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
                <td class="text-danger coutmoyenjv">{i.product.buyprice if i.product.buyprice else 0}</td>
                <td class="text-danger qtyjv">{i.qty}</td>
                <td class="totaljv">{i.total}</td>
                <td>{i.bon.client.name}</td>
                <td>{i.bon.client.code}</td>
                <td>{i.bon.salseman.name}</td>
                <td class="text-success margejv">

                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(products) == per_page,
        })
    if startdate!='0' and enddate!='0':
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        products=Livraisonitem.objects.filter(isfacture=False, date__range=[startdate, enddate], isinventaire=False).order_by('-date')[start:end]
        trs=''
        for i in products:
            trs+=f'''
            <tr class="journalvente-row" startdate={startdate} enddate={enddate}>
                <td>{i.date.strftime('%d/%m/%Y')}</td>
                <td>{i.bon.bon_no}</td>
                <td>{i.product.ref.upper()}</td>
                <td>{i.product.name}</td>
                <td>{i.price}</td>
                <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
                <td style="color:blue" class="">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
                <td class="text-danger coutmoyenjv">{i.product.buyprice if i.product.buyprice else 0}</td>
                <td class="text-danger qtyjv">{i.qty}</td>
                <td class="totaljv">{i.total}</td>
                <td>{i.bon.client.name}</td>
                <td>{i.bon.client.code}</td>
                <td>{i.bon.salseman.name}</td>
                <td class="text-success margejv">

                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(products) == per_page,
        })
    if year=='0':
        # means the year i not selected, so the records of the current year
        products = Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).order_by('-date')[start:end]
    else:
        products = Livraisonitem.objects.filter(isfacture=False, date__year=year, isinventaire=False).order_by('-date')[start:end]
    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalvente-row" year={year}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.bon.bon_no}</td>
            <td>{i.product.ref.upper()}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger coutmoyenjv">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjv">{i.qty}</td>
            <td class="totaljv">{i.total}</td>
            <td>{i.bon.client.name}</td>
            <td>{i.bon.client.code}</td>
            <td>{i.bon.salseman.name}</td>
            <td class="text-success margejv">

            </td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page
    })

def loadjournalventefc(request):
    thisyear=timezone.now().year
    page = int(request.GET.get('page', 1))
    term=request.GET.get('term')
    year=request.GET.get('year')
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    per_page = 50  # Adjust as needed
    start = (page - 1) * per_page
    end = page * per_page
    print('>>', page, start, end)
    if term != '0':
        print('>> term in term')
        search_terms = term.split('+')
        # Create a list of Q objects for each search term and combine them with &

        q_objects = Q()
        for term in search_terms:
            if term:
                q_objects &= (Q(client__name__iregex=term)|Q(ref__iregex=term)|Q(name__iregex=term)|Q(total__iregex=term)|Q(facture__facture_no__iregex=term)|Q(facture__salseman__name__iregex=term))
        if year=='0':
            print('>> interm and year')
            # means the year i not selected, so the records of the current year
            products = Outfacture.objects.filter(date__year=thisyear).filter(q_objects).order_by('-id')[start:end]
            total=round(Outfacture.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
            totalqty=Outfacture.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
        else:
            products = Outfacture.objects.filter(date__year=year).filter(q_objects).order_by('-id')[start:end]
            total=round(Outfacture.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
            totalqty=Outfacture.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
        trs=''
        for i in products:
            trs+=f'''
            <tr class="journalventefc-row" year={year} term={term}>
                <td>{i.date.strftime('%d/%m/%Y')}</td>
                <td>{i.facture.facture_no}</td>
                <td>{i.product.ref}</td>
                <td>{i.product.name}</td>
                <td>{i.price}</td>
                <td class="prnetjvfc">{i.product.prixnet if i.product.prixnet else 0}</td>
                <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
                <td class="text-danger prachatjvfc">{i.product.buyprice if i.product.buyprice else 0}</td>
                <td class="text-danger qtyjvfc">{i.qty}</td>
                <td class="totaljvfc">{i.total}</td>
                <td></td>
                <td>{i.facture.client.name}</td>
                <td>{i.facture.client.code}</td>
                <td>{i.facture.salseman.name}</td>
                <td class="text-success margejvfc">

                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(products) == per_page,
            'total':total,
            'totalqty':totalqty
        })

    if startdate != '0' and enddate != '0':
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        bons=Outfacture.objects.filter(date__range=[startdate, enddate]).order_by('-id')[start:end]
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="journalventefc-row" startdate={startdate} enddate={enddate}>
                <td>{i.date.strftime('%d/%m/%Y')}</td>
                <td>{i.facture.facture_no}</td>
                <td>{i.product.ref}</td>
                <td>{i.product.name}</td>
                <td>{i.price}</td>
                <td class="prnetjvfc">{i.product.prixnet if i.product.prixnet else 0}</td>
                <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
                <td class="text-danger prachatjvfc">{i.product.buyprice if i.product.buyprice else 0}</td>
                <td class="text-danger qtyjvfc">{i.qty}</td>
                <td class="totaljvfc">{i.total}</td>
                <td></td>
                <td>{i.facture.client.name}</td>
                <td>{i.facture.client.code}</td>
                <td>{i.facture.salseman.name}</td>
                <td class="text-success margejvfc">

                </td>
            </tr>
            '''
        ctx={
            'trs':trs,
            'has_more': len(products) == per_page,
        }
        if bons:
            ctx['total']=round(Outfacture.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'], 2)
            ctx['totalqty']=Outfacture.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('qty')).get('qty__sum')
        return JsonResponse(ctx)

    print('>> just year')
    products = Outfacture.objects.filter(date__year=year).order_by('-id')[start:end]
    total=round(Outfacture.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    totalqty=Outfacture.objects.filter(date__year=year).aggregate(Sum('qty'))['qty__sum'] or 0

    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalventefc-row" year={year}>
            <td>{i.date.strftime("%d/%m/%Y")}</td>
            <td>{i.facture.facture_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjvfc">{i.product.prixnet}</td>
            <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen}</td>
            <td class="text-danger prachatjvfc">{i.product.buyprice}</td>
            <td class="text-danger qtyjvfc">{i.qty}</td>
            <td>{i.total}</td>
            <td></td>
            <td>{i.client.name}</td>
            <td>{i.client.code}</td>
            <td>{i.facture.salseman.name}</td>
            <td class="text-success margejvfc">

            </td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page,
        'total':(Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalqty':Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('qty'))['qty__sum'] or 0,
    })

# product search selects2 for bons
def searchproduct(request):
    # get url pams
    term=request.GET.get('term')
    if not '+' in term:
        print('>> from not +')
        products=Produit.objects.filter(ref=term)

        q_objects = Q()
        q_objects &= (
            Q(ref__icontains=term) |
            Q(name__icontains=term) |
            Q(mark__name__icontains=term) |
            Q(category__name__icontains=term) |
            Q(equivalent__icontains=term) |
            Q(refeq1__icontains=term) |
            Q(refeq2__icontains=term) |
            Q(refeq3__icontains=term) |
            Q(refeq4__icontains=term) |
            Q(diametre__icontains=term)
        )
        # adding other products that have equivalent
        # if not products:
        #     products=Produit.objects.filter(q_objects).order_by('-stocktotal')
        products=products | Produit.objects.filter(q_objects).order_by('-stocktotal')
    else:
        print('>>> from +')
        # Split the term into individual words separated by '*'
        search_terms = term.split('+')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:
            if term:
                q_objects &= (
                    Q(ref__icontains=term) |
                    Q(name__icontains=term) |
                    Q(mark__name__icontains=term) |
                    Q(category__name__icontains=term) |
                    Q(equivalent__icontains=term) |
                    Q(refeq1__icontains=term) |
                    Q(refeq2__icontains=term) |
                    Q(refeq3__icontains=term) |
                    Q(refeq4__icontains=term)|
                    Q(diametre__icontains=term)
                )
        # check if term in product.ref or product.name
        products=Produit.objects.filter(q_objects).order_by('-stocktotal')

    results=[]
    for i in products:
        results.append({
            'id':f'{i.ref}§{i.name}§{i.buyprice}§{i.stocktotal}§{i.stockfacture}§{i.id}§{i.sellprice}§{i.remise}§{i.prixnet}§{i.representprice}',
            'text':f'{i.ref.upper()} - {i.name.upper()}',
            'stock':i.stocktotal,
            'stockfacture':i.stockfacture
        })
    return JsonResponse({'results': results})

def filterbldate(request):
    facture=request.GET.get('facture')=='1'
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    print(startdate, enddate)
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Bonlivraison.objects.filter(date__range=[startdate, enddate]).order_by('-bon_no')[:50]
    if facture:
        bons=Bonlivraison.objects.filter(date__range=[startdate, enddate], isfacture=True).order_by('-bon_no')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="ord {"text-danger" if i.ispaid else ''} bl-row" startdate={startdate} enddate={enddate} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')">
            <td>{ i.bon_no }</td>
            <td>{ i.date.strftime("%d/%m/%Y")}</td>
            <td>{ i.client.name }</td>
            <td>{ i.client.code }</td>
            <td>{ i.total}</td>
            <td>{ i.client.region}</td>
            <td>{ i.client.city}</td>
            <td>{ i.client.soldbl}</td>
            <td>{ i.salseman }</td>
            <td class="d-flex justify-content-between">
              <div>
              {'R0' if i.ispaid else 'N1' }

              </div>
              <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

            </td>
            <td class="text-danger">
            {'OUI' if i.isfacture else 'NON'}

            </td>

            <td>
              {i.commande.order_no if i.commande else '--'}
            </td>
            <td>
              {i.modlvrsn}
            </td>
          </tr>
        '''
    ctx={
        'trs':trs
    }
    if bons:
        ctx['total']=round(Bonlivraison.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total')).get('total__sum'), 2)
    return JsonResponse(ctx)

def searchclient(request):
    term=request.GET.get('term')
    #regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = term.split('+')

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(name__icontains=term) |
                Q(code__icontains=term) |
                Q(region__icontains=term) |
                Q(city__icontains=term) |
                Q(address__icontains=term))
    clients=Client.objects.filter(q_objects)
    # if '+' in term:
    #     term=term.split('+')
    #     for i in term:
    #         clients=Client.objects.filter(
    #             Q(name__icontains=i) |
    #             Q(code__icontains=i) |
    #             Q(region__icontains=i) |
    #             Q(city__icontains=i)
    #         )
    # else:
    #     clients=Client.objects.filter(
    #         Q(name__icontains=term) |
    #         Q(code__icontains=term) |
    #         Q(region__icontains=term) |
    #         Q(city__icontains=term)
    #     )
    results=[]
    for i in clients:
        results.append({
            'id':i.id,
            'text':f'{i.name} - {i.city}',
            'diver':i.diver
        })
    return JsonResponse({'results': results})
def searchclientrep(request):
    rep=request.user.represent
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(name__icontains=term) |
                Q(code__icontains=term) |
                Q(region__icontains=term) |
                Q(city__icontains=term) |
                Q(address__icontains=term))
    clients=Client.objects.filter(represent=rep).filter(q_objects)
    # if '+' in term:
    #     term=term.split('+')
    #     for i in term:
    #         clients=Client.objects.filter(
    #             Q(name__icontains=i) |
    #             Q(code__icontains=i) |
    #             Q(region__icontains=i) |
    #             Q(city__icontains=i)
    #         )
    # else:
    #     clients=Client.objects.filter(
    #         Q(name__icontains=term) |
    #         Q(code__icontains=term) |
    #         Q(region__icontains=term) |
    #         Q(city__icontains=term)
    #     )
    results=[]
    for i in clients:
        results.append({
            'id':i.id,
            'text':f'{i.name} - {i.city}'
        })
    return JsonResponse({'results': results})

def searchsupplier(request):
    term=request.GET.get('term')
    print(term)
    suppliers=Supplier.objects.filter(Q(name__icontains=term) | Q(phone__icontains=term)| Q(address__icontains=term))
    print('suppliers', suppliers)
    results=[]
    for i in suppliers:
        results.append({
            'id':i.id,
            'text':i.name
        })
    return JsonResponse({'results': results})

def getclientfactureprice(request):
    id=request.POST.get('id')
    clientid=request.POST.get('clientid')
    print(id, 'rr', clientid)
    try:
        clientprice=Outfacture.objects.filter(client_id=clientid, product_id=id).last()
        price=clientprice.price
        remise=clientprice.remise
        return JsonResponse({
            'price':price,
            'remise':remise
        })
    except:
        return JsonResponse({
            'price':0,
            'remise':0
        })


def updatereglebons(request):
    reglementid=request.POST.get('reglementid')
    mantant=request.POST.get('mantant')
    mode=request.POST.get('mode')
    npiece=request.POST.get('npiece')
    date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=request.POST.get('echeance')
    echeance=datetime.strptime(echeance, '%Y-%m-%d') if echeance!='' else None
    newbons=json.loads(request.POST.get('bons'))
    reglement=PaymentClientbl.objects.get(pk=reglementid)

    thisclient=reglement.client
    #print('soldbl', float(thisclient.soldbl),float(reglement.amount),float(mantant))
    thisclient.soldbl=round(float(thisclient.soldbl)+float(reglement.amount)-float(mantant), 2)
    thisclient.soldtotal=round(float(thisclient.soldtotal)+float(reglement.amount)-float(mantant), 2)
    thisclient.save()
    oldbons=reglement.bons.all()
    livraisons=Bonlivraison.objects.filter(pk__in=newbons)
    for i in oldbons:
        otherregl=PaymentClientbl.objects.filter(bons=i.id).exclude(pk=reglementid)
        if not otherregl.exists():
            i.ispaid=False
            i.statusreg='n1'
            i.save()
    #oldbons.update(ispaid=False)
    #oldbons.update(statusreg='n1')
    #print(oldbons)
    #print(livraisons)
    livraisons.update(ispaid=True)
    livraisons.update(statusreg='r0')
    reglement.bons.clear()

    # # update regleemnt amount
    reglement.date=date
    reglement.amount=mantant
    reglement.mode=mode
    reglement.npiece=npiece.lower()
    reglement.echance=echeance

    reglement.bons.set(livraisons)
    reglement.save()
    return JsonResponse({
        'success':True
    })
    # thisclient.soldbl=round(float(thisclient.soldbl)-float(reglement.amount)+float(mantant), 2)
    # substract the old total from client soldbl


def updatereglesupp(request):
    reglementid=request.POST.get('reglementid')
    mantant=request.POST.get('mantant')
    mode=request.POST.get('mode')
    npiece=request.POST.get('npiece')
    date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=request.POST.get('echeance')
    echeance=datetime.strptime(echeance, '%Y-%m-%d') if echeance!='' else None
    newbons=json.loads(request.POST.get('bons'))
    reglement=PaymentSupplier.objects.get(pk=reglementid)

    thissupplier=reglement.supplier
    thissupplier.rest=round(float(thissupplier.rest)+float(reglement.amount)-float(mantant), 2)
    thissupplier.save()
    oldbons=reglement.bons.all()
    livraisons=Itemsbysupplier.objects.filter(pk__in=newbons)
    oldbons.update(ispaid=False)
    livraisons.update(ispaid=True)
    reglement.bons.clear()

    # # update regleemnt amount
    reglement.date=date
    reglement.amount=mantant
    reglement.mode=mode
    reglement.npiece=npiece.lower()
    reglement.echance=echeance

    reglement.bons.set(livraisons)
    reglement.save()
    return JsonResponse({
        'success':True
    })
    # thisclient.soldbl=round(float(thisclient.soldbl)-float(reglement.amount)+float(mantant), 2)
    # substract the old total from client soldbl

def getreglementbl(request, id):
    reglement=PaymentClientbl.objects.get(pk=id)
    bons=reglement.bons.all().order_by('date')
    # bons without bons in reglement
    livraisons=Bonlivraison.objects.filter(client=reglement.client).exclude(pk__in=[bon.pk for bon in bons]).order_by('date')[:50]
    # livraisons=Bonlivraison.objects.filter(client=reglement.client)
    #we need bons to calculate total bl
    bonstocalculate=Bonlivraison.objects.filter(client=reglement.client)
    trs=''
    for i in livraisons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="loadblinupdateregl" reglemntid="{id}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'date':reglement.date.strftime('%Y-%m-%d'),
        'echance':reglement.echance.strftime('%Y-%m-%d') if reglement.echance else '',
        'mode':reglement.mode,
        'npiece':reglement.npiece,
        'bons':list(bons.values()),
        'client':reglement.client.name,
        'clientid':reglement.client.id,
        'mantant':reglement.amount,
        'livraisons':trs,
        'total':round(bonstocalculate.aggregate(Sum('total'))['total__sum'] or 0, 2),
        'soldclient':reglement.client.soldbl
    })



def getreglementfc(request, id):
    reglement=PaymentClientfc.objects.get(pk=id)
    # facture of this reglement

    bons=reglement.factures.all().order_by('date')
    #print([i.id for i in bons] )
    livraisons=Facture.objects.filter(client=reglement.client).exclude(pk__in=[bon.pk for bon in bons]).order_by('date')[:50]
    #livraisons=Facture.objects.filter(client=reglement.client).order_by('date')[:50]
    bonstocalculate=Facture.objects.filter(client=reglement.client)
    trs=''
    for i in livraisons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="loadblinupdatereglfc" reglemntid="{id}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'date':reglement.date.strftime('%Y-%m-%d'),
        'echance':reglement.echance.strftime('%Y-%m-%d') if reglement.echance else '',
        'mode':reglement.mode,
        'npiece':reglement.npiece,
        'bons':list(bons.values()),
        'client':reglement.client.name,
        'clientid':reglement.client.id,
        'mantant':reglement.amount,
        'livraisons':trs,
        # 'livraisons':list(livraisons.values()),
        'soldclientfc':reglement.client.soldfacture,
        'total':round(bonstocalculate.aggregate(Sum('total'))['total__sum'] or 0, 2),


    })

def updatereglefactures(request):
    reglementid=request.POST.get('reglementid')
    mantant=request.POST.get('mantant')
    mode=request.POST.get('mode')
    npiece=request.POST.get('npiece')
    date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=request.POST.get('echeance')
    echeance=datetime.strptime(echeance, '%Y-%m-%d') if echeance!='' else None
    newbons=json.loads(request.POST.get('bons'))
    reglement=PaymentClientfc.objects.get(pk=reglementid)
    thisclient=reglement.client

    # get bons
    oldbons=reglement.factures.all()
    for i in oldbons:
        # search in other reglement
        otherregl=PaymentClientfc.objects.filter(factures=i.id).exclude(pk=reglementid)

        if not otherregl.exists():
            # this means there is no other regl in this fact
            i.ispaid=False
            i.statusreg='b1'
            i.save()
    # instead of update them to not paid, above code is correct, loop over them and if fact has no other reg then make it unpaid
    # oldbons.update(ispaid=False)
    # oldbons.update(statusreg='b1')
    reglement.factures.clear()
    livraisons=Facture.objects.filter(pk__in=newbons)
    livraisons.update(ispaid=True)
    livraisons.update(statusreg='f1')

    # new code, we get all nthe

    # # update regleemnt amount
    reglement.date=date
    reglement.amount=mantant
    reglement.mode=mode
    reglement.npiece=npiece.lower()
    reglement.echance=echeance
    reglement.factures.set(livraisons)

    # reglement.bons.set(livraisons)
    reglement.save()
    print(">> adding amount to soldfacture ans soldtotal, I made it at the end so that if there will be any error above the sold will not be affected")
    thisclient.soldfacture=round(float(thisclient.soldfacture)+float(reglement.amount)-float(mantant), 2)
    thisclient.soldtotal=round(float(thisclient.soldtotal)+float(reglement.amount)-float(mantant), 2)
    thisclient.save()
    return JsonResponse({
        'success':True
    })
    # thisclient.soldbl=round(float(thisclient.soldbl)-float(reglement.amount)+float(mantant), 2)
    # substract the old total from client soldbl

def getlastsuppprice(request):
    id=request.POST.get('id')
    supplierid=request.POST.get('supplierid')
    print(id, 'rr', supplierid)
    try:
        lastprice=Stockin.objects.filter(product_id=id, supplier_id=supplierid).last()
        print(lastprice)
        price=lastprice.price
        remise=lastprice.remise

        return JsonResponse({
            'price':price,
            'remise':remise,
            'facture':lastprice.facture
        })
    except:
        return JsonResponse({
            'price':0
        })


def boncommandedetails(request, id):

    order=Order.objects.get(pk=id)
    orderitems=Orderitem.objects.filter(order=order).order_by('product__name')
    reliquat=Orderitem.objects.filter(order__client_id=order.client.id, order__note__icontains='Reliquat', product__stocktotal__gt=F('qty'), islivraison= False).exists()

    print('>>>>>>Reliquat', reliquat)
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+40] for i in range(0, len(orderitems), 40)]
    ctx={
        'hasreliquat':reliquat,
        'title':f'Bon de Commande {order.order_no}',
        'order':order,
        'reliquat':'Reliquat' in order.note,
        'orderitems':orderitems,
    }
    return render(request, 'boncommandedetails.html', ctx)


def genererbonlivraison(request, id):
    towmonthsago=timezone.now() - timedelta(days=60)
    #  date__gte=towmonthsago
    order=Order.objects.get(pk=id)
    reliquas=Orderitem.objects.filter(order__client_id=order.client.id, order__note__icontains='Reliquat', product__stocktotal__gt=F('qty'), islivraison= False).order_by('-date')
    items=Orderitem.objects.filter(order=order).order_by('name')
    client=order.client
    three_months_ago = timezone.now() - timedelta(days=90)  # Assuming 30 days per month on average

    # Query for Bonlivraison objects that have a 'date' field earlier than three months ago
    bons = Bonlivraison.objects.filter(date__lt=three_months_ago, ispaid=False, total__gt=0, client=client).first()
    factures = Facture.objects.filter(date__lt=three_months_ago, ispaid=False, total__gt=0, client=client).first()
    oldunpaid=False
    if bons or factures:
        oldunpaid=True
    print('>> unpaid', oldunpaid)
    # we need date of last invoice
    #year = timezone.now().strftime("%y")
    #latest_receipt = Bonlivraison.objects.filter(
    #    bon_no__startswith=f'BL{year}'
    #).order_by("-id").first()
    #if latest_receipt:
    #    latest_receipt_no = int(latest_receipt.bon_no[-5:])
    #    receipt_no = f"BL{year}{latest_receipt_no + 1:05}"
    #else:
    #    receipt_no = f"BL{year}00001"
    ctx={
        'reliquas':reliquas,
        'order':order,
        'items':items,
        #'receipt_no':receipt_no,
        #'clients':Client.objects.all(),
        'reps':Represent.objects.all(),
        'today':timezone.now().date(),
        'oldunpaid':oldunpaid
    }
    return render(request, 'genererbonlivraison.html', ctx)


def createclientaccount(request):
    clientid=request.POST.get('clientid')
    client= Client.objects.get(pk=clientid)
    username=request.POST.get('username')
    password=request.POST.get('password')
    #check for username
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    if serverip:
        res=req.get(f'http://{serverip}/products/createclientaccount', {
            'clientcode':client.code,
            'username':username,
            'password':password
        })
        if json.loads(res.text)['success']:
            return JsonResponse({
                'success':True
            })
        else:
            return JsonResponse({
                'success':False,
                'error':f"{json.loads(res.text)['error']}"
            })
    # user=User.objects.filter(username=username).first()
    # if user:
    return JsonResponse({
        'success':False,
        'error':'No server'
    })

    # # create user
    # user=User.objects.create_user(username=username, password=password)
    # # assign user to client
    # group, created = Group.objects.get_or_create(name="clients")
    # user.groups.add(group)
    # user.save()
    # client.user=user
    # client.save()
    # serverip = Setting.objects.only('serverip').first()
    # serverip = serverip.serverip if serverip else None
    
    # return JsonResponse({
    #     'success':True
    # })


def createrepaccount(request):
    repid=request.POST.get('repid')
    rep= Represent.objects.get(pk=repid)
    username=request.POST.get('username')
    password=request.POST.get('password')
    #check for username
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    if serverip:
        res=req.get(f'http://{serverip}/products/createrepaccount', {
            'repid':repid,
            'username':username,
            'password':password
        })
        if json.loads(res.text)['success']:
            return JsonResponse({
                'success':True
            })
        else:
            return JsonResponse({
                'success':False,
                'error':f"{json.loads(res.text)['error']}"
            })
    # user=User.objects.filter(username=username).first()
    # if user:
    #     return JsonResponse({
    #         'success':False,
    #         'error':'Username exist déja'
    #     })

    # # create user
    # user=User.objects.create_user(username=username, password=password)
    # # assign user to rep
    # group, created = Group.objects.get_or_create(name="salsemen")
    # user.groups.add(group)
    # user.save()
    # rep.user=user
    # rep.save()
    # serverip = Setting.objects.only('serverip').first()
    # serverip = serverip.serverip if serverip else None
    # if serverip:
    #     req.get(f'http://{serverip}/products/createrepaccount', {
    #         'repid':repid,
    #         'username':username,
    #         'password':password
    #     })
    return JsonResponse({
        'success':True
    })

def carlogospage(request):
    ctx={
        # maintain same names
        'categories':Carlogos.objects.all(),
        'title':'Voitures logo'
    }
    return render(request, 'carlogos.html', ctx)

def createlogo(request):
    name=request.POST.get('logoname')
    # get image file
    image=request.FILES.get('logoimage', None)
    # create category
    files={'image':image}
    serverip= Setting.objects.only('serverip').first()
    serverip=serverip.serverip if serverip else None
    if serverip:
        try:
            res=req.post(f'http://{serverip}/products/createlogo', data={'name':name}, files=files)
            res.raise_for_status()
        except req.exceptions.RequestException as e:
            return JsonResponse({
                'error':str(e)
            })
    Carlogos.objects.create(name=name, image=image)
    ctx={
        'categories':Carlogos.objects.all(),
        'title':'Voiture logo'
    }
    return JsonResponse({
        'html':render(request, 'carlogos.html', ctx).content.decode('utf-8')
    })

def updatelogo(request):
    image=request.FILES.get('updatelogoimage') or None
    id=request.POST.get('id')
    carlogo=Carlogos.objects.get(pk=id)
    files={'image':image}
    serverip= Setting.objects.only('serverip').first()
    serverip=serverip.serverip if serverip else None
    if serverip:
        try:
            res=req.post(f'http://{serverip}/products/updatelogo', data={'id':id, 'name':request.POST.get('updatelogoname')}, files=files)
            res.raise_for_status()
        except req.exceptions.RequestException as e:
            return JsonResponse({
                'success':False,
                'error':str(e)
            })
    carlogo.name=request.POST.get('updatelogoname')
    if image:
        carlogo.image=image
    carlogo.save()
    ctx={
        'categories':Carlogos.objects.all(),
        'title':'Voiture logo'
    }
    return JsonResponse({
        'html':render(request, 'carlogos.html', ctx).content.decode('utf-8')
    })


def getnotpaid(request):
    # get bon livraison not paid more than 3 months

    three_months_ago = timezone.now() - timedelta(days=90)  # Assuming 30 days per month on average

    # Query for Bonlivraison objects that have a 'date' field earlier than three months ago
    bons = Bonlivraison.objects.filter(date__lt=three_months_ago, ispaid=False, total__gt=0).order_by('-date')


    return JsonResponse({
        'html':render(request, 'bllist.html', {'bons':bons, 'notloading':True}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total')).get('total__sum'), 2),

    })

def getnotpaidfc(request):
    # get bon livraison not paid more than 3 months

    three_months_ago = timezone.now() - timedelta(days=90)  # Assuming 30 days per month on average

    # Query for Bonlivraison objects that have a 'date' field earlier than three months ago
    bons = Facture.objects.filter(date__lt=three_months_ago, ispaid=False)


    return JsonResponse({
        'html':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total')).get('total__sum') or 0, 2)
    })


def filterfcdate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    facture=request.GET.get('facture')=='1'
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Facture.objects.exclude(client_id=3731 if facture else None).filter(date__range=[startdate, enddate]).order_by('-facture_no')[:50]
    print('total', Facture.objects.filter(date__range=[startdate, enddate]).count(), Facture.objects.count())
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" startdate={startdate} enddate={enddate} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
            <td>{ i.facture_no }</td>
            <td>{ i.date.strftime("%d/%m/%Y")}</td>
            <td>{ i.total}</td>
            <td>{ i.tva}</td>
            <td>{ i.client.name }</td>
            <td>{ i.client.code }</td>
            <td>{ i.client.region}</td>
            <td>{ i.client.city}</td>
            <td>{ i.client.soldfacture}</td>
            <td>{ i.salseman }</td>
            <td class="d-flex justify-content-between">
              <div>
              {'R0' if i.ispaid else 'N1' }

              </div>
              <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

            </td>
            <td>

                {i.note}
            </td>
            <td>
              {i.bon.bon_no if i.bon else "--"}
            </td>
          </tr>
        '''
    ctx={
        'trs':trs
    }
    if bons:
        ctx['total']=round(bons.aggregate(Sum('total')).get('total__sum'), 2)
        ctx['totaltva']=round(bons.aggregate(Sum('tva')).get('tva__sum'), 2)
    return JsonResponse(ctx)
    # return JsonResponse({
    #     'html':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8'),
    #     'total':round(bons.aggregate(Sum('total')).get('total__sum'), 2),
    #     'totaltva':round(bons.aggregate(Sum('tva')).get('tva__sum'), 2),

    # })

def filterfccopydate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    facture=request.GET.get('facture')=='1'
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Facture.objects.exclude(client_id=3731 if facture else None).filter(hascopy=True, date__range=[startdate, enddate]).order_by('-facture_no')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" startdate={startdate} enddate={enddate} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetailscopy/{i.id}')">
            <td>{ i.copynumber }</td>
            <td>{ i.date.strftime("%d/%m/%Y")}</td>
            <td>{ i.total}</td>
            <td>{ i.client.name }</td>
            <td>{ i.client.code }</td>
            <td>{ i.client.region}</td>
            <td>{ i.client.city}</td>
            <td>{ i.salseman }</td>

            <td>

                {i.note}
            </td>

          </tr>
        '''
    ctx={
        'trs':trs
    }
    if bons:
        ctx['total']=round(bons.aggregate(Sum('total')).get('total__sum'), 2)
        ctx['totaltva']=round(bons.aggregate(Sum('tva')).get('tva__sum'), 2)
    return JsonResponse(ctx)
    # return JsonResponse({
    #     'html':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8'),
    #     'total':round(bons.aggregate(Sum('total')).get('total__sum'), 2),
    #     'totaltva':round(bons.aggregate(Sum('tva')).get('tva__sum'), 2),

    # })


def filterachatdate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Itemsbysupplier.objects.filter(date__range=[startdate, enddate])
    return JsonResponse({
        'html':render(request, 'achatlist.html', {'bons':bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total')).get('total__sum'), 2)
    })

def updateclientpassword(request):
    clientid=request.POST.get('clientid')
    password=request.POST.get('password')
    try:
        user=Client.objects.get(pk=clientid).user
        user.set_password(password)
        user.save()
        return JsonResponse({
            'success':True
        })
    except Exception as e:
        return JsonResponse({
            'success':False,
            'error':e
        })

def updatereppassword(request):
    repid=request.POST.get('repid')
    password=request.POST.get('password')
    try:
        user=Represent.objects.get(pk=repid).user
        print(user, 'user', password)
        user.set_password(password)
        user.save()
        # logout(user)

        return JsonResponse({
            'success':True
        })
    except Exception as e:
        return JsonResponse({
            'success':False,
            'error':e
        })



def filterreglbldate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=PaymentClientbl.objects.filter(date__range=[startdate, enddate])
    print(bons)
    return JsonResponse({
        'html':render(request, 'reglbllist.html', {'bons':bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('amount')).get('amount__sum'), 2),

    })


def filterreglfcdate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=PaymentClientfc.objects.filter(date__range=[startdate, enddate])
    return JsonResponse({
        'html':render(request, 'reglfclist.html', {'bons':bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('amount')).get('amount__sum'), 2),

    })



def sortupbl(request):
    bons=Bonlivraison.objects.all().order_by('date')
    return JsonResponse({
        'html':render(request, 'bllist.html', {'bons':bons}).content.decode('utf-8')
    })

def sortdownbl(request):
    bons=Bonlivraison.objects.all().order_by('-date')
    return JsonResponse({
        'html':render(request, 'bllist.html', {'bons':bons}).content.decode('utf-8')
    })

def sortupfc(request):
    bons=Facture.objects.all().order_by('date')
    return JsonResponse({
        'html':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8')
    })

def sortdownfc(request):
    bons=Facture.objects.all().order_by('-date')
    return JsonResponse({
        'html':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8')
    })

from django.db.models import Q
import pandas as pd

def excelclients(request):
    myfile = request.FILES['excelFile']

    # Read Excel safely
    df = pd.read_excel(myfile, dtype={'code': str})
    df = df.fillna('')

    # Normalize data
    df['code'] = df['code'].astype(str).str.strip().str.zfill(6)
    df['name'] = df['name'].astype(str).str.strip()

    # Fetch existing clients ONCE
    

    clients_to_create = []

    for d in df.itertuples(index=False):
        
        try:
            region = d.region.lower().strip()
        except Exception:
            region = d.region

        address = None if pd.isna(d.address) else d.address
        ice = None if pd.isna(d.ice) else d.ice
        city = None if pd.isna(d.city) else d.city
        phone = None if pd.isna(d.phone) else d.phone
        clientname = None if pd.isna(d.clientname) else d.clientname

        clients_to_create.append(
            Client(
                represent_id=d.rep,
                code=d.code,
                name=d.name,
                city=city,
                ice=ice,
                region=region,
                phone=phone,
                address=address,
                clientname=clientname,
            )
        )

    # 🚀 Bulk insert
    Client.objects.bulk_create(clients_to_create, batch_size=500)

    return JsonResponse({
        'success': True,
    })
def excelpdcts(request):
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    
    entries=0
    for d in df.itertuples():
        try:
            ref = d.ref.lower().strip()
        except:
            ref=d.ref
        #reps=json.dumps(d.rep)
        name = d.name
        mark = None if pd.isna(d.mark) else d.mark
        order = None if pd.isna(d.order) else d.order
        minstock = 0 if pd.isna(d.minstock) else d.minstock
        refeq = '' if pd.isna(d.refeq) else d.refeq
        diam = '' if pd.isna(d.diam) else d.diam
        qty = 0 if pd.isna(d.qty) else d.qty
        buyprice = 0 if pd.isna(d.buyprice) else d.buyprice
        cars = None if pd.isna(d.cars) else d.cars
        sellprice = 0 if pd.isna(d.sellprice) else d.sellprice
        ctg = None if pd.isna(d.ctg) else d.ctg
        img = None if pd.isna(d.img) else d.img
        total=qty*buyprice
        #prixnet=0 if pd.isna(d.prixnet) else d.prixnet
        # try:
        #     print('entering', ref)
        #     product=Produit.objects.get(ref=ref)
        #     product.representremise=int(remise)
        #     product.save()
        #     entries+=1

        # except Exception as e:
            # print('>>',e, ref)
            # with open('error.txt', 'a') as ff:
            #     ff.write(f'>> {e} {ref}')
        lastid=0 # zero cause we will add one
        if Produit.objects.last():
            lastid=Produit.objects.last().id
        uniqcode=f'pdct{lastid+1}'
        product=Produit.objects.create(
            uniqcode=uniqcode,
            minstock=minstock,
            ref=ref,
            equivalent=refeq,
            refeq1=refeq,
            isactive=True,
            name=name,
            mark_id=mark,
            code=order,
            sellprice=sellprice,
            category_id=ctg,
            stocktotal=qty,
            #stockfacture=qty,
            diametre=diam,
            buyprice=buyprice,
            cars=cars
        )
        if img:
            product.image.name=img
            product.save()
        Stockin.objects.create(isinventaire=True, product=product, price=buyprice, quantity=qty, date=timezone.now(), total=total)

    print('>>>', entries)
    return JsonResponse({
        'success':True
    })


def deactivateaccount(request):
    userid=request.GET.get('userid')
    # user=User.objects.get(id=userid)
    # user.is_active=False
    # user.save()
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    if serverip:
        try:
            res=req.get(f'http://{serverip}/products/deactivateaccount', {
                'userid':userid,
            })
            res.raise_for_status()
        except req.exceptions.RequestException as e:
            print('Error deactivating account on server:', e)
            return JsonResponse({
                'success':False,
                'error':str(e)
            })
        
    # # delete user session in django session
    # UserSession.objects.filter(user=user).delete()
    # # Clear the user's session
    # #Session.objects.filter(session_key__in=UserSession.objects.filter(user=user).values('session_key')).delete()
    
    return JsonResponse({
        'success':True,
    })

def activateaccount(request):
    userid=request.GET.get('userid')
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    if serverip:
        try:
            res=req.get(f'http://{serverip}/products/activateaccount', {
                'userid':userid,
            })
            res.raise_for_status()
        except req.exceptions.RequestException as e:
            print('Error deactivating account on server:', e)
            return JsonResponse({
                'success':False,
                'error':str(e)
            })
        
    # # delete user session in django session
    # UserSession.objects.filter(user=user).delete()
    # # Clear the user's session
    # #Session.objects.filter(session_key__in=UserSession.objects.filter(user=user).values('session_key')).delete()

    return JsonResponse({
        'success':False,
        'error':'No server'
    })
    # user=User.objects.get(id=userid)
    # user.is_active=True
    # user.save()
    # req.get(f'http://{serverip}/products/activateaccount', {
    #     'username':user.username,
    # })
    # return JsonResponse({
    #     'success':True
    # })

def stock(request):
    categories=Category.objects.all()
    products=Produit.objects.all()[:50]
    facture=request.GET.get('fc')=='1'
    # if facture is true, this means we only want to get products in facture
    print('>> facture', facture)
    ctx={'categories':categories,
        'title':'Liste des Articles',
        'products':products,
        'stocktotal':Produit.objects.all().aggregate(Sum('stocktotal'))['stocktotal__sum']or 0,
        'stockfacture':Produit.objects.all().aggregate(Sum('stockfacture'))['stockfacture__sum']or 0,
        'totalstock':sum([i.totalofstock() for i in Produit.objects.all()])
    }
    if facture:
        return render(request, 'admin/fcproducts.html', ctx)
    return render(request, 'admin/products.html', ctx)




def getreglementsupp(request, id):
    reglement=PaymentSupplier.objects.get(pk=id)
    ctx={
        'title':'Reglement',
        'reglement':reglement,
    }
    bons=reglement.bons.all()
    # bons without bons in reglement
    livraisons=Itemsbysupplier.objects.filter(supplier=reglement.supplier).exclude(pk__in=[bon.pk for bon in bons])
    # livraisons=Itemsbysupplier.objects.filter(supplier=reglement.supplier)
    #we need bons to calculate total bl
    bonstocalculate=Itemsbysupplier.objects.filter(supplier=reglement.supplier)
    trs=''
    for i in livraisons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementssupp.exists() else ""}" class="loadblinupdatereglsupp" reglemntid="{id}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.nbon}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementssupp.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'date':reglement.date.strftime('%Y-%m-%d'),
        'echance':reglement.echeance.strftime('%Y-%m-%d') if reglement.echeance else '',
        'mode':reglement.mode,
        'npiece':reglement.npiece,
        'bons':list(bons.values()),
        'supplier':reglement.supplier.name,
        'mantant':reglement.amount,
        'livraisons':list(livraisons.values()),
        'total':round(bonstocalculate.aggregate(Sum('total'))['total__sum'], 2),
        'soldsupplier':round(reglement.supplier.rest, 2)
    })


def etatblclients(request):

    current_year = datetime.now().year
    # Create a date object for the first day of the current year
    first_day_of_year = date(current_year, 1, 1)
    now = datetime.now()
    # Get the last day of the current month
    last_day_of_month = calendar.monthrange(now.year, now.month)[1]
    # Create a date object for the last day of the current month
    last_day_of_month = date(now.year, now.month, last_day_of_month)
    print('>>>>>', first_day_of_year, last_day_of_month)
    start_date_str = request.GET.get('monthtostart',first_day_of_year.strftime('%Y-%m-%d'))
    end_date_str = request.GET.get('monthtoend', last_day_of_month.strftime('%Y-%m-%d'))

    # Parse dates
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    # Generate list of months between start_date and end_date
    months = []
    current = start_date
    while current <= end_date:
        months.append(current.strftime('%m/%y'))  # e.g., "January 2024"
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)
    print('>>>', months)
    clients = Client.objects.filter().order_by('city').exclude(diver=True)

    serialized_data = []
    #client=Client.objects.get(pk=3758)
    for clientindex, client in enumerate(clients):
        sitdata=0
        client_data = {'client_name': client.name, 'client_code': client.code, 'client_city': client.city, 'client_region': client.region, 'client_represent': client.represent, 'monthly_data': [], 'totalsituation': 0}

        # Filter data for the specified date range
        bons = Bonlivraison.objects.filter(
            client=client,
            date__range=[start_date, end_date],
            total__gt=0
        )

        avoirs = Avoirclient.objects.exclude(avoirfacture=True).filter(
            client=client,
            date__range=[start_date, end_date]
        )

        regls = PaymentClientbl.objects.filter(
            client=client,
            date__range=[start_date, end_date]
        )

        # Initialize monthly data
        monthly_data = {month: {'bons': 0, 'avoirs': 0, 'regls': 0, 'situation':0} for month in months}

        # Populate monthly data with bons, avoirs, and regls
        for bon in bons:
            month = bon.date.strftime('%m/%y')
            monthly_data[month]['bons'] += bon.total

        for avoir in avoirs:
            month = avoir.date.strftime('%m/%y')
            monthly_data[month]['avoirs'] += avoir.total

        for regl in regls:
            month = regl.date.strftime('%m/%y')
            monthly_data[month]['regls'] += regl.amount
        # Calculate the client situation for each month and aggregate totals
        total_bons = total_avoirs = total_regls = 0
        # this for testing specific clients
        # if client.id==3785:
        #     print('>>>>>>>>> bons', bons, 'avoirs', avoirs, 'regls', regls, )
        for monthindex, month in enumerate(months):
            total_bons += monthly_data[month]['bons']
            total_avoirs += monthly_data[month]['avoirs']
            total_regls += monthly_data[month]['regls']
            #monthly_data[month]['situation'] = round(monthly_data[month]['bons'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
            thismonthsit=round(monthly_data[month]['bons'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
            thisreg=monthly_data[month]['regls']
            if thisreg:
                for b in range(monthindex):
                    thisreg-=client_data['monthly_data'][b]['situation']
                    client_data['monthly_data'][b]['situation']=0
            thismonthsit=round(monthly_data[month]['bons'] - monthly_data[month]['avoirs'] - thisreg, 2)
            #     print('>>>>>>>>',client_data)
            #     #thismonthsit=-3000
            #     previousmonth=months[months.index(month)-1]
            #     dpreviousmonth=months[months.index(month)-2]
            #     treviousmonth=months[months.index(month)-3]
            #     diff=monthly_data[previousmonth]['situation']+thismonthsit
            #     if diff < 0:
            #         monthly_data[month]['situation'] = diff
            #         monthly_data[previousmonth]['situation']=0
            #         for month_data in client_data['monthly_data']:
            #             if month_data['month'] == previousmonth:
            #                 if month_data['situation'] == 0.00:
            #                     break


            #     else:
            #         monthly_data[previousmonth]['situation']=diff
            #         for month_data in client_data['monthly_data']:
            #             if month_data['month'] == previousmonth:
            #                 month_data['situation'] = diff
            #                 break
            #         monthly_data[month]['situation'] = round(monthly_data[month]['bons'] - monthly_data[month]['avoirs'], 2)
            # else:
            #     monthly_data[month]['situation']=thismonthsit


            client_data['monthly_data'].append({
                'month': month,
                'bons': monthly_data[month]['bons'],
                'avoirs': monthly_data[month]['avoirs'],
                'regls': monthly_data[month]['regls'],
                'situation': thismonthsit
            })

        # Calculate total situation for the client
        client_data['totalsituation'] = round(total_bons - total_avoirs - total_regls, 2)
        serialized_data.append(client_data)
        # Define start and end months for the date range


    #return render(request, 'etatblclients.html')
    return render(request, 'etatblclients.html', {'title': 'Etat bl client', 'data': serialized_data, 'months': months, 'monthtostart': start_date_str, 'monthtoend': end_date_str})
    # clients = Client.objects.filter(soldbl__gt=0).exclude(diver=True).order_by('city')
    # current_year = timezone.now().year
    # current_month = timezone.now().month
    # monthtostart=int(request.GET.get('monthtostart', 1))
    # monthtoend=int(request.GET.get('monthtoend', 5))
    # # Initialize serialized data
    # serialized_data = []
    # months=[i for i in range(monthtostart, monthtoend+1)]
    # for client in clients:
    #     client_data = {'client_name': client.name, 'client_city': client.city, 'client_region': client.region,  'client_represent': client.represent,'monthly_data': [], 'totalsituation': 0}

    #     # Filter data for the current year from January to the current month
    #     bons = Bonlivraison.objects.filter(
    #         client=client,
    #         date__year=current_year,
    #         date__month__range=[monthtostart, monthtoend],
    #         total__gt=0
    #     ).order_by('date')

    #     avoirs = Avoirclient.objects.filter(
    #         client=client,
    #         date__year=current_year,
    #         date__month__range=[monthtostart, monthtoend]
    #     ).order_by('date')

    #     regls = PaymentClientbl.objects.filter(
    #         client=client,
    #         date__year=current_year,
    #         date__month__range=[monthtostart, monthtoend]
    #     ).order_by('date')

    #     # Initialize monthly data for each month from January to current month
    #     monthly_data = {month: {'bons': 0, 'avoirs': 0, 'regls': 0} for month in range(monthtostart, monthtoend + 1)}
    #     # Populate monthly data with bons, avoirs, and regls
    #     for bon in bons:
    #         month = bon.date.month
    #         monthly_data[month]['bons'] += bon.total

    #     for avoir in avoirs:
    #         month = avoir.date.month
    #         monthly_data[month]['avoirs'] += avoir.total

    #     for regl in regls:
    #         month = regl.date.month
    #         monthly_data[month]['regls'] += regl.amount

    #     # Calculate the client situation for each month and aggregate totals
    #     total_bons = total_avoirs = total_regls = 0

    #     for month in range(monthtostart, monthtoend + 1):
    #         monthly_data[month]['situation'] = round(monthly_data[month]['bons'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
    #         total_bons += monthly_data[month]['bons']
    #         total_avoirs += monthly_data[month]['avoirs']
    #         total_regls += monthly_data[month]['regls']
    #         client_data['monthly_data'].append({
    #             'month': month,

    #             'situation': monthly_data[month]['situation']
    #         })

    #     # Calculate total situation for the client
    #     client_data['totalsituation'] = round(total_bons - total_avoirs - total_regls, 2)
    #     serialized_data.append(client_data)
    # # Return the serialized data as a JSON response
    # return render(request, 'etatblclients.html', {'data': serialized_data, 'months':months, 'monthtostart':monthtostart, 'monthtoend':monthtoend})


def etatfcclients(request):
    current_year = datetime.now().year
    # Create a date object for the first day of the current year
    first_day_of_year = date(current_year, 1, 1)
    now = datetime.now()
    # Get the last day of the current month
    last_day_of_month = calendar.monthrange(now.year, now.month)[1]
    # Create a date object for the last day of the current month
    last_day_of_month = date(now.year, now.month, last_day_of_month)
    start_date_str = request.GET.get('monthtostart',first_day_of_year.strftime('%Y-%m-%d'))
    end_date_str = request.GET.get('monthtoend', last_day_of_month.strftime('%Y-%m-%d'))
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    # Generate list of months between start_date and end_date
    months = []
    current = start_date
    while current <= end_date:
        months.append(current.strftime('%m/%y'))  # e.g., "January 2024"
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)

    clients = Client.objects.order_by('city').exclude(diver=True)

    serialized_data = []
    #client=Client.objects.get(pk=3417)
    for clientindex, client in enumerate(clients):
        sitdata=0
        client_data = {'client_name': client.name, 'client_code': client.code, 'client_city': client.city, 'client_region': client.region, 'client_represent': client.represent, 'monthly_data': [], 'totalsituation': 0}

        # Filter data for the specified date range
        factures = Facture.objects.filter(
            client=client,
            date__range=[start_date, end_date],
            total__gt=0
        )

        avoirs = Avoirclient.objects.filter(
            client=client,
            avoirfacture=True,
            date__range=[start_date, end_date]
        )

        regls = PaymentClientfc.objects.filter(
            client=client,
            date__range=[start_date, end_date]
        )

        # Initialize monthly data
        monthly_data = {month: {'factures': 0, 'avoirs': 0, 'regls': 0, 'situation':0} for month in months}

        # Populate monthly data with factures, avoirs, and regls
        for bon in factures:
            month = bon.date.strftime('%m/%y')
            monthly_data[month]['factures'] += bon.total
            sitdata+=bon.total

        for avoir in avoirs:
            month = avoir.date.strftime('%m/%y')
            monthly_data[month]['avoirs'] += avoir.total
            sitdata-=bon.total

        for regl in regls:
            month = regl.date.strftime('%m/%y')
            monthly_data[month]['regls'] += regl.amount
            sitdata-=bon.total
        # Calculate the client situation for each month and aggregate totals
        total_factures = total_avoirs = total_regls = 0

        for monthindex, month in enumerate(months):
            total_factures += monthly_data[month]['factures']
            total_avoirs += monthly_data[month]['avoirs']
            total_regls += monthly_data[month]['regls']
            #monthly_data[month]['situation'] = round(monthly_data[month]['factures'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
            thismonthsit=round(monthly_data[month]['factures'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
            thisreg=monthly_data[month]['regls']
            if thisreg:
                for b in range(monthindex):
                    thisreg-=client_data['monthly_data'][b]['situation']
                    client_data['monthly_data'][b]['situation']=0
            thismonthsit=round(monthly_data[month]['factures'] - monthly_data[month]['avoirs'] - thisreg, 2)



            client_data['monthly_data'].append({
                'month': month,
                'factures': monthly_data[month]['factures'],
                'avoirs': monthly_data[month]['avoirs'],
                'regls': monthly_data[month]['regls'],
                'situation': thismonthsit
            })

        # Calculate total situation for the client
        client_data['totalsituation'] = round(total_factures - total_avoirs - total_regls, 2)
        serialized_data.append(client_data)
        # Define start and end months for the date range


    #return render(request, 'etatblclients.html')
    return render(request, 'etatblclients.html', {'etatfacture':True ,'title': 'Etat fc client', 'data': serialized_data, 'months': months, 'monthtostart': start_date_str, 'monthtoend': end_date_str})

def updatebonavoir(request):
    id=request.POST.get('bonid')
    avoir=Avoirclient.objects.get(pk=id)
    client=Client.objects.get(pk=request.POST.get('clientid'))
    # we need avoir n° cause delete avoir will delete id, id is used in avoir n°
    avoirno=avoir.no
    avoiritems=Returned.objects.filter(avoir=avoir)
    totalbon=request.POST.get('totalbon')
    newmode=request.POST.get('mode')
    isfacture=True if newmode=='facture' else False
    print("isfacture", isfacture)
    thisclient=avoir.client
    # regle stock
    # # regle soldclient
    # if avoir.avoirfacture:
    #     thisclient.soldtotal=round(float(thisclient.soldtotal)+float(avoir.total), 2)
    #     thisclient.soldfacture=round(float(thisclient.soldfacture)+float(avoir.total), 2)
    #     thisclient.save()
    #     for i in avoiritems:
    #         i.product.stocktotal+=i.qty
    #         i.product.stockfacture+=i.qty
    #         i.product.save()
    #         i.delete()
    # else:
    #     thisclient.soldtotal=round(float(thisclient.soldtotal)+float(avoir.total), 2)
    #     thisclient.soldbl=round(float(thisclient.soldbl)+float(avoir.total), 2)
    #     thisclient.save()
    #     for i in avoiritems:
    #         i.product.stocktotal+=i.qty
    #         i.product.save()
    #         i.delete()

    # # delete old avoir
    # # create new avoir
    if avoir.client==client:
        thisclient.soldtotal=round(float(thisclient.soldtotal)+float(avoir.total)-float(totalbon), 2)
        if avoir.avoirfacture:
            thisclient.soldfacture=round(float(thisclient.soldfacture)+float(avoir.total), 2)
        else:
            thisclient.soldbl=round(float(thisclient.soldbl)+float(avoir.total), 2)
        if isfacture:
            thisclient.soldfacture=round(float(thisclient.soldfacture)-float(totalbon), 2)
        else:
            thisclient.soldbl=round(float(thisclient.soldbl)-float(totalbon), 2)

        thisclient.save()
    else:
        # not the same client
        thisclient.soldtotal=round(float(thisclient.soldtotal)+float(avoir.total), 2)
        # add sold to old client
        if avoir.avoirfacture:
            thisclient.soldfacture=round(float(thisclient.soldfacture)+float(avoir.total), 2)
        else:
            thisclient.soldbl=round(float(thisclient.soldbl)+float(avoir.total), 2)
        thisclient.save()
        # add sold to new client
        client.soldtotal=round(float(client.soldtotal)- float(totalbon), 2)
        if isfacture:
            client.soldfacture=round(float(client.soldfacture)- float(totalbon), 2)
        else:
            client.soldbl=round(float(client.soldbl)- float(totalbon), 2)
        client.save()
        print('new', client.soldtotal)
    items=Returned.objects.filter(avoir=avoir)
    for i in items:
        product=Produit.objects.get(pk=i.product_id)
        product.stocktotal=int(product.stocktotal)-int(i.qty)
        if avoir.avoirfacture:
            product.stockfacture=int(product.stockfacture)-int(i.qty)
        product.save()
        i.delete()
    avoir.client=client
    avoir.representant_id=request.POST.get('repid')
    avoir.total=totalbon
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    avoir.date=datebon
    avoir.no=request.POST.get('orderno')
    if isfacture:
        avoir.avoirfacture=True
        #client.soldbl=5
    else:
        avoir.avoirfacture=False
        #client.soldbl=5
    avoir.save()
    # update this items


    print('client:', avoir.client.id)
    with transaction.atomic():
        for i in json.loads(request.POST.get('products')):
            product=Produit.objects.get(pk=i['productid'])
            product.stocktotal=int(product.stocktotal)+int(i['qty'])
            if isfacture:
                product.stockfacture=int(product.stockfacture)+int(i['qty'])
            product.save()
            Returned.objects.create(
                avoir=avoir,
                product=product,
                qty=i['qty'],
                source=i['source'],
                remise=i['remise'],
                price=i['price'],
                total=i['total'],
            )

    return JsonResponse({
        'success':True
    })


def updatebonavoirsupp(request):
    id=request.POST.get('bonid')
    avoir=Avoirsupplier.objects.get(pk=id)
    supplier=request.POST.get('supplierid')
    totalbon=request.POST.get('totalbon')
    newmode=request.POST.get('mode')
    isfacture=True if newmode=='facture' else False
    newsupplier=Supplier.objects.get(pk=supplier)
    thissupplier=avoir.supplier
    avoirno=avoir.no
    avoiritems=Returnedsupplier.objects.filter(avoir=avoir)
    print(">>isfacture", isfacture)
    thissupplier.rest=round(float(thissupplier.rest)+float(avoir.total), 2)

    if avoir.supplier==newsupplier:
        print('the same supplier')
        thissupplier.rest=round(float(thissupplier.rest)+float(avoir.total)-float(totalbon), 2)
        thissupplier.save()
    else:
        # not the same supplier
        print('not the same supplier')
        thissupplier.rest=round(float(thissupplier.rest)+float(avoir.total), 2)
        # add sold to old supplier
        thissupplier.save()
        # add sold to new supplier
        newsupplier.rest=round(float(newsupplier.rest)- float(totalbon), 2)
        newsupplier.save()
    items=Returnedsupplier.objects.filter(avoir=avoir)
    for i in items:
        product=Produit.objects.get(pk=i.product_id)
        print('>> product', product.ref)
        product.stocktotal=int(product.stocktotal)+int(i.qty)
        if avoir.avoirfacture:
            print('old avoir was fc')
            product.stockfacture=int(product.stockfacture)+int(i.qty)
        product.save()
        i.delete()
    avoir.supplier_id=supplier
    avoir.total=totalbon
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    avoir.date=datebon
    avoir.no=request.POST.get('orderno')
    if isfacture:
        avoir.avoirfacture=True
    else:
        avoir.avoirfacture=False
    avoir.save()


    with transaction.atomic():
        for i in json.loads(request.POST.get('products')):
            product=Produit.objects.get(pk=i['productid'])
            print('>> product from table', product.ref)
            product.stocktotal=int(product.stocktotal)-int(i['qty'])
            if isfacture:
                print('new avoir is fc')
                product.stockfacture=int(product.stockfacture)-int(i['qty'])
            product.save()
            Returnedsupplier.objects.create(
                avoir=avoir,
                product=product,
                qty=i['qty'],
                remise=0 if i['remise']=="" else i['remise'],
                price=i['price'],
                total=i['total'],
            )

    return JsonResponse({
        'success':True
    })


def notifyadmin(request):
    serverip= Setting.objects.only('serverip').first()
    serverip=serverip.serverip if serverip else None
    notification=Ordersnotif.objects.first()
    if notification:
        return JsonResponse({
            'length':notification.length,
        })
    if serverip:
        try:
            # get the number of commands in server not yet sent to local server
            res=req.get(f'http://{serverip}/products/getcommandnumber')
            length=json.loads(res.text)['length']
            if length!=0:
                Ordersnotif.objects.create(length=json.loads(res.text)['length'], orders=json.loads(res.text)['orders'])
                return JsonResponse({
                    'length':json.loads(res.text)['length'],
                    #'orders':json.loads(res.text)['orders']
                })
            res.raise_for_status()
        except req.exceptions.RequestException as e:
            print('Error notifying admin on server:', e)
            return JsonResponse({
                'length':0,
            })
    return JsonResponse({
        'length':0,
    })
    oldnotif=Ordersnotif.objects.filter(isread=True)
    oldnotif.delete()
    newnotif=Ordersnotif.objects.filter(isread=False)
    return JsonResponse({
        "length": newnotif.count(),
    })

def disablenotif(request):
    newnotif=Ordersnotif.objects.filter(isread=False)
    newnotif.update(isread=True)
    return JsonResponse({
        'success':True
    })

def listecheance(request):
    # get payments that are cheque or effet in mode
    reglbl=PaymentClientbl.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False, echance__lte=today).order_by('-refused').order_by('-echance')
    reglfc=PaymentClientfc.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False, echance__lte=today).order_by('-refused').order_by('-echance')
    echeance = chain(*[
    ((rbl, 'bl') for rbl in reglbl),
    ((rfc, 'fc') for rfc in reglfc),
    ])
    totalbl=reglbl.aggregate(Sum('amount'))['amount__sum'] or 0
    totalfc=reglfc.aggregate(Sum('amount'))['amount__sum'] or 0
    total=round(totalbl+totalfc, 2)
    # Sort the items by date
    sorted_echeance = sorted(echeance, key=lambda item: item[0].refused, reverse=True)
    ctx={
        'title':'List des echeances Actuel',
        'echeances':sorted_echeance,
        'total':total
    }
    return render(request, 'listecheance.html', ctx)


def echeancetoday(request):
    reglbl=PaymentClientbl.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False,echance__lte=today).count()
    reglfc=PaymentClientfc.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False,echance__lte=today).count()
    print('>>>>>',PaymentClientbl.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False,echance=today), PaymentClientfc.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False,echance=today), today)
    return JsonResponse({
        'length':reglbl+reglfc
    })
def tabs(request):
    return render(request, 'tabs.html')


def getconnectedusers(request):
    # more than 5 minutes means the user is not connected
    five_minutes_ago = timezone.now() - timedelta(minutes=10)
    notconnected=Connectedusers.objects.filter(lasttime__lt=five_minutes_ago)
    connected=Connectedusers.objects.filter(lasttime__gt=five_minutes_ago)
    length=connected.count()
    trs=''
    for i in connected:
        trs+=f"""
        <tr>
        <td>{i.user.username}</td>
        <td>{i.user.groups.all()[0].name}</td>
        <td>{i.activity}</td>
        <td>{(i.lasttime).strftime('%Y-%m-%d %H:%M:%S')}</td>
        </tr>
        """
    print('connected', connected)
    print('notconnected', notconnected)
    return JsonResponse({
        'length':length,
        'trs':trs
    })


def payreglbl(request):
    reglid=request.GET.get('reglid')
    regl=PaymentClientbl.objects.get(pk=reglid)
    regl.ispaid=True
    regl.save()
    return JsonResponse({
        'success':True
    })

def payreglfc(request):
    reglid=request.GET.get('reglid')
    regl=PaymentClientfc.objects.get(pk=reglid)
    print(reglid, regl)
    regl.ispaid=True
    regl.save()
    return JsonResponse({
        'success':True
    })


def boncommandes(request):
    length=Order.objects.filter(isdelivered=False).exclude(note__icontains='Reliquat').count(),
    return JsonResponse({
        'length':length
    })


def listeconnected(request):
    five_minutes_ago = timezone.now() - timedelta(minutes=10)
    serverip= Setting.objects.only('serverip').first()
    serverip=serverip.serverip if serverip else None
    connectedserver=[]
    lastactiveserver=[]
    if serverip:
        res=req.get(f'http://{serverip}/products/listeconnected')
        connectedserver=json.loads(res.text)['connected'],
        lastactiveserver=json.loads(res.text)['active']
        # notconnected=Connectedusers.objects.filter(lasttime__lt=five_minutes_ago).order_by('-lasttime')
        # connected=Connectedusers.objects.filter(lasttime__gt=five_minutes_ago)

    ctx={
        'title':'List utilisateurs Active',
        # 'connected':connected,
        # 'active':notconnected,
        'connectedserver':json.loads(res.text)['connected'],
        'lastactiveserver':json.loads(res.text)['active']
    }
    return render(request, 'listconnected.html', ctx) 


def promotionspage(request):
    ctx={
        'promotions':Promotion.objects.all(),
        'title':'List des promotions'
    }
    return render(request, 'promotions.html', ctx)

def createpromotion(request):
    name=request.POST.get('promotionname')
    # get image file
    image=request.FILES.get('promotionimage')
    # create category
    pr=Promotion.objects.create(info=name, image=image)
    req.get(f'http://{serverip}/products/createpromotion', {
        'name':name,
        # get image file
        'image':pr.image.url.replace('/media/', '') if pr.image else ''
    })
    ctx={
        'promotions':Promotion.objects.all(),
        'title':'List des promotions'
    }
    return JsonResponse({
        'html':render(request, 'promotions.html', ctx).content.decode('utf-8')
    })

def updatepromotion(request):
    image=request.FILES.get('image') or None
    id=request.POST.get('id')
    promotion=Promotion.objects.get(pk=id)
    promotion.info=request.POST.get('name')
    if image:
        promotion.image=image
    promotion.save()
    req.get(f'http://{serverip}/products/updatepromotion', {
        'name':request.POST.get('name'),
        'id':id,
        # get image file
        'image':promotion.image.url.replace('/media/', '') if promotion.image else ''
    })
    ctx={
        'promotions':Promotion.objects.all(),
        'title':'List des promotions'
    }
    return JsonResponse({
        'html':render(request, 'promotions.html', ctx).content.decode('utf-8')
    })



def searchproductsforstock(request):
    term=request.GET.get('term')
    facture=request.GET.get('facture')=='1'
    print('>> facture',term, request.GET.get('facture') )
    if(term==''):
        products=Produit.objects.all()[:50]
        trs=''
        for i in products:
            trs+=f'''
        <tr ondblclick="createtab('addpdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
            style="background:{'#f3d6d694;' if not i.isactive else '' }"
                data-product-id="{ i.id }" class="product-row">
                  <td style="padding: 5px; font-weight: bold;" >
                      {i.ref.upper()}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.name}
                  </td>

                  <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
                      {i.buyprice if i.buyprice else 0}
                  </td>
                  <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
                      {i.sellprice}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center">
                      {i.remise}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center">
                      {i.prixnet}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
                      {i.stocktotal}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
                    <span class="stockfacture invisible">{i.stockfacture}</span>
                </td>

                  <td style="padding: 5px; font-weight: bold;">
                    {i.diametre}
                </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-success">
                    {i.block}
                </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[0] if i.getequivalent() else ''}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[1] if i.getequivalent() else ''}

                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[2] if i.getequivalent() else ''}

                  </td>

                    <td style="padding: 5px; font-weight: bold;">
                        {i.mark}
                    </td>
                    <td style="padding: 5px; font-weight: bold;">
                        {i.code}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-danger"><span class="percentage invisible"> {round(i.getpercentage(), 2)}</span></td>

              </tr>
        '''
        print('>> facturesection', facture)
        return JsonResponse({
            'trs':render(request, 'product_search.html', {'products':products, 'facture':facture, 'loadmore':True, 'facturesection':facture}).content.decode('utf-8'),
            'stocktotal':Produit.objects.all().aggregate(Sum('stocktotal'))['stocktotal__sum']or 0,
            'stockfacture':Produit.objects.all().aggregate(Sum('stockfacture'))['stockfacture__sum']or 0,
        })
    term = request.GET.get('term').lower()

    # Remove non-alphanumeric characters and convert to lowercase

    if not '+' in term:
        products=Produit.objects.filter(ref__istartswith=term).order_by('-stocktotal')

        q_objects = Q()
        q_objects &= (
            Q(ref__icontains=term) |
            Q(name__icontains=term) |
            Q(mark__name__icontains=term) |
            Q(category__name__icontains=term) |
            Q(equivalent__icontains=term) |
            Q(refeq1__icontains=term) |
            Q(refeq2__icontains=term) |
            Q(refeq3__icontains=term) |
            Q(refeq4__icontains=term)|
            Q(diametre__icontains=term)
        )
        # check if term in product.ref or product.name
        products=Produit.objects.filter(ref__istartswith=term).order_by('-stocktotal')

        q_objects = Q()
        q_objects &= (
            Q(ref__icontains=term) |
            Q(name__icontains=term) |
            Q(mark__name__icontains=term) |
            Q(category__name__icontains=term) |
            Q(equivalent__icontains=term) |
            Q(refeq1__icontains=term) |
            Q(refeq2__icontains=term) |
            Q(refeq3__icontains=term) |
            Q(refeq4__icontains=term) |
            Q(diametre__icontains=term)
        )
            # adding other products that have equivalent
        products=products | Produit.objects.filter(q_objects).order_by('-stocktotal')
    else:
        # Split the cleaned term into individual words separated by '*'
        search_terms = term.split('+')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:
            if term:

                # term = ''.join(char for char in term if char.isalnum())
                q_objects &= (Q(ref__icontains=term) | Q(coderef__icontains=term) | Q(name__icontains=term) | Q(category__name__icontains=term) |  Q(mark__name__icontains=term) |  Q(equivalent__icontains=term)  |  Q(refeq1__icontains=term) |  Q(refeq2__icontains=term)  |  Q(block__icontains=term) | Q(refeq3__icontains=term) | Q(refeq4__icontains=term) | Q(sellprice__icontains=term)  | Q(buyprice__icontains=term)  | Q(cars__icontains=term)  | Q(diametre__icontains=term))
        products=Produit.objects.filter(q_objects)[:50]
    trs=''
    for i in products:
        trs+=f'''
        <tr ondblclick="createtab('addpdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
              style="background:{'#f3d6d694;' if not i.isactive else '' }"
              class="product-row" data-product-id="{ i.id }"
              >
                  <td style="padding: 5px; font-weight: bold;" >
                      {i.ref.upper()}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.name}
                  </td>

                  <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
                      {i.buyprice if i.buyprice else 0}
                  </td>
                  <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
                      {i.sellprice}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center">
                      {i.remise}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center">
                      {i.prixnet}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
                      {i.stocktotal}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
                    <span class="stockfacture invisible">{i.stockfacture}</span>
                  </td>

                  <td style="padding: 5px; font-weight: bold;">
                    {i.diametre}
                </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-success">
                    {i.block}
                </td>
                  <td style="padding: 5px; font-weight: bold;">
                    {i.coderef}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[0] if i.getequivalent() else ''}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[1] if i.getequivalent() else ''}

                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[2] if i.getequivalent() else ''}

                  </td>
                <td style="padding: 5px; font-weight: bold;">
                      {i.mark}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.code}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-danger"><span class="percentage invisible"> {round(i.getpercentage(), 2)}</span></td>
        '''
    return JsonResponse({
        'trs':render(request, 'product_search.html', {'products':products, 'facture':facture, 'loadmore':True, 'facturesection':facture}).content.decode('utf-8'),
        #'stocktotal':Produit.objects.filter(q_objects).aggregate(Sum('stocktotal'))['stocktotal__sum']or 0,
        #'stockfacture':Produit.objects.filter(q_objects).aggregate(Sum('stockfacture'))['stockfacture__sum']or 0,
    })

# loading stock by 50, 50records at a time
# def loadstock(request):
#     page = int(request.GET.get('page', 1))
#     term = request.GET.get('term')
#     facture = request.GET.get('facture')=='1'
#     notactive = request.GET.get('notactive')
#     per_page = 50  # Adjust as needed
#     print('>> not active', notactive)
#     start = (page - 1) * per_page
#     end = page * per_page
#     print('>> fature', facture)
#     if term=='0':
#         print('>>>>>>>>>>>>', term=='0')
#         if notactive=='1':
#             print('from notactive')
#             products = Produit.objects.filter(isactive=False)[start:end]
#             trs=''
#             for i in products:
#                 trs+=f"""
#                     <tr ondblclick="createtab('addpdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
#                         style="background:{'#f3d6d694;' if not i.isactive else '' }"
#                             data-product-id="{ i.id }" class="product-row notactive">
#                             <td style="padding: 5px; font-weight: bold;" class="pe-2">
#                                 {i.ref.upper()}
#                             </td>
#                             <td style="padding: 5px; font-weight: bold;">
#                                 {i.name}
#                             </td>
#
#                             <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
#                                 {i.buyprice if i.buyprice else 0}
#                             </td>
#                             <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
#                                 {i.sellprice if i.sellprice else 0}
#                             </td>
#                             <td style="padding: 5px; font-weight: bold;" class="text-center">
#                                 {i.remise}
#                             </td>
#                             <td style="padding: 5px; font-weight: bold;" class="text-center">
#                                 {i.prixnet}
#                             </td>
#                             <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
#                                 {i.stocktotal}
#                             </td>
#                             <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
#                                 <span class="stockfacture invisible">{i.stockfacture}</span>
#                             </td>
#
#                             <td style="padding: 5px; font-weight: bold;">
#                                 {i.diametre}
#                             </td>
#                             <td style="padding: 5px; font-weight: bold;" class="text-success">
#                                 {i.block}
#                             </td>
#                             <td style="padding: 5px; font-weight: bold;">
#                                 {i.coderef}
#                             </td>
#                             <td style="padding: 5px; font-weight: bold;">
#                                 {i.getequivalent()[0] if i.getequivalent() else ''}
#                             </td>
#                             <td style="padding: 5px; font-weight: bold;">
#                                 {i.getequivalent()[1] if i.getequivalent() else ''}
#
#                             </td>
#                             <td style="padding: 5px; font-weight: bold;">
#                                 {i.getequivalent()[2] if i.getequivalent() else ''}
#
#                             </td>
#                             <td style="padding: 5px; font-weight: bold;">
#                                 {i.mark}
#                             </td>
#                              <td style="padding: 5px; font-weight: bold;"  class="text-danger">
#                                 {i.code}
#                             </td>
#
#                         </tr>
#                 """
#
#             return JsonResponse({
#                 'trs':trs,
#                 'has_more': len(products) == per_page
#             })
#         products = Produit.objects.all()[start:end]
#         trs=''
#         for i in products:
#             trs+=f'''
#             <tr ondblclick="createtab('addpdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
#                 style="background:{'#f3d6d694;' if not i.isactive else '' }"
#                     data-product-id="{ i.id }" class="product-row">
#                     <td style="padding: 5px; font-weight: bold;" class="pe-2">
#                         {i.ref.upper()}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;">
#                         {i.name}
#                     </td>
#
#                     <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
#                         {i.buyprice if i.buyprice else 0}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
#                         {i.sellprice}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;" class="text-center">
#                         {i.remise}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;" class="text-center">
#                         {i.prixnet}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
#                         {i.stocktotal}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
#                         <span class="stockfacture invisible">{i.stockfacture}</span>
#                     </td>
#
#                     <td style="padding: 5px; font-weight: bold;">
#                         {i.diametre}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;" class="text-success">
#                         {i.block}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;">
#                         {i.coderef}
#                     </td>
#
#                   <td style="padding: 5px; font-weight: bold;">
#                       {i.getequivalent()[0] if i.getequivalent() else ''}
#                   </td>
#                   <td style="padding: 5px; font-weight: bold;">
#                       {i.getequivalent()[1] if i.getequivalent() else ''}
#
#                   </td>
#                   <td style="padding: 5px; font-weight: bold;">
#                       {i.getequivalent()[2] if i.getequivalent() else ''}
#
#                   </td>
#
#                     <td style="padding: 5px; font-weight: bold;">
#                         {i.mark}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;"  class="text-danger">
#                         {i.code}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;" class="text-danger"><span class="percentage invisible"> {round(i.getpercentage(), 2)}</span></td>
#                 </tr>
#             '''
#         return JsonResponse({
#             'trs':trs,
#             'has_more': len(products) == per_page
#         })
#     else:
#         print('>>>>>>>>>>>>', term=='0')
#
#         search_terms = term.split('+')
#         print(search_terms)
#
#         # Create a list of Q objects for each search term and combine them with &
#         q_objects = Q()
#         for term in search_terms:
#             q_objects &= (Q(ref__iregex=term) | Q(name__iregex=term) | Q(category__name__iregex=term) |  Q(mark__name__iregex=term))
#         products=Produit.objects.filter(q_objects)[start:end]
#         trs=''
#         for i in products:
#             trs+=f'''
#             <tr ondblclick="createtab('addpdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
#                 style="background:{'#f3d6d694;' if not i.isactive else '' }"
#                     data-product-id="{ i.id }" class="product-row ">
#                     <td style="padding: 5px; font-weight: bold;" >
#                         {i.ref.upper()}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;">
#                         {i.name}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
#                         {i.buyprice if i.buyprice else 0}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
#                         {i.sellprice}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;" class="text-center">
#                         {i.remise}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;" class="text-center">
#                         {i.prixnet}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
#                         {i.stocktotal}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
#                         <span class="stockfacture invisible">{i.stockfacture}</span>
#                     </td>
#
#                     <td style="padding: 5px; font-weight: bold;">
#                         {i.diametre}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;" class="text-success">
#                         {i.block}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;">
#                         {i.coderef}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;">
#                       {i.getequivalent()[0] if i.getequivalent() else ''}
#                   </td>
#                   <td style="padding: 5px; font-weight: bold;">
#                       {i.getequivalent()[1] if i.getequivalent() else ''}
#
#                   </td>
#                   <td style="padding: 5px; font-weight: bold;">
#                       {i.getequivalent()[2] if i.getequivalent() else ''}
#
#                   </td>
#
#                     <td style="padding: 5px; font-weight: bold;">
#                         {i.mark}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;"  class="text-danger">
#                         {i.code}
#                     </td>
#                     <td style="padding: 5px; font-weight: bold;" class="text-danger"><span class="percentage invisible"> {round(i.getpercentage(), 2)}</span></td>
#                 </tr>
#             '''
#         return JsonResponse({
#             'trs':trs,
#             'has_more': len(products) == per_page
#         })
#
def loadstock(request):
    page = int(request.GET.get('page', 1))
    term = request.GET.get('term')
    facture = request.GET.get('facture')=='1'
    notactive = request.GET.get('notactive')
    per_page = 50  # Adjust as needed
    print('>> not active', notactive)
    start = (page - 1) * per_page
    end = page * per_page
    print('>> fature', facture)
    if term=='0':
        print('>>>>>>>>>>>>', term=='0')
        if notactive=='1':
            print('from notactive')
            products = Produit.objects.filter(isactive=False)[start:end]
            trs=''
            for i in products:
                trs+=f"""
                    <tr ondblclick="createtab('addpdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
                        style="background:{'#f3d6d694;' if not i.isactive else '' }"
                            data-product-id="{ i.id }" class="product-row notactive">
                            <td style="padding: 5px; font-weight: bold;" class="pe-2">
                                {i.ref.upper()}
                            </td>
                            <td style="padding: 5px; font-weight: bold;">
                                {i.name}
                            </td>

                            <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
                                {i.buyprice if i.buyprice else 0}
                            </td>
                            <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
                                {i.sellprice if i.sellprice else 0}
                            </td>
                            <td style="padding: 5px; font-weight: bold;" class="text-center">
                                {i.remise}
                            </td>
                            <td style="padding: 5px; font-weight: bold;" class="text-center">
                                {i.prixnet}
                            </td>
                            <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
                                {i.stocktotal}
                            </td>
                            <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
                                <span class="stockfacture invisible">{i.stockfacture}</span>
                            </td>

                            <td style="padding: 5px; font-weight: bold;">
                                {i.diametre}
                            </td>
                            <td style="padding: 5px; font-weight: bold;" class="text-success">
                                {i.block}
                            </td>
                            <td style="padding: 5px; font-weight: bold;">
                                {i.coderef}
                            </td>
                            <td style="padding: 5px; font-weight: bold;">
                                {i.getequivalent()[0] if i.getequivalent() else ''}
                            </td>
                            <td style="padding: 5px; font-weight: bold;">
                                {i.getequivalent()[1] if i.getequivalent() else ''}

                            </td>
                            <td style="padding: 5px; font-weight: bold;">
                                {i.getequivalent()[2] if i.getequivalent() else ''}

                            </td>
                            <td style="padding: 5px; font-weight: bold;">
                                {i.mark}
                            </td>
                             <td style="padding: 5px; font-weight: bold;"  class="text-danger">
                                {i.code}
                            </td>

                        </tr>
                """

            return JsonResponse({
                'trs':render(request, 'product_search.html', {'products':products, 'facture':facture, 'loadmore':True}).content.decode('utf-8'),
                'has_more': len(products) == per_page
            })
        products = Produit.objects.all()[start:end]
        trs=''
        for i in products:
            trs+=f'''
            <tr ondblclick="createtab('addpdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
                style="background:{'#f3d6d694;' if not i.isactive else '' }"
                    data-product-id="{ i.id }" class="product-row">
                    <td style="padding: 5px; font-weight: bold;" class="pe-2">
                        {i.ref.upper()}
                    </td>
                    <td style="padding: 5px; font-weight: bold;">
                        {i.name}
                    </td>

                    <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
                        {i.buyprice if i.buyprice else 0}
                    </td>
                    <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
                        {i.sellprice}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center">
                        {i.remise}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center">
                        {i.prixnet}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
                        {i.stocktotal}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
                        <span class="stockfacture invisible">{i.stockfacture}</span>
                    </td>

                    <td style="padding: 5px; font-weight: bold;">
                        {i.diametre}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-success">
                        {i.block}
                    </td>
                    <td style="padding: 5px; font-weight: bold;">
                        {i.coderef}
                    </td>

                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[0] if i.getequivalent() else ''}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[1] if i.getequivalent() else ''}

                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[2] if i.getequivalent() else ''}

                  </td>

                    <td style="padding: 5px; font-weight: bold;">
                        {i.mark}
                    </td>
                    <td style="padding: 5px; font-weight: bold;"  class="text-danger">
                        {i.code}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-danger"><span class="percentage invisible"> {round(i.getpercentage(), 2)}</span></td>
                </tr>
            '''
        return JsonResponse({
            'trs':render(request, 'product_search.html', {'products':products, 'facture':facture, 'loadmore':True}).content.decode('utf-8'),
            'has_more': len(products) == per_page
        })
    else:
        print('>>>>>>>>>>>>', term=='0')

        search_terms = term.split('+')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:
            q_objects &= (Q(ref__iregex=term) | Q(name__iregex=term) | Q(category__name__iregex=term) |  Q(mark__name__iregex=term))
        products=Produit.objects.filter(q_objects)[start:end]
        trs=''
        for i in products:
            trs+=f'''
            <tr ondblclick="createtab('addpdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
                style="background:{'#f3d6d694;' if not i.isactive else '' }"
                    data-product-id="{ i.id }" class="product-row ">
                    <td style="padding: 5px; font-weight: bold;" >
                        {i.ref.upper()}
                    </td>
                    <td style="padding: 5px; font-weight: bold;">
                        {i.name}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
                        {i.buyprice if i.buyprice else 0}
                    </td>
                    <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
                        {i.sellprice}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center">
                        {i.remise}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center">
                        {i.prixnet}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
                        {i.stocktotal}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
                        <span class="stockfacture invisible">{i.stockfacture}</span>
                    </td>

                    <td style="padding: 5px; font-weight: bold;">
                        {i.diametre}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-success">
                        {i.block}
                    </td>
                    <td style="padding: 5px; font-weight: bold;">
                        {i.coderef}
                    </td>
                    <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[0] if i.getequivalent() else ''}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[1] if i.getequivalent() else ''}

                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[2] if i.getequivalent() else ''}

                  </td>

                    <td style="padding: 5px; font-weight: bold;">
                        {i.mark}
                    </td>
                    <td style="padding: 5px; font-weight: bold;"  class="text-danger">
                        {i.code}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-danger"><span class="percentage invisible"> {round(i.getpercentage(), 2)}</span></td>
                </tr>
            '''
        return JsonResponse({
            'trs':render(request, 'product_search.html', {'products':products, 'facture':facture, 'loadmore':True}).content.decode('utf-8'),
            'has_more': len(products) == per_page
        })

def loadlistachat(request):
    thisyear=timezone.now().year
    page = int(request.GET.get('page', 1))
    year =request.GET.get('year')
    startdate =request.GET.get('startdate')
    enddate =request.GET.get('enddate')
    term =request.GET.get('term')
    per_page = 50  # Adjust as needed
    print(term, year, startdate, enddate)
    trs=''
    start = (page - 1) * per_page
    end = page * per_page
    if term != '0':
        print('>>term', term)
        # Split the term into individual words separated by '*'

        print('>>>>>>>>> here bl')
        # Split the term into individual words separated by '*'
        search_terms = term.split('+')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:


            q_objects &= (Q(supplier__name__iregex=term) |
                Q(nbon__iregex=term) |
                Q(total__iregex=term)
                )
        print(startdate, enddate)
        if startdate=='0' and enddate=='0':
            bons=Itemsbysupplier.objects.filter(q_objects).filter(date__year=thisyear).order_by('-date')[start:end]
            total=round(Bonlivraison.objects.filter(q_objects).filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
        else:
            bons=Itemsbysupplier.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-date')[start:end]
            total=round(Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'] or 0, 2)
    # if startdate != '0' and enddate != '0':
    #     startdate = datetime.strptime(startdate, '%Y-%m-%d')
    #     enddate = datetime.strptime(enddate, '%Y-%m-%d')
    #     bons=Itemsbysupplier.objects.filter(date__range=[startdate, enddate]).order_by('-date')[start:end]
    #     total=round(Itemsbysupplier.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'] or 0, 2)
    if year=="0":
        bons= Itemsbysupplier.objects.filter(date__year=thisyear).order_by('-date')[start:end]
        total=round(Itemsbysupplier.objects.filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
    else:
        bons= Itemsbysupplier.objects.filter(date__year=year).order_by('-date')[start:end]
        total=round(Itemsbysupplier.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    for order in bons:
        trs+=f'''
        <tr class="ord achat-row" orderid="{order.id}" ondblclick="createtab('bonachat{order.id}', 'Bon achat {order.nbon}', '/products/bonachatdetails/{order.id}')">
            <td>{ order.nbon }</td>
            <td>{ order.date.strftime("%d/%m/%Y") }</td>
            <td>{ order.supplier.name }</td>
            <td>{ order.total}</td>
            <td>{ order.tva}</td>
            <td>{"Facture"if order.isfacture else "Bl"}</td>
            <td>{ round(order.supplier.rest, 2) }</td>
            <td class="d-flex">
                <div>{"R0"if order.ispaid else "N1"}</div>

              <div style="width:15px; height:15px; border-radius:50%; background:{"green"if order.ispaid else "red"};" ></div>



            </td>
            <td>
              <button class="btn bi bi-download" onclick="printbonachat('{order.id}')"></button>
            </td>

          </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })

def loadlistbl(request):
    thisyear=timezone.now().year
    page = int(request.GET.get('page', 1))
    year =request.GET.get('year')
    facture =request.GET.get('facture')=='1'
    startdate =request.GET.get('startdate')
    enddate =request.GET.get('enddate')
    term =request.GET.get('term')
    per_page = 50  # Adjust as needed
    print(term, year, startdate, enddate)
    trs=''
    start = (page - 1) * per_page
    end = page * per_page
    if term != '0':
        print('>>term', term)
        # Split the term into individual words separated by '*'

        print('>>>>>>>>> here bl')
        # Split the term into individual words separated by '*'
        search_terms = term.split('+')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:

            # if '-' in term:
            #     date_range = term.split('-')
            #     start_date = datetime.strptime(date_range[0].strip(), '%d/%m/%Y')
            #     end_date = datetime.strptime(date_range[1].strip(), '%d/%m/%Y')
            #     q_objects &= (Q(client__name__iregex=term) |
            #         Q(salseman__name__iregex=term) |
            #         Q(bon_no__iregex=term) |
            #         Q(client__region__iregex=term)|
            #         Q(client__city__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(total__iregex=term)|
            #         Q(date__range=[start_date, end_date])
            #     )
            # else:
            #     q_objects &= (Q(client__name__iregex=term) |
            #         Q(salseman__name__iregex=term) |
            #         Q(bon_no__iregex=term) |
            #         Q(client__region__iregex=term)|
            #         Q(client__city__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(total__iregex=term)
            #     )
            q_objects &= (Q(client__name__iregex=term) |
                    Q(salseman__name__iregex=term) |
                Q(bon_no__iregex=term) |
                Q(client__region__iregex=term)|
                Q(client__city__iregex=term)|
                Q(client__code__iregex=term)|
                Q(total__iregex=term)|
                Q(commande__order_no__iregex=term)|
                Q(statusreg__iregex=term)|
                Q(note__iregex=term)
                )
        print(startdate, enddate)
        if startdate=='0' and enddate=='0':
            bons=Bonlivraison.objects.filter(q_objects).filter(date__year=year).order_by('-bon_no')[start:end]
            if facture:
                bons=Bonlivraison.objects.filter(q_objects).filter(date__year=year, isfacture=True).exclude(total__gt=0).order_by('-bon_no')[start:end]
            total=round(Bonlivraison.objects.filter(q_objects).filter(date__year=year).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        else:
            bons=Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-bon_no')[start:end]
            if facture:
                bons=Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfacture=True).exclude(total__gt=0).order_by('-bon_no')[start:end]
            total=round(Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        for i in bons:
            trs+=f'''
            <tr
            style="background: {"lightgreen;" if i.isdelivered else ""} color:{"blue" if i.isfacture else ""} "
            class="ord {"text-danger" if i.ispaid else ''} bl-row"
            year={year}
            orderid="{i.id}"
            ondblclick="createtab('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')"
            term="{term}">
                <td>{ i.bon_no }</td>
                    <td>{ i.date.strftime("%d/%m/%Y")}</td>
                    <td>{ i.client.name }</td>
                    <td>{ i.client.code }</td>
                    <td style="color: blue;">{ i.total}</td>
                    <td>{ i.client.region}</td>
                    <td>{ i.client.city}</td>
                    <td>{ "%.2f" % i.client.soldbl} </td>
                    <td>{ i.salseman }</td>
                    <td class="d-flex justify-content-between">
                    <div>
                    {'R0' if i.ispaid else 'N1' }

                    </div>
                    <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                    </td>
                    <td>
                    {'OUI' if i.isfacture else 'NON'}
                    </td>

                    <td>
                        {i.commande.order_no if i.commande else "---"}
                    </td>
                    <td>
                    {i.modlvrsn}
                    </td>
                    <td>
                    {i.note}
                    </td>
                    <td class="d-flex">
                  <button class="btn btn-sm btn-info" onclick="makedelivered('{i.id}', event)"></button>
                <button class="btn btn-sm bi bi-download" onclick="printlivraison('{i.id}')"></button>
                    </td>

              </tr>
            '''
        print('>>>load bl term')
        return JsonResponse({
            'trs':trs,
            'has_more': len(bons) == per_page
        })
    if startdate != '0' and enddate != '0':
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        bons=Bonlivraison.objects.filter(date__range=[startdate, enddate]).order_by('-bon_no')[start:end]
        if facture:
            bons=Bonlivraison.objects.filter(date__range=[startdate, enddate], isfacture=True).exclude(total__gt=0).order_by('-bon_no')[start:end]
        total=round(Bonlivraison.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'] or 0, 2)
        for i in bons:
            trs+=f'''
            <tr
            style="background: {"lightgreen;" if i.isdelivered else ""} color:{"blue" if i.isfacture else ""} "
            class="ord {"text-danger" if i.ispaid else ''} bl-row"
            year={year} orderid="{i.id}"
            ondblclick="createtab('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')"
            startdate={startdate} enddate={enddate}>
                <td>{ i.bon_no }</td>
                    <td>{ i.date.strftime("%d/%m/%Y")}</td>
                    <td>{ i.client.name }</td>
                    <td>{ i.client.code }</td>
                    <td style="color: blue;">{ i.total}</td>
                    <td>{ i.client.region}</td>
                    <td>{ i.client.city}</td>
                    <td>{ "%.2f" % i.client.soldbl }</td>
                    <td>{ i.salseman }</td>
                    <td class="d-flex justify-content-between">
                    <div>
                    {'R0' if i.ispaid else 'N1' }

                    </div>
                    <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                    </td>
                    <td>
                    {'OUI' if i.isfacture else 'NON'}

                    </td>

                    <td>

                    </td>
                    <td>
                    {i.modlvrsn}
                    </td>
                    <td>
                    {i.note}
                    </td>
                    <td class="d-flex">
                  <button class="btn btn-sm btn-info" onclick="makedelivered('{i.id}', event)"></button>
                <button class="btn btn-sm bi bi-download" onclick="printlivraison('{i.id}')"></button>
                    </td>

              </tr>
            '''
        print('>>>load bl date f')
        return JsonResponse({
            'trs':trs,
            'has_more': len(bons) == per_page
        })
    if year=="0":
        bons= Bonlivraison.objects.filter(date__year=thisyear).order_by('-bon_no')[start:end]
        if facture:
            bons= Bonlivraison.objects.filter(date__year=thisyear, isfacture=True).exclude(total__gt=0).order_by('-bon_no')[start:end]
        total=round(Bonlivraison.objects.filter(date__year=thisyear).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
    else:
        bons= Bonlivraison.objects.filter(date__year=year).order_by('-bon_no')[start:end]
        if facture:
            bons= Bonlivraison.objects.filter(date__year=year, isfacture=True).exclude(total__gt=0).order_by('-bon_no')[start:end]
        total=round(Bonlivraison.objects.filter(date__year=year).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)

    for i in bons:
        trs+=f'''
        <tr style="background: {"lightgreen;" if i.isdelivered else ""} color:{"blue" if i.isfacture else ""} " class="ord {"text-danger" if i.ispaid else ''} bl-row" year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')">
            <td>{ i.bon_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td style="color: blue;">{ i.total}</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ "%.2f" % i.client.soldbl}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td>
                {'OUI' if i.isfacture else 'NON'}

                </td>

                <td>

                </td>
                <td>
                {i.modlvrsn}
                </td>
                <td>
                {i.note}
                </td>
                <td class="d-flex">
                  <button class="btn btn-sm btn-info" onclick="makedelivered('{i.id}', event)"></button>
                <button class="btn btn-sm bi bi-download" onclick="printlivraison('{i.id}')"></button>

                </td>
          </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })


def loadlistbc(request):
    thisyear=timezone.now().year
    # each block needs a return statement
    page = int(request.GET.get('page', 1))
    year =request.GET.get('year')
    startdate =request.GET.get('startdate')
    enddate =request.GET.get('enddate')
    term =request.GET.get('term')
    print(year, startdate, enddate, term)
    per_page = 50  # Adjust as needed
    trs=''

    start = (page - 1) * per_page
    end = page * per_page
    if term != '0':
        print('>>>>> in term')

        # Split the term into individual words separated by '*'
        search_terms = term.split('+')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for i in search_terms:
            q_objects &= (Q(client__name__iregex=i) |
                    Q(salseman__name__iregex=i) |
                    Q(order_no__iregex=i) |
                    Q(client__region__iregex=i)|
                    Q(client__city__iregex=i)|
                    Q(client__code__iregex=i)|
                    Q(total__iregex=i)|
                    Q(note__iregex=i)
                )
        if startdate=='0' and enddate=='0':
            print("in  term  no filter data")
            bons=Order.objects.filter(q_objects).filter(date__year=year).order_by('-order_no')[start:end]
            total=round(Order.objects.filter(q_objects).filter(date__year=year).order_by('-order_no').aggregate(Sum('total'))['total__sum'] or 0, 2)

        else:
            print('in term and filterdate')
            bons=Order.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-order_no')[start:end]
            total=round(Order.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-order_no').aggregate(Sum('total'))['total__sum'] or 0, 2)

        return JsonResponse({
            'trs':render(request, 'bclist.html', {"bons":bons, 'loadmore':True, 'term':term, 'startdate':request.GET.get('startdate'), 'enddate':request.GET.get('enddate'), 'year':year}).content.decode('utf-8'),
            'has_more': len(bons) == per_page,
            'total':total
        })
    if startdate != '0' and enddate != '0':
        print('>>>>> in date fil')
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')+timedelta(1)
        bons=Order.objects.filter(date__range=[startdate, enddate]).order_by('-order_no')[start:end]
        total=round(Order.objects.filter(date__range=[startdate, enddate]).order_by('-order_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        return JsonResponse({
            'trs':render(request, 'bclist.html', {"bons":bons, 'loadmore':True, 'term':term, 'startdate':request.GET.get('startdate'), 'enddate':request.GET.get('enddate'), 'year':year}).content.decode('utf-8'),
            'has_more': len(bons) == per_page,
            'total':total
        })

    if year=="0":
        print('>>>>> in year list bc is null')

        bons= Order.objects.filter(date__year=thisyear).order_by('-order_no')[start:end]
        total=round(Order.objects.filter(date__year=thisyear).order_by('-order_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
    else:
        print('>>>>> in year')

        bons= Order.objects.filter(date__year=year).order_by('-order_no')[start:end]
        total=round(Order.objects.filter(date__year=year).order_by('-order_no').aggregate(Sum('total'))['total__sum'] or 0, 2)



    return JsonResponse({
        'trs':render(request, 'bclist.html', {"bons":bons, 'loadmore':True, 'term':term, 'startdate':request.GET.get('startdate'), 'enddate':request.GET.get('enddate'), 'year':year}).content.decode('utf-8'),
        'has_more': len(bons) == per_page,
        'total':total
    })


def searchforlistachat(request):
    thisyear=timezone.now().year
    term=request.GET.get('term')
    year=request.GET.get('year')
    facture=request.GET.get('facture')=='1'
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    search_terms = term.split('+')
    print('>> startdate==0', startdate=='0' and enddate=='0', 'year', year)
    print('>> term', search_terms)
    #and combine them with &
    q_objects = Q()
    for i in search_terms:
        q_objects &= (Q(supplier__name__iregex=i) |
            Q(nbon__iregex=i) |
            Q(total__iregex=i)
        )
    if startdate=='0' and enddate=='0':
        if facture:
            bons=Itemsbysupplier.objects.filter(q_objects).filter(date__year=year, isfacture=True).order_by('-date')
            total=round(Itemsbysupplier.objects.filter(q_objects).filter(date__year=year, isfacture=True).order_by('-date').aggregate(Sum('total'))['total__sum'] or 0, 2)
            print('>> no dates and facture bons', bons)
        else:

            bons=Itemsbysupplier.objects.filter(q_objects).filter(date__year=year).order_by('-date')
            total=round(Itemsbysupplier.objects.filter(q_objects).filter(date__year=year).order_by('-date').aggregate(Sum('total'))['total__sum'] or 0, 2)
            print('>> no dates and not facture bons', bons)
    else:
        if facture:
            bons=Itemsbysupplier.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfacture=True).order_by('-date')
            total=round(Itemsbysupplier.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfacture=True).order_by('-date').aggregate(Sum('total'))['total__sum'] or 0, 2)
            print('>> w dates and facture bons', bons)

        else:
            bons=Itemsbysupplier.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-date')
            total=round(Itemsbysupplier.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-date').aggregate(Sum('total'))['total__sum'] or 0, 2)
            print('>> w dates and no facture', bons)
    trs=''
    for order in bons:
        trs+=f'''
            <tr class="ord " orderid="{order.id}" ondblclick="createtab('bonachat{order.id}', 'Bon achat {order.nbon}', '/products/bonachatdetails/{order.id}')">
            <td>{ order.nbon }</td>
            <td>{ order.date.strftime("%d/%m/%Y") }</td>
            <td>{ order.supplier.name }</td>
            <td>{ order.total}</td>
            <td>{ order.tva}</td>
            <td>{"Facture"if order.isfacture else "Bl"}</td>
            <td>{ round(order.supplier.rest, 2) }</td>
            <td class="d-flex">
                <div>{"R0"if order.ispaid else "N1"}</div>
                <div style="width:15px; height:15px; border-radius:50%; background:{"green"if order.ispaid else "red"};" ></div>
            </td>
            <td>
              <button class="btn bi bi-download" onclick="printbonachat('{order.id}')"></button>
            </td>

          </tr>
            '''

    return JsonResponse({
        'trs':trs,
        'total':total
    })

def searchforlistbl(request):
    #thisyear=timezone.now().year
    term=request.GET.get('term')
    year=request.GET.get('year')
    facture=request.GET.get('facture')=='1'
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    # we dont need this
    if(term==''):

        bons=Bonlivraison.objects.filter(date__year=year).order_by('-bon_no')[:50]
        total=round(Bonlivraison.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''} bl-row" orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')">
                <td>{ i.bon_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.total}</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldbl}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td class="text-danger">

                </td>
                <td class="text-danger">
                {'OUI' if i.isfacture else 'NON'}

                </td>

                <td>


                </td>
                <td>
                {i.note}
                </td>
                <td>
                {i.modlvrsn}
                </td>
                <td class="d-flex">
                  <button class="btn btn-sm btn-info" onclick="makedelivered('{i.id}', event)"></button>
                <button class="btn btn-sm bi bi-download" onclick="printlivraison('{i.id}')"></button>
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':render(request, 'bllist.html', {'bons':bons}).content.decode('utf-8')
        })

    # Split the term into individual words separated by '*'
    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for i in search_terms:

            # if '-' in term:
            #     date_range = term.split('-')
            #     start_date = datetime.strptime(date_range[0].strip(), '%d/%m/%Y')
            #     end_date = datetime.strptime(date_range[1].strip(), '%d/%m/%Y')
            #     q_objects &= (Q(client__name__iregex=term) |
            #         Q(salseman__name__iregex=term) |
            #         Q(bon_no__iregex=term) |
            #         Q(client__region__iregex=term)|
            #         Q(client__city__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(total__iregex=term)|
            #         Q(date__range=[start_date, end_date])
            #     )
            # else:
            #     q_objects &= (Q(client__name__iregex=term) |
            #         Q(salseman__name__iregex=term) |
            #         Q(bon_no__iregex=term) |
            #         Q(client__region__iregex=term)|
            #         Q(client__city__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(total__iregex=term)
            #     )
        q_objects &= (Q(client__name__iregex=i) |
                Q(salseman__name__iregex=i) |
                Q(bon_no__iregex=i) |
                Q(client__region__iregex=i)|
                Q(client__city__iregex=i)|
                Q(client__code__iregex=i)|
                Q(total__iregex=i)|
                Q(commande__order_no__iregex=i)|
                Q(statusreg__iregex=i)|
                Q(note__iregex=i)

            )
    print(startdate, enddate)
    if startdate=='0' and enddate=='0':
        bons=Bonlivraison.objects.filter(q_objects).filter(date__year=year).order_by('-bon_no')[:50]
        if facture:
            bons=Bonlivraison.objects.filter(q_objects).filter(date__year=year, isfacture=True).order_by('-bon_no')[:50]
        total=round(Bonlivraison.objects.filter(q_objects).filter(date__year=year).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
    else:
        bons=Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-bon_no')[:50]
        if facture:
            bons=Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate], isfacture=True).order_by('-bon_no')[:50]
        total=round(Bonlivraison.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
    trs=''
    for i in bons:
        trs+=f'''
            <tr
            style="background: {"lightgreen;" if i.isdelivered else ""} color:{"blue" if i.isfacture else ""} "
            class="ord bl-row {"text-danger" if i.ispaid else ''}" term={term} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')"
            >
                <td>{ i.bon_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td style="color: blue;">{ i.total:.2f}</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldbl:.2f}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td>
                {'OUI' if i.isfacture else 'NON'}

                </td>

                <td>
                    {i.commande.order_no if i.commande else '--'}
                </td>
                <td>
                {i.note}
                </td>
                <td>
                {i.modlvrsn}
                </td>
                <td class="d-flex">
                    <button class="btn btn-sm btn-info" onclick="makedelivered('{i.id}', event)"></button>
                    <button class="btn btn-sm bi bi-download" onclick="printlivraison('{i.id}')"></button>
                </td>
            </tr>
            '''

    return JsonResponse({
        'trs':trs,
        'total':total
    })

def searchforlistbc(request):
    term=request.GET.get('term')
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    year=request.GET.get('year')
    # we dont need this
    if(term==''):

        bons=Order.objects.filter(date__year=year)[:50]
        total=round(Order.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''} bl-row" orderid="{i.id}" ondblclick="createtab('command{i.id}', 'Commande {i.order_no}', '/products/boncommandedetails/{i.id}')" term={term}>
                <td>{ i.order_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.total}</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldbl}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td class="text-danger">

                </td>
                <td class="text-danger">
                {'OUI' if i.isfacture else 'NON'}

                </td>

                <td>


                </td>
                <td>
                {i.note}
                </td>
                <td>
                {i.modlvrsn}
                </td>
                <td class="d-flex">
                  <button class="btn btn-sm btn-info" onclick="makedelivered('{i.id}', event)"></button>
                <button class="btn btn-sm bi bi-download" onclick="printlivraison('{i.id}')"></button>
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs
        })

    # Split the term into individual words separated by '*'
    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for i in search_terms:
        q_objects &= (Q(client__name__iregex=i) |
                Q(salseman__name__iregex=i) |
                Q(order_no__iregex=i) |
                Q(client__region__iregex=i)|
                Q(client__city__iregex=i)|
                Q(client__code__iregex=i)|
                Q(total__iregex=i)|
                Q(note__iregex=i)

            )
    print(startdate, enddate)
    if startdate=='0' and enddate=='0':
        bons=Order.objects.filter(q_objects).filter(date__year=year).order_by('-order_no')[:50]
        total=round(Order.objects.filter(q_objects).filter(date__year=year).order_by('-order_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
    else:
        bons=Order.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-order_no')[:50]
        total=round(Order.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-order_no').aggregate(Sum('total'))['total__sum'] or 0, 2)



    return JsonResponse({
        'trs':render(request, 'bclist.html', {'bons':bons,
        'loadmore':True,
        'startdate':startdate, 'enddate':enddate, 'term':term, 'year':year}).content.decode('utf-8'),
        'total':total
    })

def loadlistfc(request):
    facture=request.GET.get('facture')=='1'
    thisyear=timezone.now().year
    page = int(request.GET.get('page', 1))
    year =request.GET.get('year')
    startdate =request.GET.get('startdate')
    enddate =request.GET.get('enddate')
    term =request.GET.get('term')
    comptable =request.GET.get('comptable')
    print('>> page, year, term, startdate, enddate, comptable', page, year, term, startdate, enddate, comptable)
    per_page = 50  # Adjust as needed
    start = (page - 1) * per_page
    end = page * per_page
    print('>>>>>', start, end, page)
    print('>>>>> term', term)
    if comptable=='1':
        if year=='0':
            bons=Facture.objects.filter(date__year=thisyear, isaccount=True).exclude(client_id=3731 if facture else None).order_by('-facture_no')[start:end]
        else:
            bons=Facture.objects.filter(date__year=year, isaccount=True).exclude(client_id=3731 if facture else None).order_by('-facture_no')[start:end]

        trs=''
        for i in bons:
            trs+=f'''
                <tr class="ord {"text-danger" if i.ispaid else ''}
                 fc-row"
                    style="color:{"blue" if i.bon else ""} " comptable="1" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                    <td>{ i.facture_no }</td>
                    <td>{ i.date.strftime("%d/%m/%Y")}</td>
                    <td>{ i.total}</td>
                    <td>{ i.tva}</td>
                    <td>{ i.client.name }</td>
                    <td>{ i.client.code }</td>
                    <td>{ i.client.region}</td>
                    <td>{ i.client.city}</td>
                    <td>{ i.client.soldfacture}</td>
                    <td>{ i.salseman }</td>
                    <td class="d-flex justify-content-between">
                    <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
                    <button title="Facture Comptabilisé" class="btn border border-success border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
                    </td>
                    <td >
                        {i.note}
                    </td>

                    <td>
                    {i.bon.bon_no if i.bon else "--"}
                    </td>
                    <td class="d-flex">
                    <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                    <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
                    </td>
                </tr>
                '''
        return JsonResponse({
            'trs':render(request, 'fclist.html', {'bons':bons, 'loadmore':True, 'facturesection':facture}).content.decode('utf-8'),
            'has_more': len(bons) == per_page,
            'facturesection':facture
        })
    if term != '0':

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in term.split('+'):
            # print('>>>>>>>term ', term)
            # if '-' in term[0]:
            #     date_range = term.split('-')
            #     start_date = datetime.strptime(date_range[0].strip(), '%d/%m/%Y')
            #     end_date = datetime.strptime(date_range[1].strip(), '%d/%m/%Y')
            #     q_objects &= (
            #         Q(client__name__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(salseman__name__iregex=term)|
            #         Q(facture_no__iregex=term)|
            #         Q(bon__bon_no__iregex=term)|
            #         Q(client__region__iregex=term)|
            #         Q(client__city__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(total__iregex=term)|
            #         Q(date__range=[start_date, end_date])
            #         )
            # else:
            #     q_objects &= (
            #         Q(client__name__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(salseman__name__iregex=term)|
            #         Q(facture_no__iregex=term)|
            #         Q(bon__bon_no__iregex=term)|
            #         Q(client__region__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(total__iregex=term)
            #     )
            q_objects &= (
                Q(client__name__iregex=term)|
                Q(client__city__iregex=term)|
                Q(salseman__name__iregex=term)|
                Q(facture_no__iregex=term)|
                Q(bon__bon_no__iregex=term)|
                Q(client__region__iregex=term)|
                Q(client__code__iregex=term)|
                Q(note__iregex=term)|
                Q(statusreg__iregex=term)|
                Q(total__iregex=term)
            )

        if startdate=='0' and enddate=='0':
            print('>>>>>in term without dates term, year', term, year)
            bons=Facture.objects.filter(q_objects).filter(date__year=year).exclude(client_id=3731 if facture else None).order_by('-facture_no')[start:end]
            total=round(Facture.objects.filter(q_objects).filter(date__year=year).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
            totaltva=round(Facture.objects.filter(q_objects).filter(date__year=year).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
            trs=''
            for i in bons:
                trs+=f'''
                <tr class="ord {"text-danger" if i.ispaid else ''}
                 fc-row"
                    style="color:{"blue" if i.bon else ""} "
                  year="{year}" term="{term}" orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                    <td>{ i.facture_no }</td>
                    <td>{ i.date.strftime("%d/%m/%Y")}</td>
                    <td>{ i.total}</td>
                    <td>{ i.tva}</td>
                    <td>{ i.client.name }</td>
                    <td>{ i.client.code }</td>
                    <td>{ i.client.region}</td>
                    <td>{ i.client.city}</td>
                    <td>{ i.client.soldfacture}</td>
                    <td>{ i.salseman }</td>
                    <td class="d-flex justify-content-between">
                    <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
                    <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
                    </td>
                    <td >
                        {i.note}
                    </td>

                    <td>
                    {i.bon.bon_no if i.bon else "--"}
                    </td>
                    <td class="d-flex">
                        <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                        <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
                    </td>
                </tr>
                '''
            return JsonResponse({
                'trs':render(request, 'fclist.html', {'bons':bons, 'loadmore':True, 'term':term, 'year':year, 'startdate':startdate, 'enddate':enddate, 'comptable':comptable, 'facturesection':facture}).content.decode('utf-8'),
                'has_more': len(bons) == per_page,
                'total':total,
                'totaltva':totaltva,
                'facturesection':facture
            })
        else:
            print('>>>>>in term with daterange ')
            bons=Facture.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-facture_no')[start:end]
            total=round(Facture.objects.filter(q_objects).filter(date__range=[startdate, enddate]).exclude(client_id=3731 if facture else None).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
            totaltva=round(Facture.objects.filter(q_objects).filter(date__range=[startdate, enddate]).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
            trs=''
            for i in bons:
                trs+=f'''
                <tr class="ord {"text-danger" if i.ispaid else ''}
                 fc-row"
                    style="color:{"blue" if i.bon else ""} "
                  year={year} term={term} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                    <td>{ i.facture_no }</td>
                    <td>{ i.date.strftime("%d/%m/%Y")}</td>
                    <td>{ i.total}</td>
                    <td>{ i.tva}</td>
                    <td>{ i.client.name }</td>
                    <td>{ i.client.code }</td>
                    <td>{ i.client.region}</td>
                    <td>{ i.client.city}</td>
                    <td>{ i.client.soldfacture}</td>
                    <td>{ i.salseman }</td>
                    <td class="d-flex justify-content-between">
                    <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
                    <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
                    </td>
                    <td >
                        {i.note}
                    </td>

                    <td>
                    {i.bon.bon_no if i.bon else "--"}
                    </td>
                    <td class="d-flex">
                        <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                        <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
                    </td>
                </tr>
                '''
            return JsonResponse({
                'trs':render(request, 'fclist.html', {'bons':bons, 'loadmore':True, 'term':term, 'year':year, 'startdate':startdate, 'enddate':enddate, 'comptable':comptable, 'facturesection':facture}).content.decode('utf-8'),
                'has_more': len(bons) == per_page,
                'total':total,
                'totaltva':totaltva,
                'facturesection':facture
            })
    if startdate != '0' and enddate != '0':
        print('>>>>>>>>>> in start end dat')
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        print(startdate, enddate)
        bons=Facture.objects.filter(date__range=[startdate, enddate]).exclude(client_id=3731 if facture else None).order_by('-facture_no')[start:end]
        total=round(Facture.objects.filter(date__range=[startdate, enddate]).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        totaltva=round(Facture.objects.filter(date__range=[startdate, enddate]).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''}
             fc-row"
             term={term}
                style="color:{"blue" if i.bon else ""} "
              startdate={startdate} enddate={enddate} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture:.2f}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
                <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
                </td>
                <td>
                    {i.note}
                </td>

                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
                <td class="d-flex">
                    <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                    <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':render(request, 'fclist.html', {'bons':bons, 'loadmore':True, 'term':term, 'year':year, 'startdate':startdate, 'enddate':enddate, 'comptable':comptable, 'facturesection':facture}).content.decode('utf-8'),
            'has_more': len(bons) == per_page,
            'total':total,
            'totaltva':totaltva,
            'facturesection':facture
        })
    if year=="0":
        print('>>> in nothing')
        bons= Facture.objects.filter(date__year=timezone.now().year).exclude(client_id=3731 if facture else None).order_by('-facture_no')[start:end]
        total=round(Facture.objects.filter(date__year=timezone.now().year).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        totaltva=round(Facture.objects.filter(date__year=timezone.now().year).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''}
             fc-row"
                style="color:{"blue" if i.bon else ""} "
              year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture:.2f}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
                <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
                </td>
                <td>
                    {i.note}
                </td>

                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
                <td class="d-flex">
                    <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                    <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':render(request, 'fclist.html', {'bons':bons, 'loadmore':True, 'term':term, 'year':year, 'startdate':startdate, 'enddate':enddate, 'comptable':comptable, 'facturesection':facture}).content.decode('utf-8'),
            'has_more': len(bons) == per_page,
            'total':total,
            'totaltva':totaltva,
            'facturesection':facture
        })
    else:
        print('in year >> facture', facture)
        bons= Facture.objects.filter(date__year=year).exclude(client_id=3731 if facture else None).order_by('-facture_no')[start:end]
        total=round(Facture.objects.filter(date__year=year).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        totaltva=round(Facture.objects.filter(date__year=year).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture:.2f}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
                <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
                </td>
                <td>
                    {i.note}
                </td>

                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
                <td class="d-flex">
                    <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                    <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':render(request, 'fclist.html', {'bons':bons, 'loadmore':True, 'term':term, 'year':year, 'startdate':startdate, 'enddate':enddate, 'comptable':comptable, 'facturesection':facture}).content.decode('utf-8'),
            'has_more': len(bons) == per_page,
            'total':total,
            'totaltva':totaltva,

        })

def loadlistfccopy(request):
    facture=request.GET.get('facture')=='1'
    thisyear=timezone.now().year
    page = int(request.GET.get('page', 1))
    year =request.GET.get('year')
    startdate =request.GET.get('startdate')
    enddate =request.GET.get('enddate')
    term =request.GET.get('term')
    comptable =request.GET.get('comptable')
    print('>> page, year, term, startdate, enddate, comptable', page, year, term, startdate, enddate, comptable)
    per_page = 50  # Adjust as needed
    start = (page - 1) * per_page
    end = page * per_page
    print('>>>>>', start, end, page)
    print('>>>>> term', term)
    if comptable=='1':
        if year=='0':
            bons=Facture.objects.filter(hascopy=True, date__year=thisyear, isaccount=True).exclude(client_id=3731 if facture else None).order_by('-facture_no')[start:end]
        else:
            bons=Facture.objects.filter(hascopy=True, date__year=year, isaccount=True).exclude(client_id=3731 if facture else None).order_by('-facture_no')[start:end]

        trs=''
        for i in bons:
            trs+=f'''
                <tr class="ord {"text-danger" if i.ispaid else ''}
                 fccopy-row"
                    style="color:{"blue" if i.bon else ""} " comptable="1" ondblclick="createtab('bonl{i.id}', 'BL {i.facture_no}', '/products/facturedetailscopy/{i.id}')">
                    <td>{ i.facture_no }</td>
                    <td>{ i.date.strftime("%d/%m/%Y")}</td>
                    <td>{ i.total}</td>
                    <td>{ i.tva}</td>
                    <td>{ i.client.name }</td>
                    <td>{ i.client.code }</td>
                    <td>{ i.client.region}</td>
                    <td>{ i.client.city}</td>
                    <td>{ i.client.soldfacture}</td>
                    <td>{ i.salseman }</td>
                    <td class="d-flex justify-content-between">
                    <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
                    <button title="Facture Comptabilisé" class="btn border border-success border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
                    </td>
                    <td >
                        {i.note}
                    </td>

                    <td>
                    {i.bon.bon_no if i.bon else "--"}
                    </td>
                    <td class="d-flex">
                    <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                    <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
                    </td>
                </tr>
                '''
        return JsonResponse({
            'trs':render(request, 'fclistcopy.html', {'bons':bons, 'loadmore':True, 'facturesection':facture}).content.decode('utf-8'),
            'has_more': len(bons) == per_page,
            'facturesection':facture
        })
    if term != '0':

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in term.split('+'):
            # print('>>>>>>>term ', term)
            # if '-' in term[0]:
            #     date_range = term.split('-')
            #     start_date = datetime.strptime(date_range[0].strip(), '%d/%m/%Y')
            #     end_date = datetime.strptime(date_range[1].strip(), '%d/%m/%Y')
            #     q_objects &= (
            #         Q(client__name__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(salseman__name__iregex=term)|
            #         Q(facture_no__iregex=term)|
            #         Q(bon__bon_no__iregex=term)|
            #         Q(client__region__iregex=term)|
            #         Q(client__city__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(total__iregex=term)|
            #         Q(date__range=[start_date, end_date])
            #         )
            # else:
            #     q_objects &= (
            #         Q(client__name__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(salseman__name__iregex=term)|
            #         Q(facture_no__iregex=term)|
            #         Q(bon__bon_no__iregex=term)|
            #         Q(client__region__iregex=term)|
            #         Q(client__code__iregex=term)|
            #         Q(total__iregex=term)
            #     )
            q_objects &= (
                Q(client__name__iregex=term)|
                Q(client__city__iregex=term)|
                Q(salseman__name__iregex=term)|
                Q(facture_no__iregex=term)|
                Q(bon__bon_no__iregex=term)|
                Q(client__region__iregex=term)|
                Q(client__code__iregex=term)|
                Q(note__iregex=term)|
                Q(statusreg__iregex=term)|
                Q(total__iregex=term)
            )

        if startdate=='0' and enddate=='0':
            print('>>>>>in term without dates term, year', term, year)
            bons=Facture.objects.filter(q_objects).filter(hascopy=True, date__year=year).exclude(client_id=3731 if facture else None).order_by('-facture_no')[start:end]
            total=round(Facture.objects.filter(q_objects).filter(hascopy=True, date__year=year).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
            totaltva=round(Facture.objects.filter(q_objects).filter(hascopy=True, date__year=year).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
            trs=''
            for i in bons:
                trs+=f'''
                <tr class="ord {"text-danger" if i.ispaid else ''}
                 fc-row"
                    style="color:{"blue" if i.bon else ""} "
                  year="{year}" term="{term}" orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                    <td>{ i.facture_no }</td>
                    <td>{ i.date.strftime("%d/%m/%Y")}</td>
                    <td>{ i.total}</td>
                    <td>{ i.tva}</td>
                    <td>{ i.client.name }</td>
                    <td>{ i.client.code }</td>
                    <td>{ i.client.region}</td>
                    <td>{ i.client.city}</td>
                    <td>{ i.client.soldfacture}</td>
                    <td>{ i.salseman }</td>
                    <td class="d-flex justify-content-between">
                    <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
                    <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
                    </td>
                    <td >
                        {i.note}
                    </td>

                    <td>
                    {i.bon.bon_no if i.bon else "--"}
                    </td>
                    <td class="d-flex">
                        <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                        <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
                    </td>
                </tr>
                '''
            return JsonResponse({
                'trs':render(request, 'fclistcopy.html', {'bons':bons, 'loadmore':True, 'term':term, 'year':year, 'startdate':startdate, 'enddate':enddate, 'comptable':comptable, 'facturesection':facture}).content.decode('utf-8'),
                'has_more': len(bons) == per_page,
                'total':total,
                'totaltva':totaltva,
                'facturesection':facture
            })
        else:
            print('>>>>>in term with daterange ')
            bons=Facture.objects.filter(q_objects).filter(hascopy=True, date__range=[startdate, enddate]).order_by('-facture_no')[start:end]
            total=round(Facture.objects.filter(q_objects).filter(hascopy=True, date__range=[startdate, enddate]).exclude(client_id=3731 if facture else None).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
            totaltva=round(Facture.objects.filter(q_objects).filter(hascopy=True, date__range=[startdate, enddate]).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
            trs=''
            for i in bons:
                trs+=f'''
                <tr class="ord {"text-danger" if i.ispaid else ''}
                 fc-row"
                    style="color:{"blue" if i.bon else ""} "
                  year={year} term={term} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                    <td>{ i.facture_no }</td>
                    <td>{ i.date.strftime("%d/%m/%Y")}</td>
                    <td>{ i.total}</td>
                    <td>{ i.tva}</td>
                    <td>{ i.client.name }</td>
                    <td>{ i.client.code }</td>
                    <td>{ i.client.region}</td>
                    <td>{ i.client.city}</td>
                    <td>{ i.client.soldfacture}</td>
                    <td>{ i.salseman }</td>
                    <td class="d-flex justify-content-between">
                    <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
                    <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
                    </td>
                    <td >
                        {i.note}
                    </td>

                    <td>
                    {i.bon.bon_no if i.bon else "--"}
                    </td>
                    <td class="d-flex">
                        <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                        <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
                    </td>
                </tr>
                '''
            return JsonResponse({
                'trs':render(request, 'fclistcopy.html', {'bons':bons, 'loadmore':True, 'term':term, 'year':year, 'startdate':startdate, 'enddate':enddate, 'comptable':comptable, 'facturesection':facture}).content.decode('utf-8'),
                'has_more': len(bons) == per_page,
                'total':total,
                'totaltva':totaltva,
                'facturesection':facture
            })
    if startdate != '0' and enddate != '0':
        print('>>>>>>>>>> in start end dat')
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        print(startdate, enddate)
        bons=Facture.objects.filter(hascopy=True, date__range=[startdate, enddate]).exclude(client_id=3731 if facture else None).order_by('-facture_no')[start:end]
        total=round(Facture.objects.filter(hascopy=True, date__range=[startdate, enddate]).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        totaltva=round(Facture.objects.filter(hascopy=True, date__range=[startdate, enddate]).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''}
             fc-row"
             term={term}
                style="color:{"blue" if i.bon else ""} "
              startdate={startdate} enddate={enddate} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture:.2f}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
                <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
                </td>
                <td>
                    {i.note}
                </td>

                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
                <td class="d-flex">
                    <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                    <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':render(request, 'fclistcopy.html', {'bons':bons, 'loadmore':True, 'term':term, 'year':year, 'startdate':startdate, 'enddate':enddate, 'comptable':comptable, 'facturesection':facture}).content.decode('utf-8'),
            'has_more': len(bons) == per_page,
            'total':total,
            'totaltva':totaltva,
            'facturesection':facture
        })
    if year=="0":
        print('>>> in nothing')
        bons= Facture.objects.filter(hascopy=True, date__year=timezone.now().year).exclude(client_id=3731 if facture else None).order_by('-facture_no')[start:end]
        total=round(Facture.objects.filter(hascopy=True, date__year=timezone.now().year).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        totaltva=round(Facture.objects.filter(hascopy=True, date__year=timezone.now().year).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''}
             fc-row"
                style="color:{"blue" if i.bon else ""} "
              year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture:.2f}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
                <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
                </td>
                <td>
                    {i.note}
                </td>

                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
                <td class="d-flex">
                    <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                    <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':render(request, 'fclistcopy.html', {'bons':bons, 'loadmore':True, 'term':term, 'year':year, 'startdate':startdate, 'enddate':enddate, 'comptable':comptable, 'facturesection':facture}).content.decode('utf-8'),
            'has_more': len(bons) == per_page,
            'total':total,
            'totaltva':totaltva,
            'facturesection':facture
        })
    else:
        print('in year >> facture', facture)
        bons= Facture.objects.filter(hascopy=True, date__year=year).exclude(client_id=3731 if facture else None).order_by('-facture_no')[start:end]
        total=round(Facture.objects.filter(hascopy=True, date__year=year).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        totaltva=round(Facture.objects.filter(hascopy=True, date__year=year).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture:.2f}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
                <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
                </td>
                <td>
                    {i.note}
                </td>

                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
                <td class="d-flex">
                    <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                    <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':render(request, 'fclistcopy.html', {'bons':bons, 'loadmore':True, 'term':term, 'year':year, 'startdate':startdate, 'enddate':enddate, 'comptable':comptable, 'facturesection':facture}).content.decode('utf-8'),
            'has_more': len(bons) == per_page,
            'total':total,
            'totaltva':totaltva,

        })


def searchforlistfc(request):
    term=request.GET.get('term')
    facture=request.GET.get('facture')=='1'
    searchedterm=request.GET.get('term')
    startdate=request.GET.get('startdate') or '0'
    enddate=request.GET.get('enddate') or '0'
    print('>>', startdate, enddate)
    year=request.GET.get('year')
    if(term==''):
        bons=Facture.objects.filter(date__year=year).order_by('-facture_no')[:50]
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')" style="{'background:yellowgreen;' if i.isaccount else ''}">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div></div>
                <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>

                </td>
                <td>
                    {i.note}
                </td>
                <td>
                    {i.getreglement()[:16]}
                </td>
                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
                <td class="d-flex">
                    <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                    <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs
        })
    print('>>>>term', term)

    # Split the term into individual words separated by '*'
    search_terms = term.split('+')

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        print('>>>>>> term', term)
        q_objects &= (
            Q(client__name__iregex=term)|
            Q(client__city__iregex=term)|
            Q(salseman__name__iregex=term)|
            Q(facture_no__iregex=term)|
            Q(bon__bon_no__iregex=term)|
            Q(client__region__iregex=term)|
            Q(client__code__iregex=term)|
            Q(note__iregex=term)|
            Q(statusreg__iregex=term)|
            Q(total__iregex=term)
        )

    if startdate=='0' and enddate=='0':
        print('>>>> search list fc, startdate and enddate are 0')
        bons=Facture.objects.filter(q_objects).filter(date__year=year).exclude(client_id=3731 if facture else None).order_by('-facture_no')[:50]
        total=round(Facture.objects.filter(q_objects).filter(date__year=year).exclude(client_id=3731 if facture else None).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)

    else:
        bons=Facture.objects.filter(q_objects).filter(date__range=[startdate, enddate]).exclude(client_id=3731 if facture else None).order_by('-facture_no')[:50]
        total=round(Facture.objects.filter(q_objects).filter(date__range=[startdate, enddate]).exclude(client_id=3731 if facture else None).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
    # when it's facture section, we need to get factures but hide the factures with client is diver
    # if year=='0':
    #     bons=Facture.objects.filter(q_objects).filter(date__year=thisyear).order_by('-facture_no')[:50]
    #     total=round(Facture.objects.filter(q_objects).filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
    # else:
    #     bons=Facture.objects.filter(q_objects).filter(date__year=year).order_by('-facture_no')[:50]
    #     total=round(Facture.objects.filter(q_objects).filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    trs=''
    for i in bons:
        trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" term={searchedterm} year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')" style="{'background:yellowgreen;' if i.isaccount else ''}">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                    <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>
                    <button title="Facture Comptabilisé" class="btn border border-success" onclick="makefacturecompta(event, '{i.id}')"></button>
                </td>
                <td>
                    {i.note}
                </td>
                <td>
                    {i.getreglement()[:16]}
                </td>
                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
                <td class="d-flex">
                    <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                    <button title="Imprimer" class="btn btn-sm bi bi-download" onclick="printfacture('{i.id}')"></button>
                </td>
            </tr>
            '''
    return JsonResponse({
        'trs':trs,
        'total':total,

    })

def searchforlistfccopy(request):
    term=request.GET.get('term')
    facture=request.GET.get('facture')=='1'
    searchedterm=request.GET.get('term')
    startdate=request.GET.get('startdate') or '0'
    enddate=request.GET.get('enddate') or '0'
    print('>>', startdate, enddate)
    year=request.GET.get('year')
    if(term==''):
        bons=Facture.objects.filter(hascopy=True, date__year=year).order_by('-facture_no')[:50]
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetailscopy/{i.id}')" style="{'background:yellowgreen;' if i.isaccount else ''}">
                <td>{ i.copynumber }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.salseman }</td>

                <td>
                    {i.note}
                </td>



            </tr>
            '''
        return JsonResponse({
            'trs':trs
        })
    print('>>>>term', term)

    # Split the term into individual words separated by '*'
    search_terms = term.split('+')

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        print('>>>>>> term', term)
        q_objects &= (
            Q(client__name__iregex=term)|
            Q(client__city__iregex=term)|
            Q(salseman__name__iregex=term)|
            Q(facture_no__iregex=term)|
            Q(bon__bon_no__iregex=term)|
            Q(client__region__iregex=term)|
            Q(client__code__iregex=term)|
            Q(note__iregex=term)|
            Q(statusreg__iregex=term)|
            Q(total__iregex=term)
        )

    if startdate=='0' and enddate=='0':
        print('>>>> search list fc, startdate and enddate are 0')
        bons=Facture.objects.filter(q_objects).filter(hascopy=True, date__year=year).exclude(client_id=3731 if facture else None).order_by('-facture_no')[:50]
        total=round(Facture.objects.filter(q_objects).filter(hascopy=True, date__year=year).exclude(client_id=3731 if facture else None).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)

    else:
        bons=Facture.objects.filter(q_objects).filter(hascopy=True, date__range=[startdate, enddate]).exclude(client_id=3731 if facture else None).order_by('-facture_no')[:50]
        total=round(Facture.objects.filter(q_objects).filter(hascopy=True, date__range=[startdate, enddate]).exclude(client_id=3731 if facture else None).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
    # when it's facture section, we need to get factures but hide the factures with client is diver
    # if year=='0':
    #     bons=Facture.objects.filter(q_objects).filter(hascopy=True, date__year=thisyear).order_by('-facture_no')[:50]
    #     total=round(Facture.objects.filter(q_objects).filter(hascopy=True, date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
    # else:
    #     bons=Facture.objects.filter(q_objects).filter(hascopy=True, date__year=year).order_by('-facture_no')[:50]
    #     total=round(Facture.objects.filter(q_objects).filter(hascopy=True, date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    trs=''
    for i in bons:
        trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetailscopy/{i.id}')" style="{'background:yellowgreen;' if i.isaccount else ''}">
                <td>{ i.copynumber }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.salseman }</td>

                <td>
                    {i.note}
                </td>



            </tr>
            '''
    return JsonResponse({
        'trs':trs,
        'total':total,

    })

def createnewclientaccount(request):
    clientid=request.POST.get('clientid')
    client= Client.objects.get(pk=clientid)
    olduser=client.user
    username=request.POST.get('username').strip()
    password=request.POST.get('password')
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    if serverip:
        res=req.get(f'http://{serverip}/products/createnewclientaccount', {
            'username':username,
            'password':password,
            'clientcode':client.code
        })
        if json.loads(res.text)['success']:
            return JsonResponse({
                'success':True
            })
        else:
            return JsonResponse({
                'success':False,
                'error':f"{json.loads(res.text)['error']}"
            })
    # user=User.objects.filter(username=username).first()
    # if user:
    return JsonResponse({
        'success':False,
        'error':'No server'
    })
    #check for username
    user=User.objects.filter(username=username).first()
    print('>>>>>>>>>>>', username, client.code)
    if user:
        return JsonResponse({
            'success':False,
            'error':'Username exist déja'
        })
    user=User.objects.create_user(username=username, password=password)
    try:
        response=req.get(f'http://{serverip}/products/createnewclientaccount', {
            'username':username,
            'password':password,
            'clientcode':client.code
        })
        response.raise_for_status()
        cart=Cart.objects.filter(user=olduser).first()
        wich=Wich.objects.filter(user=olduser).first()
        if cart:
            cart.user=user
            cart.save()
        if wich:
            wich.user=user
            wich.save()
        olduser.delete()
        # create user
        # assign user to client
        group, created = Group.objects.get_or_create(name="clients")
        user.groups.add(group)
        user.save()
        client.user=user
        client.save()

        return JsonResponse({
            'success':True
        })
    except Exception as e:
        print('>>>>>>>', e)
        return JsonResponse({
            'success':False,
            'error':'ERROR SERVER'
        })


def createnewrepaccount(request):
    repid=request.POST.get('repid')
    rep= Represent.objects.get(pk=repid)
    olduser=rep.user
    username=request.POST.get('username')
    password=request.POST.get('password')
    serverip = Setting.objects.only('serverip').first()
    serverip = serverip.serverip if serverip else None
    if serverip:
        res=req.get(f'http://{serverip}/products/createnewrepaccount', {
            'username':username,
            'password':password,
            'repid':repid
        })
        if json.loads(res.text)['success']:
            return JsonResponse({
                'success':True
            })
        else:
            return JsonResponse({
                'success':False,
                'error':f"{json.loads(res.text)['error']}"
            })
    # user=User.objects.filter(username=username).first()
    # if user:
    return JsonResponse({
        'success':False,
        'error':'No server'
    })
    #     response.raise_for_status()
    #     user=User.objects.create_user(username=username, password=password)
    #     # assign user to rep
    #     cart=Cart.objects.filter(user=olduser).first()
    #     wich=Wich.objects.filter(user=olduser).first()
    #     if cart:
    #         cart.user=user
    #         cart.save()
    #     if wich:
    #         wich.user=user
    #         wich.save()
    #     olduser.delete()
    #     group, created = Group.objects.get_or_create(name="salsemen")
    #     user.groups.add(group)
    #     user.save()
    #     rep.user=user
    #     rep.save()
    #     return JsonResponse({
    #         'success':True
    #     })
    # except Exception as e:
    #     print('>>>>>>>', e)
    #     return JsonResponse({
    #         'success':False,
    #         'error':'ERROR SERVER'
    #     })



def yeardatabl(request):
    year=request.GET.get('year')
    # get all bls of that year
    bls=Bonlivraison.objects.filter(date__year=year).order_by('-id')[:50]
    trs=''
    for i in bls:
        trs+=f'''
        <tr class="ord {"text-danger" if i.ispaid else ''} bl-row" year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')">
            <td>{ i.bon_no }</td>
            <td>{ i.date.strftime("%d/%m/%Y")}</td>
            <td>{ i.client.name }</td>
            <td>{ i.client.code }</td>
            <td>{ i.total}</td>
            <td>{ i.client.region}</td>
            <td>{ i.client.city}</td>
            <td>{ i.client.soldbl}</td>
            <td>{ i.salseman }</td>
            <td class="d-flex justify-content-between">
              <div>
              {'R0' if i.ispaid else 'N1' }

              </div>
              <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

            </td>
            <td class="text-danger">
            {'OUI' if i.isfacture else 'NON'}

            </td>

            <td>
              {i.commande.order_no if i.commande else '--'}
            </td>
            <td>
              {i.modlvrsn}
            </td>
          </tr>
        '''
    return JsonResponse({
        'trs':render(request, 'bllist.html', {'bons':bls}).content.decode('utf-8'),
        'total':round(Bonlivraison.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    })

def yeardatabachat(request):
    year=request.GET.get('year') or this_year
    # get all bls of that year
    facture=request.GET.get('facture')=='1'
    if facture:
        bons=Itemsbysupplier.objects.filter(date__year=year, isfacture=True).order_by('-date')
    else:
        bons=Itemsbysupplier.objects.filter(date__year=year).order_by('-date')
    trs=''
    for order in bons:
        trs+=f'''
        <tr class="ord {'' if facture else 'achat-row'}" orderid="{order.id}" ondblclick="createtab('bonachat{order.id}', 'Bon achat {order.nbon}', '/products/bonachatdetails/{order.id}')">
            <td>{ order.nbon }</td>
            <td>{ order.date.strftime("%d/%m/%Y") }</td>
            <td>{ order.supplier.name }</td>
            <td>{ order.total}</td>
            <td>{ order.tva}</td>
            <td>{"Facture"if order.isfacture else "Bl"}</td>
            <td>{ round(order.supplier.rest, 2) }</td>
            <td class="d-flex">
                <div>{"R0"if order.ispaid else "N1"}</div>

              <div style="width:15px; height:15px; border-radius:50%; background:{"green"if order.ispaid else "red"};" ></div>



            </td>
            <td>
              <button class="btn bi bi-download" onclick="printbonachat('{order.id}')"></button>
            </td>

          </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':round(Itemsbysupplier.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    })

def yeardatabc(request):
    year=request.GET.get('year')
    print(year)
    # get all bls of that year
    bons=Order.objects.filter(date__year=year).order_by('-id')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="orderrow {'text-danger' if not i.isdelivered else ''}" year={year} orderid="{i.code}" ondblclick="createtab('command{i.id}', 'Commande {i.order_no}', '/products/boncommandedetails/{i.id}')">
            <td>{ i.order_no }</td>
            <td>{ i.date.strftime('%d/%m/%Y') }</td>
            <td>{ i.client.name if i.client else '' }</td>
            <td>{ i.client.code if i.client else '' }</td>
            <td>{ i.total}</td>
            <td>{ i.client.region if i.client else '' }</td>
            <td>{ i.client.city if i.client else '' }</td>
            <td>{ i.client.soldbl if i.client else '' }</td>
            <td>{ i.salseman }</td>
            <td>
            {"Non" if i.isclientcommnd else "OUI"}
            </td>
            <td>
              {"R1" if i.isdelivered else  "R0" }
            </td>
            <td>
              {i.note}
            </td>


          </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':round(Order.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    })



def yeardatafccopy(request):
    year=request.GET.get('year')
    print(year)
    # get all bls of that year
    bls=Facture.objects.filter(date__year=year, hascopy=True).order_by('-facture_no')[:50]
    trs=''
    for i in bls:
        trs+=f'''
        <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Bl {i.facture_no}', '/products/facturedetailscopy/{i.id}')">
            <td>{ i.copynumber }</td>
            <td>{ i.date.strftime("%d/%m/%Y")}</td>
            <td>{ i.total}</td>
            <td>{ i.client.name }</td>
            <td>{ i.client.code }</td>
            <td>{ i.client.region}</td>
            <td>{ i.client.city}</td>
            <td>{ i.salseman }</td>

            <td>
                {i.note}
            </td>
            <td></td>


        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        # 'trs':render(request, 'fclist.html', {'bons':bls, 'loadmore':True}).content.decode('utf-8'),
        'total':round(bls.aggregate(Sum('total'))['total__sum'] or 0, 2)
    })

def yeardatafc(request):
    year=request.GET.get('year')
    print(year)
    # get all bls of that year
    bls=Facture.objects.filter(date__year=year, hascopy=True).order_by('-facture_no')[:50]
    trs=''
    for i in bls:
        trs+=f'''
        <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetailscopy/{i.id}')">
            <td>{ i.copynumber }</td>
            <td>{ i.date.strftime("%d/%m/%Y")}</td>
            <td>{ i.total}</td>
            <td>{ i.client.name }</td>
            <td>{ i.client.code }</td>
            <td>{ i.client.region}</td>
            <td>{ i.client.city}</td>
            <td>{ i.salseman }</td>

            <td>
                {i.note}
            </td>
            <td></td>


        </tr>
        '''
    return JsonResponse({
        'trs':render(request, 'fclist.html', {'bons':bls}).content.decode('utf-8'),
        # 'trs':render(request, 'fclist.html', {'bons':bls, 'loadmore':True}).content.decode('utf-8'),
        'total':round(bls.aggregate(Sum('total'))['total__sum'] or 0, 2)
    })



def loadreglbl(request):
    page = int(request.GET.get('page', 1))
    per_page = 50  # Adjust as needed
    term=request.GET.get('term')
    year=request.GET.get('year')
    start = (page - 1) * per_page
    end = page * per_page
    print('start', start, end)
    if term=="":
        reblbls = PaymentClientbl.objects.filter(date__year=year).order_by('-echance')[start:end]
    else:
        q_objects = Q()
        for term in term.split('+'):
            q_objects &= (Q(client__name__icontains=term) | Q(client__code__icontains=term) | Q(mode__icontains=term) | Q(npiece__icontains=term) | Q(amount__icontains=term))
        reblbls=PaymentClientbl.objects.filter(date__year=year).filter(q_objects).order_by('-echance')[start:end]
    return JsonResponse({
        'trs':render(request, 'reglbllist.html', {'bons':reblbls}).content.decode('utf-8'),
        'has_more': len(reblbls) == per_page
    })


def laodreglfc(request):
    page = int(request.GET.get('page', 1))
    term=request.GET.get('term')
    year=request.GET.get('year')
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    if term=='':
        reblbls = PaymentClientfc.objects.all().order_by('-echance')[start:end]
    # trs=''
    # for i in reblbls:
    #     tooltip_content = ''.join([bon.bon_no for bon in i.bons.all()])
    #     trs += f'''
    #     <tr style="{ "background: yellow;" if (i.echance and i.echance == today) else ("color: red;" if (i.echance and i.echance < today) else "") }" class="reglbl-row">
    #         <td>
    #             { i.mode } <br>
    #             <!-- <select name="updatemodereglbl{ i.id }" id=""></select> -->
    #         </td>
    #         <td>
    #             { i.npiece } <br>
    #             <input type="text" class="d-none updatenpiecereglbl{ i.id }">
    #         </td>
    #         <td>
    #             { i.amount }<br>
    #             <!-- <input type="text" class="d-none updateamountreglbl{ i.id }"> -->
    #         </td>
    #         <td>
    #             { i.date }<br>
    #             <input type="date" class="d-none updatetdatereglbl{ i.id }">
    #         </td>
    #         <td>
    #             { i.client.name }<br>
    #         </td>
    #         <td>
    #             { i.client.code }<br>
    #         </td>
    #         <td>
    #             { i.echance} <br>
    #         </td>
    #         <td class="d-flex justify-content-between">
    #             <button type="button" class="btn btn-secondary btn-sm" data-toggle="tooltip" data-placement="top" data-tooltip="Bon Nos: {tooltip_content}">
    #             </button>
    #             <button class="btn btn-success btn-sm" onclick="updatereglementbl({ i.id })">✐</button>
    #         </td>
    #     </tr>
    # '''
    else:
        q_objects = Q()
        for term in term.split('+'):
            q_objects &= (Q(client__name__iregex=term) | Q(client__code__iregex=term) | Q(mode__iregex=term) | Q(npiece__iregex=term) | Q(amount__iregex=term))
        reblbls=PaymentClientfc.objects.filter(date__year=year).filter(q_objects).order_by('echance')[start:end]
    return JsonResponse({
        'trs':render(request, 'reglfclist.html', {'bons':reblbls}).content.decode('utf-8'),
        'has_more': len(reblbls) == per_page
    })



def loadclients(request):
    page = int(request.GET.get('page', 1))
    per_page = 50  # Adjust as needed
    term=request.GET.get('term')
    facture=request.GET.get('facture')=='1'
    start = (page - 1) * per_page
    end = page * per_page
    if term=="":
        print('>> search without term')
        if facture:
          clients = Client.objects.all().order_by('-soldfacture')[start:end]
        else:
          clients = Client.objects.all().order_by('-soldtotal')[start:end]
        trs=''
        for i in clients:
            trs+=f'''
            <tr class="client-row" style="background:{'#a8dfc5' if i.diver else ''};">
                <td>
                    <button class="btn editsuppbtn border" id="{i.id}" data-toggle="modal" data-target="#editclientmodal" onclick="populateclientfields({i.id})">
                        ✐
                    </button>
                </td>
                <td onclick="createtab('client{i.id}', 'Client: {i.name} ', '/clients/client/{i.id}')">{i.name} </td>
                <td>{i.code}</td>
                <td>{i.phone}</td>
                <td>{i.city}</td>
                <td>{i.region}</td>
                <td>

                    {i.represent.name if i.represent else ''}


                </td>
                <td>{i.soldtotal:.2f}</td>
                <td style="background: yellowgreen;">{i.soldbl:.2f}</td>
                <td style="background: aliceblue;">{i.soldfacture:.2f}</td>
                <td>{i.ice}</td>
            </tr>
            '''
        return JsonResponse({
            'trs':render(request, 'clienttrs.html', {'clients':clients, 'facturesection':facture}).content.decode('utf-8'),
            'has_more': len(clients) == per_page
        })
    print('>> search with term', term)
    q_objects = Q()
    for i in term.split('+'):
        q_objects &= (Q(city__icontains=term) |
        Q(name__icontains=term) |
        Q(ice__icontains=term) |
        Q(phone__icontains=term) |
        Q(region__icontains=term) |
        Q(code__icontains=term) |
        Q(represent__name__icontains=term) |
        Q(address__icontains=term))
    if facture:
      clients=Client.objects.filter(q_objects).order_by('-soldfacture')[start:end]
    else:
      clients=Client.objects.filter(q_objects).order_by('-soldtotal')[start:end]

    trs=''
    for i in clients:
        trs+=f'''
        <tr class="client-row" style="background:{'#a8dfc5' if i.diver else ''};">
            <td>
                <button class="btn editsuppbtn border" id="{i.id}" data-toggle="modal" data-target="#editclientmodal" onclick="populateclientfields({i.id})">
                    ✐
                </button>
            </td>
            <td onclick="createtab('client{i.id}', 'Client: {i.name} ', '/clients/client/{i.id}')">{i.name} </td>
            <td>{i.code}</td>
            <td>{i.phone}</td>
            <td>{i.city}</td>
            <td>{i.region}</td>
            <td>

                {i.represent.name if i.represent else ''}


            </td>
            <td>{i.soldtotal:.2f}</td>
            <td style="background: yellowgreen;">{i.soldbl:.2f}</td>
            <td style="background: aliceblue;">{i.soldfacture:.2f}</td>
            <td>{i.ice}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':render(request, 'clienttrs.html', {'clients':clients, 'facturesection':facture}).content.decode('utf-8'),
        'has_more': len(clients) == per_page
    })

def exportproducts(request):
    categoryid=request.GET.get('categoryid')
    if categoryid=='0':
        products=Produit.objects.all()
        filename='Produit_tous'+today.strftime('%d/%m/%y')+'.xlsx'
    else:
        category=Category.objects.get(pk=categoryid)
        products=Produit.objects.filter(category__id=categoryid)
        filename='Produit_'+category.name+today.strftime('%d%m%y')+'.xlsx'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


    # Create a new Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write column headers
    ws.append([
    'id',
    'ref',
    'name',
    'category',
    'devise',
    'buyprice',
    'sellprice',
    'remise',
    'prixnet',
    'stockfacture',
    'mark',
    'diametre',
    'block',
    'refeq1',
    'refeq2',
    'stocktotal',
    # 'entre',
    # 'sortie',
    # 'initial',
    'commession'])

    # Write product data
    for product in products:
        entr=round(Stockin.objects.filter(product=product).aggregate(Sum('quantity'))['quantity__sum'] or 0, 2)
        sortbl=round(Livraisonitem.objects.filter(product=product, isfacture=False).aggregate(Sum('qty'))['qty__sum'] or 0, 2)
        sortfc=round(Outfacture.objects.filter(product=product).exclude(facture__bon__isnull=True).aggregate(Sum('qty'))['qty__sum'] or 0, 2)
        sort=sortbl+sortfc
        ws.append([
            product.id,
            product.ref,
            product.name,
            product.category.name if product.category else '',  # Extract category name
            product.devise,
            product.buyprice,
            product.sellprice,
            product.remise,
            product.prixnet,
            product.stockfacture,
            product.mark.name if product.mark else '',
            product.diametre,
            product.block,
            product.refeq1,
            product.refeq2,
            product.stocktotal,
            product.getpercentage()
            # entr,
            # sort,
            # entr-sort+product.stocktotal,

        ])

    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    # Save the workbook to the response
    wb.save(response)

    return response

def datepdct(request):
    datefrom=request.GET.get('datefrom')
    dateto=request.GET.get('dateto')
    pdctid=request.GET.get('pdctid')
    print('dates>>>>>>', datefrom, dateto)
    datefrom=datetime.strptime(datefrom, '%Y-%m-%d')
    dateto=datetime.strptime(dateto, '%Y-%m-%d')
    stockin=Stockin.objects.filter(product_id=pdctid, date__range=[datefrom, dateto]).order_by('date')
    stockout=Livraisonitem.objects.filter(product_id=pdctid, bon__date__range=[datefrom, dateto]).order_by('date')
    totalqtyin=stockin.aggregate(Sum('quantity'))['quantity__sum'] or 0
    totalqtyout=stockout.aggregate(Sum('qty'))['qty__sum'] or 0
    totalcoutin=stockin.aggregate(Sum('total'))['total__sum'] or 0
    totalrevenu=stockout.aggregate(Sum('total'))['total__sum'] or 0
    trin=''
    trout=''
    for i in stockin:
        trin+=f'''
            <tr>
                <td>
                    {i.date.strftime('%d/%m/%Y')}
                </td>
                <td>
                    {i.nbon}
                </td>
                <td>
                    {i.supplier.name}
                </td>
                <td>
                    {i.quantity}
                </td>
                <td>
                    {i.price}
                </td>
                <td>
                    {i.total}
                </td>
            </tr>
        '''
    for i in stockout:
        trout+=f'''
            <tr>
                <td>
                    {i.date.strftime('%d/%m/%Y')}
                </td>
                <td>
                    {i.bon.bon_no}
                </td>
                <td>
                    {i.bon.client.name}
                </td>
                <td>
                    {i.qty}
                </td>
                <td>
                    {i.price}
                </td>
                <td>
                    {i.remise}%
                </td>
                <td>
                    {i.total}
                </td>
            </tr>
        '''
    return JsonResponse({
        'trin':trin,
        'trout':trout,
        'totalqtyin':totalqtyin,
        'totalqtyout':totalqtyout,
        'totalcoutin':totalcoutin,
        'totalrevenu':totalrevenu
    })


def showdeactivated(request):
    products=Produit.objects.filter(isactive=False)[:100]
    trs=''
    for i in products:
        trs+=f"""
            <tr ondblclick="createtab('addpdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
                style="background:{'#f3d6d694;' if not i.isactive else '' }"
                    data-product-id="{ i.id }" class="product-row notactive">
                    <td >
                        {i.ref.upper()}
                    </td>
                    <td>
                        {i.name}
                    </td>

                    <td class="text-center prachat">
                        {i.buyprice if i.buyprice else 0}
                    </td>
                    <td class="text-center">
                        {i.sellprice}
                    </td>
                    <td class="text-center">
                        {i.remise}
                    </td>
                    <td class="text-center">
                        {i.prixnet}
                    </td>
                    <td class="text-center text-danger stock">
                        {i.stocktotal}
                    </td>
                    <td class="text-center stockfacture" style="color: blue;">
                        <span class="stockfacture invisible">{i.stockfacture}</span>
                    </td>

                    <td>
                        {i.diametre}
                    </td>
                    <td class="text-success">
                        {i.block}
                    </td>
                    <td>
                        {i.coderef}
                    </td>
                    <td>
                        {i.getequivalent()[0] if i.getequivalent() else ''}
                    </td>
                    <td>
                        {i.getequivalent()[1] if i.getequivalent() else ''}
                    </td>
                    <td>
                        {i.getequivalent()[2] if i.getequivalent() else ''}
                    </td>
                    <td>
                        {i.mark}
                    </td>

                </tr>
        """

    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == 100
    })

def searchforlistclient(request):
    term=request.GET.get('term').strip()
    facture=request.GET.get('facture')=='1'

    # Split the term into individual words separated by '*'
    print('>> in term')

    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:

            # term = ''.join(char for char in term if char.isalnum())
            q_objects &= (Q(city__icontains=term) |
                Q(name__icontains=term) |
                Q(ice__icontains=term) |
                Q(phone__icontains=term) |
                Q(region__icontains=term) |
                Q(code__icontains=term) |
                Q(represent__name__icontains=term) |
                Q(address__icontains=term))
    if facture:
      clients=Client.objects.filter(q_objects).order_by('-soldfacture')[:50]
    else:
      clients=Client.objects.filter(q_objects).order_by('-soldtotal')[:50]
    return JsonResponse({
        'trs':render(request, 'clienttrs.html', {'clients':clients, 'facturesection':facture}).content.decode('utf-8'),
    })


def laodblregl(request):
    page = int(request.GET.get('page', 1))
    clientid=request.GET.get('clientid')
    #nr: non reglé
    nr=request.GET.get('nr')
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    print('>>>>>>>>>', nr)
    if nr=='0':
        print('all', page)
        bons=Bonlivraison.objects.filter(client_id=clientid).order_by('date')[start:end]
        for i in bons:
            trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)"></td></tr>'
    else:
        print('non regl')
        bons=Bonlivraison.objects.filter(client_id=clientid, ispaid=False).order_by('date')[start:end]
        for i in bons:
            trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow nr" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)"></td></tr>'



    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })


def laodblinupdateregl(request):
    page = int(request.GET.get('page', 1))

    reglementid=request.GET.get('reglementid')
    print('>>>>>>', reglementid)
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    reglement=PaymentClientbl.objects.get(pk=reglementid)

    bons=reglement.bons.all()
    # bons without bons in reglement
    bons=Bonlivraison.objects.filter(client=reglement.client).exclude(pk__in=[bon.pk for bon in bons]).order_by('date')[start:end]
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="loadblinupdateregl" reglemntid="{reglementid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)"></td></tr>'



    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })
#here
def laodfcinupdateregl(request):
    page = int(request.GET.get('page', 1))

    reglementid=request.GET.get('reglementid')
    print('>>>>>>', reglementid)
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    reglement=PaymentClientfc.objects.get(pk=reglementid)

    bons=reglement.factures.all().order_by('date')
    # bons without bons in reglement
    bons=Facture.objects.filter(client=reglement.client).exclude(pk__in=[bon.pk for bon in bons]).order_by('date')[start:end]
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="loadblinupdatereglfc" reglemntid="{reglementid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'



    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })

# the facture load in the add reglement
def laodfcregl(request):
    page = int(request.GET.get('page', 1))
    clientid=request.GET.get('clientid')
    #nr: non reglé
    nr=request.GET.get('nr')
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    if nr=='0':
        bons=Facture.objects.filter(client_id=clientid).order_by('date')[start:end]
        for i in bons:
            trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="fcreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    else:
        bons=Facture.objects.filter(client_id=clientid, ispaid=False).order_by('date')[start:end]
        for i in bons:
            trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="fcreglrow nr" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'



    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })

def searchclientbls(request):
    clientid=request.GET.get('clientid')
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Use list comprehension for building search terms
    search_terms = [term for term in regex_search_term.split('*') if term]

    # Use __icontains for case-insensitive searches
    q_objects = Q()
    q_objects &= (Q(salseman__name__icontains=term) | Q(bon_no__icontains=term) | Q(total__icontains=term))

    # Combine filter conditions in one line
    bons = Bonlivraison.objects.filter(client_id=clientid).filter(q_objects)[:10]

    trs=''
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)" {"checked" if i.reglements.exists() else ""}></td></tr>'
    return JsonResponse({
        'trs':trs
    })


def searchclientfcs(request):
    clientid=request.GET.get('clientid')
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(salseman__name__iregex=term) | Q(facture_no__iregex=term) | Q(total__iregex=term))
    bons=Facture.objects.filter(client_id=clientid).filter(q_objects)[:10]
    print(bons)
    trs=''
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)" {"checked" if i.reglementsfc.exists() else ""}></td></tr>'
    return JsonResponse({
        'trs':trs
    })

# search fc in update reglement
def searchclientfcsupdatereg(request):
    clientid=request.GET.get('clientid')
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(salseman__name__iregex=term) | Q(facture_no__iregex=term) | Q(total__iregex=term))
    bons=Facture.objects.filter(client_id=clientid).filter(q_objects)[:10]
    print(bons)
    trs=''
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'trs':trs
    })

def searchclientblsupdatereg(request):
    clientid=request.GET.get('clientid')
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(salseman__name__iregex=term) | Q(bon_no__iregex=term) | Q(total__iregex=term))
    bons=Bonlivraison.objects.filter(client_id=clientid).filter(q_objects)[:10]
    print(bons, clientid)
    trs=''
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'trs':trs
    })



def searchregl(request):
    term=request.GET.get('term')
    year=request.GET.get('year')

    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        q_objects &= (Q(client__name__icontains=term) | Q(client__code__icontains=term) | Q(mode__icontains=term) | Q(npiece__icontains=term) | Q(amount__icontains=term))
    regls=PaymentClientbl.objects.filter(date__year=year).filter(q_objects).order_by('-echance')[:50]
    return JsonResponse({
        'trs':render(request, 'reglbllist.html', {'bons':regls}).content.decode('utf-8'),

    })


def searchreglfc(request):
    thisyear=timezone.now().year
    term=request.GET.get('term')
    year=request.GET.get('year')
    q_objects = Q()
    if term=='':
        if year=='0':
            regls=PaymentClientfc.objects.filter(date__year=thisyear).filter(q_objects).order_by('echance')[:50]
        else:
            regls=PaymentClientfc.objects.filter(date__year=year).filter(q_objects).order_by('echance')[:50]
        return JsonResponse({
            'trs':render(request, 'reglfclist.html', {'bons':regls, 'today':timezone.now().date()}).content.decode('utf-8'),
        })

    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &

    for term in search_terms:
        q_objects &= (Q(client__name__iregex=term) | Q(client__code__iregex=term) | Q(mode__iregex=term) | Q(npiece__iregex=term) | Q(amount__iregex=term))
    regls=PaymentClientfc.objects.filter(date__year=year).filter(q_objects).order_by('echance')[:50]
    return JsonResponse({
        'trs':render(request, 'reglfclist.html', {'bons':regls, 'today':timezone.now().date()}).content.decode('utf-8'),

    })

def brahim(request):
    # if post method
    if request.method=='POST':
        # get user and password
        username=request.POST.get('username')
        password=request.POST.get('password')
        # check if user exist
        user=authenticate(username=username, password=password)
        if user:
            group=user.groups.all().first().name
            if group == 'admin':
                login(request, user)
                return redirect('products:system')
    if request.user.groups.all():
        group=request.user.groups.all().first().name
        if group == 'admin':
            return redirect('products:system')
    return render(request, 'brahim.html')

def yeardatareglfc(request):
    year=request.GET.get('year')
    regls=PaymentClientfc.objects.filter(date__year=year).order_by('-date')[:50]
    return JsonResponse({
        'trs':render(request, 'reglfclist.html', {'bons':regls, 'today':timezone.now().date(), 'year':{year}}).content.decode('utf-8'),
    })
def yeardatareglbl(request):
    year=request.GET.get('year')
    regls=PaymentClientbl.objects.filter(date__year=year).order_by('-echance')[:50]
    return JsonResponse({
        'trs':render(request, 'reglbllist.html', {'bons':regls, 'today':timezone.now().date(), 'year':{year}}).content.decode('utf-8'),
    })

def filterbcdate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    print(startdate, enddate)
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')+ timedelta(days=1)
    bons=Order.objects.filter(date__range=[startdate, enddate]).order_by('-id')[:50]
    trs=''
    # for i in bons:
    #     trs+=f'''
    #     <tr class="ord orderrow {'text-danger' if not i.isdelivered else ''}"  startdate={startdate} enddate={enddate} orderid="{i.id}" ondblclick="createtab('command{i.id}', 'Commande {i.order_no}', '/products/boncommandedetails/{i.id}')">
    #         <td>{ i.order_no }</td>
    #         <td>{ i.date.strftime('%d/%m/%Y') }</td>
    #         <td>{ i.client.name if i.client else '' }</td>
    #         <td>{ i.client.code if i.client else '' }</td>
    #         <td>{ i.total}</td>
    #         <td>{ i.client.region if i.client else '' }</td>
    #         <td>{ i.client.city if i.client else '' }</td>
    #         <td>{ i.client.soldbl if i.client else '' }</td>
    #         <td>{ i.salseman }</td>
    #         <td>
    #         {'Non' if i.isclientcommnd else 'Oui'}

    #         </td>
    #         <td>
    #           {'R1' if i.isdelivered else 'R0'}
    #         </td>
    #       </tr>
    #     '''
    print('>>>>>', bons)
    ctx={
        'bons':render(request, 'bclist.html', {'bons':bons, 'loadmore':True, 'startdate':request.GET.get('startdate'), 'enddate':request.GET.get('enddate')}).content.decode('utf-8')
    }
    if bons:
        ctx['total']=round(Order.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total')).get('total__sum'), 2)
    return JsonResponse(ctx)


def deletebonachat(request):
    # manual bon achat should never be deleted using this function
    bon=Itemsbysupplier.objects.get(pk=request.GET.get('id'))
    bontotal=bon.total
    bonsupplier=bon.supplier
    # supplier sold
    bonsupplier.rest=round(float(bonsupplier.rest)-float(bontotal), 2)
    # supplier total trunsactions
    bonsupplier.total=round(float(bonsupplier.total)-float(bontotal), 2)
    bonsupplier.save()
    items=Stockin.objects.filter(nbon=bon)
    for i in items:
        print(i.product.ref, int(i.product.stocktotal)-int(i.quantity))
        i.product.stocktotal=int(i.product.stocktotal)-int(i.quantity)

        if bon.isfacture:
            print('>> bon is facture')
            i.product.stockfacture=int(i.product.stockfacture)-int(i.quantity)
        i.product.save()

    items.delete()
    bon.delete()
    return JsonResponse({
        'success':True
    })



def searchforjv(request):
    term=request.GET.get('term')
    year=request.GET.get('year')
    print('term, year', term, year)
    # Split the term into individual words separated by '*'
    search_terms = term.split('+')
    # Create a list of Q objects for each search term and combine them with &
    print('getting corr')
    q_objects = Q()
    for term in search_terms:
        print('>> term', term)
        q_objects &= (Q(bon__client__name__icontains=term)|Q(bon__client__code__icontains=term)|Q(ref__icontains=term)|Q(name__icontains=term)|Q(total__icontains=term)|Q(bon__bon_no__icontains=term)|Q(bon__salseman__name__icontains=term))
    print('>> with year')
    searched=Livraisonitem.objects.filter(isfacture=False, date__year=year, isinventaire=False).filter(q_objects)
    products = searched.order_by('-date')[:50]
    total=round(searched.aggregate(Sum('total'))['total__sum'] or 0, 2)
    totalqty=searched.aggregate(Sum('qty'))['qty__sum'] or 0
    print('>> ', products.count())
    # trs=''
    # for i in products:
    #     trs+=f'''
    #     <tr class="journalvente-row" year={year} term={term}>
    #         <td>{i.date.strftime('%d/%m/%Y')}</td>
    #         <td>{i.bon.bon_no}</td>
    #         <td>{i.product.ref.upper()}</td>
    #         <td>{i.product.name}</td>
    #         <td>{i.price}</td>
    #         <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
    #         <td style="color:blue" class="coutmoyenjv">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
    #         <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
    #         <td class="text-danger qtyjv">{i.qty}</td>
    #         <td class="totaljv">{i.total}</td>
    #         <td>{i.bon.client.name}</td>
    #         <td>{i.bon.client.code}</td>
    #         <td>{i.bon.salseman.name}</td>
    #         <td class="text-success margejv">
    #
    #         </td>
    #     </tr>
    #     '''

    return JsonResponse({
        'trs':render(request, 'journalventtrs.html', {'year':year, 'term':term, 'products':products}).content.decode('utf-8'),
        'total':total,
        'totalqty':totalqty
    })

def searchforjvfc(request):
    thisyear=timezone.now().year
    term=request.GET.get('term')
    year=request.GET.get('year')

    # Split the term into individual words separated by '*'
    search_terms = term.split('+')
    # Create a list of Q objects for each search term and combine them with &

    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(client__name__iregex=term)|Q(client__code__iregex=term)|Q(ref__iregex=term)|Q(name__iregex=term)|Q(total__iregex=term)|Q(facture__facture_no__iregex=term)|Q(facture__salseman__name__iregex=term))
    if year=='0':
        # means the year i not selected, so the records of the current year
        products = Outfacture.objects.filter(date__year=thisyear).filter(q_objects).order_by('-date')[:50]
        total=round(Outfacture.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Outfacture.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
    else:
        products = Outfacture.objects.filter(date__year=year).filter(q_objects).order_by('-date')[:50]
        total=round(Outfacture.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=round(Outfacture.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0, 2)
    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalventefc-row" year={year} term={term}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.facture.facture_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjvfc">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger prachatjvfc">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjvfc">{i.qty}</td>
            <td class="">{round(i.product.prixnet*i.qty, 2)}</td>
            <td class="totalmoyenjvfc"></td>
            <td>{i.facture.client.name}</td>
            <td>{i.facture.client.code}</td>
            <td>{i.facture.salseman.name}</td>
            <td class="text-success margejvfc">

            </td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':total,
        'totalqty':totalqty
    })

def filterjvdate(request):
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Livraisonitem.objects.filter(isfacture=False, date__range=[startdate, enddate]).order_by('-date')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="journalvente-row" startdate={startdate} enddate={enddate}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.bon.bon_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjv">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjv">{i.qty}</td>
            <td class="totaljv">{i.total}</td>
            <td>{i.bon.client.name}</td>
            <td>{i.bon.salseman.name}</td>
            <td class="text-success margejv">

            </td>
        </tr>
        '''
    ctx={
        'trs':trs
    }
    if bons:
        ctx['total']=round(Livraisonitem.objects.filter(isfacture=False, date__range=[startdate, enddate]).aggregate(Sum('total')).get('total__sum'), 2)
        ctx['qty']=round(Livraisonitem.objects.filter(isfacture=False, date__range=[startdate, enddate]).aggregate(Sum('qty')).get('qty__sum'), 2)
    return JsonResponse(ctx)

def filterjvfcdate(request):
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Outfacture.objects.filter(date__range=[startdate, enddate]).order_by('-date')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="journalventefc-row" startdate={startdate} enddate={enddate}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.facture.facture_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjvfc">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjvfc">{i.qty}</td>
            <td class="totaljvfc">{i.total}</td>
            <td></td>
            <td>{i.facture.client.name}</td>
            <td>{i.facture.salseman.name}</td>
            <td class="text-success margejvfc">

            </td>
        </tr>
        '''
    ctx={
        'trs':trs,
    }
    if bons:
        ctx['total']=round(Outfacture.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'], 2)
        ctx['totalqty']=Outfacture.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('qty')).get('qty__sum')
    return JsonResponse(ctx)

def filterjachdate(request):
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Stockin.objects.filter(isinventaire=False, date__range=[startdate, enddate]).order_by('-date')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="journalacha-row" startdat={startdate} enddate={enddate}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''
    ctx={
        'trs':trs,
    }
    if bons:
        ctx['total']=round(Stockin.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'], 2)
        ctx['totalqty']=Stockin.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('quantity'))['quantity__sum']
    return JsonResponse(ctx)

def searchforjach(request):
    term=request.GET.get('term')
    year=request.GET.get('year')

    # Split the term into individual words separated by '*'
    search_terms = term.split('+')
    # Create a list of Q objects for each search term and combine them with &

    q_objects = Q()
    q_forhistory = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(supplier__name__iregex=term)|Q(product__ref__iregex=term)|Q(product__name__iregex=term)|Q(total__iregex=term)|Q(nbon__nbon__iregex=term))
            q_forhistory &= (Q(fournisseur__iregex=term)|Q(designation__iregex=term)|Q(ref__iregex=term)|Q(mantant__iregex=term))
    # means the year i not selected, so the records of the current year
    products1 = Stockin.objects.filter(isinventaire=False).filter(q_objects)
    producthistory = Achathistory.objects.filter(q_forhistory)
    products = chain(*[
        ((bon, 0) for bon in products1),
        ((b, 1) for b in producthistory)
    ])
    products=sorted(products, key=lambda item: item[0].date)
    total=round(products1.aggregate(Sum('total'))['total__sum'] or 0, 2)+round(producthistory.aggregate(Sum('total'))['total__sum'] or 0, 2)
    totalqty=(producthistory.aggregate(Sum('quantity'))['quantity__sum'] or 0)+(products1.aggregate(Sum('quantity'))['quantity__sum'] or 0)
    # if year=='0':
    #     print('thisyear')
    #     # means the year i not selected, so the records of the current year
    #     products = Stockin.objects.filter(date__year=thisyear).filter(q_objects).order_by('-date')
    #     products = Achathistory.objects.filter(date__year=thisyear).filter(q_objects).order_by('-date')
    #     total=round(Stockin.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
    #     totalqty=Stockin.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('quantity'))['quantity__sum'] or 0
    # else:
    #     products = Stockin.objects.filter(date__year=year).filter(q_objects).order_by('-date')
    #     products = Achathistory.objects.filter(date__year=year).filter(q_objects).order_by('-date')
    #     total=round(Stockin.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
    #     totalqty=Stockin.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('quantity'))['quantity__sum'] or 0
    trs=''
    print('his>>>', products)
    for i in products:
        print(i[0])
        # trs+=f'''
        # <tr class="jach-row" year={year} term={term}>
        #     <td>{i.date.strftime('%d/%m/%Y')}</td>
        #     <td>{i.product.ref}</td>
        #     <td>{i.product.name}</td>
        #     <td>{i.price}</td>
        #     <td>{i.supplier.name}</td>
        #     <td>{i.devise}</td>
        #     <td class="qtyjournalachat">{i.quantity}</td>
        #     <td class="totaljournalachat">{i.total}</td>
        # </tr>
        # '''
        trs+=f'''
        <tr style="background:{'#ffd79c'if i[1]==1 else ''};"">
            <td>{i[0].date.strftime('%d/%m/%Y')}</td>
            <td>{i[0].product.ref if i[1]==0 else i[0].ref}</td>
            <td>{i[0].product.name if i[1]==0 else i[0].designation}</td>
            <td>{i[0].price if i[1]==0 else i[0].prixunitaire}</td>
            <td>{i[0].supplier.name if i[1]==0 else i[0].fournisseur}</td>
            <td>{i[0].devise}</td>
            <td class="qtyjournalachat">{i[0].quantity}</td>
            <td class="totaljournalachat">{i[0].total}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':total,
        'totalqty':totalqty
    })

def yeardatajach(request):
    year=request.GET.get('year')
    print(year)
    items=Stockin.objects.filter(isinventaire=False, date__year=year).order_by('-date')[:50]
    trs=''
    totalmarge=0
    for i in items:
        try:
            marge_value = round((i.product.prixnet - (i.product.coutmoyen if i.product.coutmoyen else 0)) * i.qty, 2)
        except:
            marge_value = 0
        totalmarge+=marge_value
        trs+=f'''
        <tr class="journalacha-row" year={year}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''

    return JsonResponse({
        'trs':trs,
        'totalqty':Stockin.objects.filter(date__year=year).aggregate(Sum('quantity'))['quantity__sum'] or 0,
        'total':round(Stockin.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalmarge':round(totalmarge, 2)
    })

def filterjachfcdate(request):
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Stockin.objects.filter(facture=True, date__range=[startdate, enddate]).order_by('-date')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="jachfc-row" startdat={startdate} enddate={enddate}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''
    ctx={
        'trs':trs,
    }
    if bons:
        ctx['total']=round(Stockin.objects.filter(facture=True, date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'], 2)
        ctx['totalqty']=Stockin.objects.filter(facture=True, date__range=[startdate, enddate]).aggregate(Sum('quantity'))['quantity__sum']
    return JsonResponse(ctx)

def searchforjachfc(request):
    thisyear=timezone.now().year
    term=request.GET.get('term')
    year=request.GET.get('year')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    # Create a list of Q objects for each search term and combine them with &

    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(supplier__name__iregex=term)|Q(product__ref__iregex=term)|Q(product__name__iregex=term)|Q(total__iregex=term)|Q(nbon__nbon__iregex=term))
    # if year=='0':
    #     print('thisyear')
    #     # means the year i not selected, so the records of the current year
    #     products = Stockin.objects.filter(facture=True, date__year=thisyear).filter(q_objects).order_by('-date')[:50]
    #     total=round(Stockin.objects.filter(facture=True, date__year=thisyear).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
    #     totalqty=Stockin.objects.filter(facture=True, date__year=thisyear).filter(q_objects).aggregate(Sum('quantity'))['quantity__sum'] or 0
    # else:
    #     products = Stockin.objects.filter(facture=True, isinventaire=False, date__year=year).filter(q_objects).order_by('-date')
    #     total=round(Stockin.objects.filter(facture=True, isinventaire=False, date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
    #     totalqty=Stockin.objects.filter(facture=True, date__year=year).filter(q_objects).aggregate(Sum('quantity'))['quantity__sum'] or 0


    products = Stockin.objects.filter(facture=True, isinventaire=False).filter(q_objects).order_by('-date')
    total=round(Stockin.objects.filter(facture=True, isinventaire=False).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
    totalqty=Stockin.objects.filter(facture=True).filter(q_objects).aggregate(Sum('quantity'))['quantity__sum'] or 0
    trs=''
    for i in products:
        trs+=f'''
        <tr class="jachfc-row" year={year} term={term}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':total,
        'totalqty':totalqty
    })

def yeardatajachfc(request):
    year=request.GET.get('year')
    print(year)
    items=Stockin.objects.filter(facture=True, date__year=year).order_by('-id')[:50]
    trs=''
    totalmarge=0
    for i in items:
        try:
            marge_value = round((i.product.prixnet - (i.product.coutmoyen if i.product.coutmoyen else 0)) * i.qty, 2)
        except:
            marge_value = 0
        totalmarge+=marge_value
        trs+=f'''
        <tr class="jachfc-row journalachafc-row" year={year}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{round(i.total, 2)}</td>
        </tr>
        '''

    return JsonResponse({
        'trs':trs,
        'totalqty':Stockin.objects.filter(facture=True, date__year=year).aggregate(Sum('quantity'))['quantity__sum'] or 0,
        'total':round(Stockin.objects.filter(facture=True, date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalmarge':round(totalmarge, 2)
    })

def updaterepdata(request):
    region=request.POST.get('region')
    caneditprice=True if request.POST.get('caneditprice')=='on' else False
    slides=True if request.POST.get('slides')=='on' else False
    repid=request.POST.get('repid')
    data={
    'repid':repid,
    'caneditprice':caneditprice
    }
    try:
        res=req.get(f'http://{serverip}/products/updaterepdata', data)
        res.raise_for_status()
    except req.exceptions.RequestException as e:
        # in case connection failed
        return JsonResponse({
        'success':False,
        'here':'rr'
        })
    print(region, caneditprice, slides, repid)
    rep=Represent.objects.get(pk=repid)
    rep.region=region
    rep.caneditprice=caneditprice
    rep.slides=slides
    rep.save()

    return JsonResponse({
        'success':True,
        'here':'rr'

    })

def makebondelivered(request):
    id=request.GET.get('id')
    bon=Bonlivraison.objects.get(pk=id)
    bon.isdelivered=True
    bon.save()
    return JsonResponse({
        'success':True
    })

def getitemsforlistbl(request):
    term=request.GET.get('term')
    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (
                Q(ref__icontains=term) |
                Q(name__icontains=term) |
                Q(mark__name__icontains=term) |
                Q(category__name__icontains=term) |
                Q(equivalent__icontains=term) |
                Q(refeq1__icontains=term) |
                Q(refeq2__icontains=term) |
                Q(refeq3__icontains=term) |
                Q(refeq4__icontains=term) |
                Q(diametre__icontains=term)|
                Q(cars__icontains=term)
            )
    # check if term in product.ref or product.name
    products=Produit.objects.filter(q_objects).order_by('-stocktotal')
    brands = [product.mark for product in products]
    categories = [product.category for product in products]
    print('>>>>', brands)
    unique_categories = set(categories)
    unique_brands = set(brands)
    brands = [{'id': mark.id if mark.id else None, 'name': mark.name if mark.name else None, 'image':mark.image.url if mark.image else '/media/default.png'} for mark in unique_brands]
    categories = [{'id': category.id, 'name': category.name, 'image':category.image.url if category.image else '/media/default.png'} for category in unique_categories]
    trs=[f'''<tr class="productsbrand{i.mark.id if i.mark else ''}">
    <td><img style="width:50%" src={i.image.url if i.image else ''}></td>
    <td>{i.ref.upper()}</td>
    <td>{i.name.upper()}</td>
    <td style="color: #ff6409;
    font-weight: bold;">{i.stocktotal}</td>
    <td style="color:blue;font-weight: bold;">{i.sellprice}</td>
    <td>{i.remise}</td>
    <td>{i.prixnet}</td>
    <td>{i.diametre}</td>
    </tr>
    ''' for i in products]
    return JsonResponse({
        'trs':trs,
        'brands':brands,
        'categories':categories
    })
def refspage(request):
    res=req.get(f'http://{serverip}/products/refspage')
    print(res)

    refs=Refstats.objects.all().order_by('-lastdate')
    return render(request, 'refspage.html', {'refs':refs, 'refserver':json.loads(res.text)['refs']})

def updateadmindata(request):
    from django.contrib.auth.hashers import make_password
    username=request.POST.get('adminusername')
    password=request.POST.get('adminpassword')
    # Check if the username already exists
    existing_user = User.objects.filter(username=username).exclude(id=request.user.id).first()
    print('>>>>>>>',existing_user)
    if existing_user:
        return JsonResponse({
            'success': False,
            'error': 'Username already exists.'
        })
    else:
        # Update user data
        print('>>> updating')
        user = User.objects.get(pk=request.user.id)
        user.username = username
        user.password = make_password(password)  # Ensure the password is securely hashed
        user.save()

        return JsonResponse({
            'success': True,
        })

def notavailable(request):
    ctx={
    'products':Notavailable.objects.all()
    }
    return render(request, 'notavailable.html', ctx)

def cartpage(request):
    res=req.get(f'http://{serverip}/products/getcarts')

    ctx={
        'carts':Cart.objects.all().order_by('-total').exclude(total=0),
        'cartsserver':list(json.loads(res.text)['carts'])
    }
    return render(request, 'cartspage.html', ctx)

def reliquatpage(request):
    res=req.get(f'http://{serverip}/products/getwishs')
    print(list(json.loads(res.text)['carts']))
    ctx={
        'carts':Wich.objects.all().order_by('-total'),
        'wishserver':list(json.loads(res.text)['carts'])
    }
    return render(request, 'reliquatpage.html', ctx)

def excelnotav(request):
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    entries=0
    for d in df.itertuples():
        try:
            ref = d.ref.lower().strip()
        except:
            ref=d.ref
        #reps=json.dumps(d.rep)
        name = d.name
        mark = None if pd.isna(d.mark) else d.mark
        #refeq = '' if pd.isna(d.refeq) else d.refeq
        #status = False if pd.isna(d.status) else True
        #coderef = '' if pd.isna(d.code) else d.code
        #diam = '' if pd.isna(d.diam) else d.diam
        #qty = 0 if pd.isna(d.qty) else d.qty
        #buyprice = 0 if pd.isna(d.buyprice) else d.buyprice
        #devise = 0 if pd.isna(d.devise) else d.devise
        # car = None if pd.isna(d.car) else d.car
        #prixbrut = 0 if pd.isna(d.prixbrut) else d.prixbrut
        #ctg = None if pd.isna(d.ctg) else d.ctg
        #order = '' if pd.isna(d.order) else d.order
        #img = None if pd.isna(d.img) else d.img
        #remise = 0 if pd.isna(d.remise) else d.remise
        #prixnet=0 if pd.isna(d.prixnet) else d.prixnet
        product=Produit.objects.filter(ref=ref).first()
        if not product:

            exis=Notavailable.objects.filter(ref=ref).first()
            if not exis:
                product=Notavailable.objects.create(
                    ref=ref,
                #     equivalent=refeq,
                #     isactive=False,
                #     coderef=coderef,
                    name=name,
                    mark_id=mark,
                #     sellprice=prixbrut,
                #     prixnet=prixnet,
                #     remise=remise,
                #     category_id=ctg,
                    image=f'/products_imags/tette.jpg',
                #     stockfacture=qty,
                #     #diametre=diam,
                #     buyprice=buyprice,
                #     devise=devise
                )
            else:
                print(ref, 'exist in hors stock')
        else:
            print(ref, 'exist in products')

    return JsonResponse({
        'success':True
    })

def deletereglbl(request):
    reglid=request.GET.get('reglid')
    password=request.GET.get('password')
    regl=PaymentClientbl.objects.get(pk=reglid)
    client=regl.client
    if password=='0000':
        # client.soldbl=round(float(client.soldbl)+float(regl.amount))
        client.soldbl=client.mehtodsoldbl
        client.soldtotal=client.mehtodsoldbl+float(regl.amount)
        client.save()
        bons=regl.bons.all()
        for i in bons:
            otherregl=PaymentClientbl.objects.filter(bons=i.id).exclude(pk=reglid)
            if not otherregl.exists():
                print(i.bon_no, 'not mentioned so it will be paid=false')
                i.ispaid=False
                i.statusreg='n1'
                i.save()
        regl.delete()
    return JsonResponse({
        'success':True
    })

def deletereglsupp(request):
    reglid=request.GET.get('reglid')
    password=request.GET.get('password')
    # get reglement supp
    regl=PaymentSupplier.objects.get(pk=reglid)
    #get supplier
    supp=regl.supplier
    print(reglid, password, regl.amount, supp)
    if password=='0000':
        # add amount to supplier sold
        print('adding amount to supplier sold')
        supp.rest+=regl.amount
        supp.save()
        bons=regl.bons.all()
        for i in bons:
            otherregl=PaymentSupplier.objects.filter(bons=i.id).exclude(pk=reglid)
            if not otherregl.exists():
                print('bon not exist')
                i.ispaid=False
                i.save()
        regl.delete()
    return JsonResponse({
        'success':True
    })



def deletereglfc(request):
    reglid=request.GET.get('reglid')
    password=request.GET.get('password')
    regl=PaymentClientfc.objects.get(pk=reglid)
    print('reglem', regl.client.name, regl.amount)
    client=regl.client
    if password=='0000':
        client.soldfacture=round(float(client.soldfacture)+float(regl.amount))
        client.soldtotal=round(float(client.soldtotal)+float(regl.amount))
        client.save()
        bons=regl.factures.all()
        for i in bons:
            # search in other reglement
            otherregl=PaymentClientfc.objects.filter(factures=i.id).exclude(pk=reglid)

            if not otherregl.exists():
                i.ispaid=False
                i.statusreg='n1'
                i.save()
        regl.delete()
    return JsonResponse({
        'success':True
    })
def changeclientbl(request):
    oldclient=request.GET.get('oldclient')
    orderid=request.GET.get('orderid')
    newclient=request.GET.get('newclient')
    total=request.GET.get('total')
    oldclient=Client.objects.get(pk=oldclient)
    newclient=Client.objects.get(pk=newclient)
    order=Bonlivraison.objects.get(pk=orderid)
    order.client=newclient
    order.save()
    print(oldclient.name, newclient.name, total)
    #extract total from soldbl  and soldtotal of old client
    oldclient.soldbl=float(oldclient.soldbl)-float(total)
    oldclient.soldtotal=float(oldclient.soldtotal)-float(total)
    oldclient.save()
    # add total to soldbl and soldtotal of newclient
    newclient.soldbl=float(newclient.soldbl)+float(total)
    newclient.soldtotal=float(newclient.soldtotal)+float(total)
    newclient.save()
    return JsonResponse({
        'success':True
    })

def comptable(request):
    factureid=request.GET.get('factureid')
    facture=Facture.objects.get(pk=factureid)
    facture.isaccount=not facture.isaccount
    facture.save()
    return JsonResponse({
        'success':True
    })

def deletsearchedref(request):
    refid=request.GET.get('refid')
    ref=Refstats.objects.get(pk=refid)
    ref.delete()
    return JsonResponse({
        'success':True
    })
def getcompatbilse(request):
    year=request.GET.get('year') or '2024'
    print('>>>>>>>', year, Facture.objects.filter(date__year=year, isaccount=True))
    factures=Facture.objects.filter(date__year=year, isaccount=True).order_by('-facture_no')[:50]
    total=round(Facture.objects.filter(date__year=year, isaccount=True).aggregate(Sum('total'))['total__sum'] or 0, 2)
    totaltva=round(Facture.objects.filter(date__year=year, isaccount=True).aggregate(Sum('tva'))['tva__sum'] or 0, 2)
    trs=''
    for i in factures:
        trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''}
             fc-row"
                style="color:{"blue" if i.bon else ""} " comptable="1" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td >
                    {i.note}
                </td>

                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
                <td>
                <i class="bi {"bi-check" if i.isaccount else ''} h3"></i>{"c" if i.isaccount else ''}
                </td>
            </tr>
            '''
    return JsonResponse({
        'trs':trs,
        'total':total,
        'totaltva':totaltva,
    })

def listnotifications(request):
    ctx={
        'title':'List Notifications',
        'notifications':Notification.objects.all()
    }
    return render(request, 'listnotifications.html', ctx)

def addnotification(request):
    notification=request.GET.get('notification')
    try:
        req.get(f'http://{serverip}/products/addnotification', {'notificationid':notificationid,'notification':notification})

        Notification.objects.create(notification=notification)
        return JsonResponse({
            'success':True
            })
    except:
        print('>>>>>>> error')
        return JsonResponse({
            'success':False
            })

def updatenotification(request):
    notificationid=request.GET.get('notificationid')
    notification=request.GET.get('notification')
    print(notificationid, notification)
    return JsonResponse({
        'success':True
    })
    # try:
    #     req.get(f'http://{serverip}/products/updatenotification', {'notificationid':notificationid,'notification':notification})
    #     notif=Notification.objects.get(pk=notificationid)
    #     notif.notification=notification
    #     notif.save()
    #     return JsonResponse({
    #         'success':True
    #     })
    # except:
    #     print('>>>>>>> error')
    #     return JsonResponse({
    #         'success':False
    #         })

def updatefacturenote(request):
    factureid=request.GET.get('factureid')
    note=request.GET.get('note')
    facture=Facture.objects.get(pk=factureid)
    facture.note=note
    facture.save()
    return JsonResponse({
        'success':True
    })

# update facture client
def updatefactureclient(request):
    factureid=request.GET.get('factureid')
    clientid=request.GET.get('clientid')
    print('>>> factureid, clientid', factureid, clientid)
    facture=Facture.objects.get(pk=factureid)
    # get current client
    thisclient=facture.client
    print('>>> thisclient, thisclient.soldtotal, facture.total', thisclient, thisclient.soldtotal, facture.total)
    # substract total from sold facture of current client
    thisclient.soldtotal=round(thisclient.soldtotal-facture.total, 2)
    thisclient.soldfacture=round(thisclient.soldfacture-facture.total, 2)
    print(' >>> thisclient.soldtotal', thisclient.soldtotal)
    print(' >>> thisclient.soldfacture', thisclient.soldfacture)
    # save client
    thisclient.save()
    # get new client
    newclient=Client.objects.get(pk=clientid)
    print('>>> newclient', newclient)
    # add total from sold facture of new client
    newclient.soldtotal=round(newclient.soldtotal+facture.total, 2)
    newclient.soldfacture=round(newclient.soldfacture+facture.total, 2)
    print(' >>> newclient.soldtotal', newclient.soldtotal)
    print(' >>> newclient.soldfacture', newclient.soldfacture)
    # save new client
    newclient.save()
    # assign new client to facture
    facture.client=newclient
    # save facture
    facture.save()
    return JsonResponse({
        'success':True
    })

def getfacturedata(request):
    facture_no=request.GET.get('facture_no')
    input=request.GET.get('input')
    total=0
    for i in input.split(','):
        f=Facture.objects.get(facture_no=f'FC{i.strip()}')
        print('>> ', f.client.name)
        total+=f.total
    print('>> ', input.split(','), 'total: ', total)
    try:
        facture=Facture.objects.get(facture_no=f'FC{facture_no}')
        print('>>> facture found', facture)
        # this used to exit and not processding when client is not diver, means when facture is already assigned with a client => no, because the client is not showing in the select item => fixed it , now it's showing the client
        # if not facture.client.id==3731:
        #     return JsonResponse({
        #     'success':False,
        #     'error':"Facture deja modifier"
        #     })
        return JsonResponse({
        'success':True,
        'facture_no':facture.facture_no,
        'date':facture.date.strftime('%d/%m/%Y'),
        'client':facture.client.name,
        'client_id':facture.client.id,
        'total':facture.total,
        'tva':facture.tva,
        'id':facture.id,
        'note':facture.note,
        'salseman':facture.salseman.name,
        })
    except Exception as e:
        print('>> ', e)
        return JsonResponse({
        'success':False,
        'error':"Numero n'exist pas"
        })

def updatefacturerep(request):
    factureid=request.GET.get('factureid')
    rep=request.GET.get('rep')
    facture=Facture.objects.get(pk=factureid)
    facture.salseman_id=rep
    facture.save()
    return JsonResponse({
        'success':True
        })

def updateproductstock(request):
    stock=request.GET.get('stock')
    productid=request.GET.get('productid')
    product=Produit.objects.get(pk=productid)
    diff=int(stock)-int(product.stocktotal)
    Modifierstock.objects.create(stock=diff, product=product)
    # req.get(f'http://{serverip}/products/updateproduct', {

    #     'id':productid,
    #     'ref':product.ref,
    #     'stocktotal':stock,
    # })
    product.stocktotal=stock
    product.save()
    return JsonResponse({
        'success':True
    })

def updateproductstockfc(request):
    stock=request.GET.get('stock')
    productid=request.GET.get('productid')
    product=Produit.objects.get(pk=productid)
    diff=int(stock)-int(product.stockfacture)
    Modifierstock.objects.create(stock=diff, product=product, stockfc=True)

    product.stockfacture=stock
    product.save()
    return JsonResponse({
        'success':True
    })

def updatetransportbl(request):
    transport=request.GET.get('transport')
    blid=request.GET.get('blid')
    bon=Bonlivraison.objects.get(pk=blid)
    bon.modlvrsn=transport
    bon.save()
    return JsonResponse({
        'success':True
        })
def updatenotebl(request):
    note=request.GET.get('note')
    blid=request.GET.get('blid')
    bon=Bonlivraison.objects.get(pk=blid)
    bon.note=note
    bon.save()
    return JsonResponse({
        'success':True
       } )
def updaterepbl(request):
    rep=request.GET.get('rep')
    blid=request.GET.get('blid')
    bon=Bonlivraison.objects.get(pk=blid)
    bon.salseman_id=rep
    bon.save()
    return JsonResponse({
        'success':True
        })

def stockupdated(request):
    ctx={
    'products':Modifierstock.objects.all()
    }
    return render(request, 'stockupdated.html', ctx)

def getclientcode(request):
    factureno=request.GET.get('factureno')
    facture=Facture.objects.get(facture_no__icontains=factureno)
    return JsonResponse({
        'code':facture.client.code
    })

def allowcatalog(request):
    clientcode=request.GET.get('clientcode')
    serverip= Setting.objects.only('serverip').first()
    serverip=serverip.serverip if serverip else None
    if serverip:
        res=req.get(f'http://{serverip}/products/allowcatalog', {
            'clientcode':clientcode,
        })
        if json.loads(res.text)['success']:
            return JsonResponse({
                'success':True
            })
        else:
            return JsonResponse({
                'success':False,
                'error':f"{json.loads(res.text)['error']}"
            })
    # user=User.objects.filter(username=username).first()
    # if user:
    return JsonResponse({
        'success':False,
        'error':'No server'
    })
    #     res.raise_for_status()
    #     print('>><W>>>>>>>>>>>>>>>>>')
    #     client=Client.objects.get(code=clientcode)
    #     client.accesscatalog=not client.accesscatalog
    #     client.save()
    #     return JsonResponse({
    #         'success':True
    #     })
    # except req.exceptions.RequestException as e:
    #     return JsonResponse({
    #         'success':False
    #     })

def filterepbons(request):
    repid=request.GET.get('repid')
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    bontarget=request.GET.get('bontarget')

    if bontarget=='1':
        bons=Bonlivraison.objects.filter(date__range=[startdate, enddate], salseman_id=repid, ispaid=True).order_by('-id')
        repbons=bons.filter(commande__isnull= False, commande__isclientcommnd=False)



        factures=Facture.objects.filter(date__range=[startdate, enddate], salseman_id=repid, ispaid=True).order_by('-id')
        repfactures=factures.filter(bon__commande__isnull= False, bon__commande__isclientcommnd=False)


    elif bontarget=='0':
        bons=Bonlivraison.objects.filter(date__range=[startdate, enddate], salseman_id=repid, ispaid=False).order_by('-id')
        repbons=bons.filter(commande__isnull= False, commande__isclientcommnd=False)



        factures=Facture.objects.filter(date__range=[startdate, enddate], salseman_id=repid, ispaid=False).order_by('-id')
        repfactures=factures.filter(bon__commande__isnull= False, bon__commande__isclientcommnd=False)

    else:
        bons=Bonlivraison.objects.filter(date__range=[startdate, enddate], salseman_id=repid).exclude(total=0).order_by('-id')
        # this gets only bons from tablete
        repbons=bons.filter(commande__isnull= False, commande__isclientcommnd=False)

        systembons=bons.exclude(pk__in=[bon.pk for bon in repbons])
        print('>>>>>>>>>>> system bons', systembons)
        factures=Facture.objects.filter(date__range=[startdate, enddate], salseman_id=repid).order_by('-id')
        # this gets only bons from tablete
        repfactures=factures.filter(bon__commande__isnull= False, bon__commande__isclientcommnd=False)
        systemfactures=factures.exclude(pk__in=[i.pk for i in repfactures])
    totalbl=bons.aggregate(Sum('total'))['total__sum'] or 0
    totalfc=factures.aggregate(Sum('total'))['total__sum'] or 0
    totalblfctable=round(totalbl+totalfc, 2)
    avoirs=Avoirclient.objects.filter(date__range=[startdate, enddate], representant_id=repid)
    systemtotalblfc=round(systembons.aggregate(Sum('total'))['total__sum'] or 0, 2)+round(systemfactures.aggregate(Sum('total'))['total__sum'] or 0, 2)
    reptotalblfc=round(repbons.aggregate(Sum('total'))['total__sum'] or 0, 2)+round(repfactures.aggregate(Sum('total'))['total__sum'] or 0, 2)
    totalavoirs=round(avoirs.aggregate(Sum('total'))['total__sum'] or 0, 2)
    systemresultttc=systemtotalblfc-totalavoirs
    represultttc=reptotalblfc-totalavoirs
    systemresultht=round(systemtotalblfc/1.2, 2)
    #systemresultht=round(systemresultttc/1.2, 2)
    #represultht=round(represultttc/1.2, 2)
    represultht=round(represultttc/1.2, 2)
    print('>>>', bons)
    trsbons=[f"""
    <tr
    style="background: {"lightgreen;" if i.isdelivered else ""} color:{"blue" if i.isfacture else ""} "
    class="ord {"text-danger" if i.ispaid else ''}"
    orderid="{i.id}">
        <td>{ i.bon_no }</td>
            <td>{ i.date.strftime("%d/%m/%Y")}</td>
            <td>{ i.client.name }</td>
            <td>{ i.client.code }</td>
            <td style="color: blue;">{ i.total}</td>
            <td>{ i.client.region}</td>
            <td>{ i.client.city}</td>
            <td>{ "%.2f" % i.client.soldbl} </td>
            <td>{ i.salseman }</td>
            <td class="d-flex justify-content-between">
            <div>
            {'R0' if i.ispaid else 'N1' }

            </div>
            <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

            </td>
            <td>
            {'OUI' if i.isfacture else 'NON'}
            </td>

            <td>
                {i.commande.order_no if i.commande else "---"}
            </td>
            <td>
            {i.modlvrsn}
            </td>
            <td>
            {i.note}
            </td>
            <td>

            </td>

      </tr>
    """ for i in bons]
    trsfactures=[f"""
    <tr class="ord {"text-danger" if i.ispaid else ''}">
        <td>{ i.facture_no }</td>
        <td>{ i.date.strftime("%d/%m/%Y")}</td>
        <td>{ i.client.name }</td>
        <td>{ i.client.code }</td>
        <td>{ i.total}</td>
        <td>{ i.client.region}</td>
        <td>{ i.client.city}</td>
        <td>{ i.client.soldfacture:.2f}</td>
        <td>{ i.salseman }</td>
        <td class="d-flex justify-content-between">
        <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

        </td>
        <td></td>
        <td>
        {i.bon.bon_no if i.bon else "--"}
        </td>

        <td class="d-flex">

        </td>
        <td>
        {i.note}
        </td>
    </tr>
    """ for i in factures]
    return JsonResponse({
        'totalbons':round(systembons.aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalfactures':round(systemfactures.aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalblfc':systemtotalblfc,
        'totalavoirs':totalavoirs,
        'resultttc':systemtotalblfc,
        'resultht':systemresultht,

        'reptotalblfc':reptotalblfc,
        #'represultttc':represultttc,
        'represultttc':represultttc,
        #'represultht':represultht,
        'represultht':represultht,
        'totalrepbons':round(repbons.aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalrepfactures':round(repfactures.aggregate(Sum('total'))['total__sum'] or 0, 2),
        'bons':''.join(trsbons),
        'factures':''.join(trsfactures),
        'totalblfctable':totalblfctable
    })
def updateavoirnote(request):
    avoirid=request.GET.get('avoirid')
    note=request.GET.get('note')
    print('>>>>>>>', avoirid, note)
    avoir=Avoirclient.objects.get(pk=avoirid)
    avoir.note= note
    avoir.save()
    return JsonResponse({
        'success':True
    })

def getreliquatcommande(request):
    bons=Order.objects.filter(note__icontains='Reliquat', isdelivered=False)
    print(bons)
    return JsonResponse({
        'success':True,
        'trs':render(request, 'bclist.html', {"bons":bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total'))['total__sum'] or 0, 2)
    })
#this to indicate if the facture is given to client
def printed(request):
    factureid=request.GET.get('factureid')
    facture=Facture.objects.get(pk=factureid)
    facture.printed=True
    facture.save()
    return JsonResponse({
        'success':True
    })

def getnongenerer(request):
    bons=Order.objects.filter(isdelivered=False).exclude(note__icontains='Reliquat')
    return JsonResponse({
        'success':True,
        'trs':render(request, 'bclist.html', {"bons":bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total'))['total__sum'] or 0, 2)
    })

def addrepnote(request):
    note=request.GET.get('note')
    repid=request.GET.get('repid')
    repnote=Notesrepresentant.objects.filter(represent_id=repid).first()
    if repnote:
        repnote.note=note
        repnote.save()
    else:
        Notesrepresentant.objects.create(represent_id=repid, note=note)
    return JsonResponse({
        'success':True
    })

def alertreliquatcommande(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    orders=Orderitem.objects.filter(date__range=[startdate, enddate], order__note__icontains='Reliquat', product__stocktotal__gt=F('qty'), islivraison= False)
    bons=[i.order for i in orders]
    bons=set(bons)
    return JsonResponse({
        'success':True,
        'trs':render(request, 'bclist.html', {"bons":bons}).content.decode('utf-8'),

    })

def makeitemsent(request):
    orderitemid=request.GET.get('orderitemid')
    password=request.GET.get('password')
    if password=="1574":
        orderitem=Orderitem.objects.get(pk=orderitemid)
        orderitem.islivraison=True
        orderitem.save()
        return JsonResponse({
            'success':True
        })
    return JsonResponse({
        'success':False,
        'error':'Mot de pass incorrect'
    })

def updatedateavsupp(request):
    avoirid=request.GET.get('avoirsupid')
    date=request.GET.get('date')
    datebon=datetime.strptime(date, '%Y-%m-%d')
    avoirsupp=Avoirsupplier.objects.get(pk=avoirid)
    avoirsupp.date=datebon
    avoirsupp.save()
    print('>>>>', datebon, avoirsupp.date, avoirsupp.total)
    return JsonResponse({
        'success':True
    })

def tsgs(request):
    thisyear=timezone.now().year

    import random
    most_sold_items = Livraisonitem.objects.values('product__id', 'product__name').annotate(total_sold=Sum('qty')).order_by('-total_sold')[:10]
    most_sold_products = [(item['product__name'], item['total_sold']) for item in most_sold_items]
    for product_name, total_sold in most_sold_products:
        print(f">>* {''.join([random.sample(product_name, 1)[0] for i in range(len(product_name))])} {product_name}: {total_sold} units")
    sldsbl=round(PaymentClientbl.objects.aggregate(Sum('amount'))['amount__sum'] or 0, 2)
    sldsfc=round(PaymentClientfc.objects.aggregate(Sum('amount'))['amount__sum'] or 0, 2)
    sldsup=round(PaymentSupplier.objects.filter(date__year=thisyear).aggregate(Sum('amount'))['amount__sum'], 2)
    nt=sldsbl+sldsfc-sldsup

    # Annotate and aggregate to get the total per day
    totals_per_day = (Bonlivraison.objects
                      .annotate(day=TruncDay('date'))
                      .values('day')
                      .annotate(total=Sum('total'))
                      .order_by('-total'))

    # Get the day with the highest total
    if totals_per_day.exists():
        highest_total_day = totals_per_day.first()
        print(">>> Day with the highest total:", highest_total_day['day'], "with total:", highest_total_day['total'])
    else:
        print("No data available.")
    # Annotate and aggregate to get the total per client
    totals_per_client = (Bonlivraison.objects
                         .values('client')
                         .annotate(total=Sum('total'))
                         .order_by('-total'))

    # Get the client with the highest total
    if totals_per_client.exists():
        highest_total_client_data = totals_per_client.first()
        highest_total_client = Client.objects.get(id=highest_total_client_data['client'])
        print(">>> Client with the highest total:", highest_total_client, "with total:", highest_total_client_data['total'])
    else:
        print("No data available.")
    # Annotate and aggregate to get the total per product
    totals_per_product = (Livraisonitem.objects
                          .values('product')
                          .annotate(total=Sum('total'))
                          .order_by('-total'))

    # Get the product with the highest total
    if totals_per_product.exists():
        highest_total_product_data = totals_per_product.first()
        highest_total_product = Produit.objects.get(id=highest_total_product_data['product'])
        print(">>> Product with the highest total:", highest_total_product, "with total:", highest_total_product_data['total'])
    else:
        print("No data available.")
    # Annotate and aggregate to get the total per client
    totals_per_client = (Avoirclient.objects
                         .values('client')
                         .annotate(total=Sum('total'))
                         .order_by('-total'))

    # Get the client with the highest total
    if totals_per_client.exists():
        highest_total_client_data = totals_per_client.first()
        highest_total_client = Client.objects.get(id=highest_total_client_data['client'])
        print(">>> Client with the highest avoir total:", highest_total_client, "with total:", highest_total_client_data['total'])
    else:
        print("No data available.")
    print(nt)
    print(nt/8)
    return JsonResponse({
        'rr':'rr'
    })
def bonlivraisonprint(request, id):
    order=Bonlivraison.objects.get(pk=id)
    orderitems=Livraisonitem.objects.filter(bon=order, isfacture=False).order_by('product__name')
    reglements=PaymentClientbl.objects.filter(bons__in=[order])
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+41] for i in range(0, len(orderitems), 41)]
    ctx={
        'title':f'Bon de livraison {order.bon_no}',
        'order':order,
        'bon_no':order.bon_no.replace('BL', ''),
        'orderitems':orderitems,
        'reglements':reglements,
        'reps':Represent.objects.all()
    }
    return render(request, 'bonlivraisonprint.html', ctx)

def bonavoirprint(request, id):
    order=Avoirclient.objects.get(pk=id)
    orderitems=Returned.objects.filter(avoir=order).order_by('product__name')
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+33] for i in range(0, len(orderitems), 33)]
    ht=round(order.total/1.2, 2)
    tva=order.total-ht

    ctx={
        'title':f'Bon avoir {order.no}',
        'order':order,
        'orderitems':orderitems,
        'ht':ht,
        'tva':tva,
    }
    return render(request, 'bonavoirprint.html', ctx)

def boncmndprint(request, id):
    order=Order.objects.get(pk=id)
    orderitems=Orderitem.objects.filter(order=order).order_by('product__name')


    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+45] for i in range(0, len(orderitems), 45)]
    ctx={
        'title':f'Bon commande {order.order_no}',
        'order':order,
        'orderitems':orderitems,
    }
    return render(request, 'boncmndprint.html', ctx)

def factureprint(request, id):
    order=Facture.objects.get(pk=id)
    orderitems=Outfacture.objects.filter(facture=order).order_by('product__name')
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+30] for i in range(0, len(orderitems), 30)]

    ctx={
        'title':f'Facture {order.facture_no}',
        'facture':order,
        'orderitems':orderitems,
        'tva':order.tva,
        'ttc':order.total,
        'ht':round(order.total-order.tva, 2),
    }
    return render(request, 'factureprint.html', ctx)

def factureprintcopy(request, id):
    order=Facture.objects.get(pk=id)
    orderitems=Outfacture.objects.filter(facture=order).order_by('product__name')
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+30] for i in range(0, len(orderitems), 30)]

    ctx={
        'title':f'Bon livraison {order.facture_no}',
        'facture':order,
        'orderitems':orderitems,
        'tva':order.tva,
        'ttc':order.total,
        'ht':round(order.total-order.tva, 2),
    }
    return render(request, 'factureprintcopy.html', ctx)


def achatprint(request, id):
    bon=Itemsbysupplier.objects.get(pk=id)
    items=Stockin.objects.filter(nbon=bon)
    payments=PaymentSupplier.objects.filter(bons__in=[bon])
    orderitems=[items[i:i+36] for i in range(0, len(items), 36)]
    # ctx={
    #     'title':f'Bon de livraison {order.bon_no}',
    #     'order':order,
    #     'orderitems':orderitems,
    #     'reglements':reglements,
    #     'reps':Represent.objects.all()
    # }
    ctx={
        'title':f'Bon achat {bon.nbon}',
        'bon':bon,
        'items':items,
        'orderitems':orderitems,
        'payments':payments
    }
    return render(request, 'bonachatprint.html', ctx)

def refusedreglfc(request):
    reglfcid=request.GET.get('reglfcid')
    regl=PaymentClientfc.objects.get(pk=reglfcid)
    regl.refused=True
    regl.save()
    return JsonResponse({
        'success':True
    })

def refusedreglbl(request):
    reglblid=request.GET.get('reglblid')
    regl=PaymentClientbl.objects.get(pk=reglblid)
    regl.refused=True
    regl.save()
    return JsonResponse({
        'success':True
    })

def relevblprint(request):
    clientid=request.GET.get('clientid')
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    # print(clientid, startdate, enddate)
    # return JsonResponse({
    #     'rr':'rr'
    #     })
    client=Client.objects.get(pk=clientid)
    avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate])
    reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate])
    bons=Bonlivraison.objects.filter(client_id=clientid, date__range=[startdate, enddate], total__gt=0)

    releve = chain(*[
    ((bon, 'Bonlivraison') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return render(request, 'releveclprint.html', {
            'releve':[sorted_releve[i:i+36] for i in range(0, len(sorted_releve), 36)],
            'client':client,

            'startdate':startdate,
            'enddate':enddate,

        })

def relevsuppprint(request):
    supplierid=request.GET.get('supplierid')
    print('>>>', supplierid)
    supplier=Supplier.objects.get(pk=supplierid)
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    avoirs=Avoirsupplier.objects.filter(supplier_id=supplierid, avoirfacture=False, date__range=[startdate, enddate])
    reglementsbl = PaymentSupplier.objects.filter(
        supplier_id=supplierid,
        date__range=[startdate, enddate]
    ).filter(
        Q(bons__isfacture=False) | Q(bons__isnull=True)
    ).distinct()
    bons=Itemsbysupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfacture=False)

    releve = chain(*[
    ((bon, 'Bonlivraison') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return render(request, 'relevesuppprint.html', {
            'releve':[sorted_releve[i:i+30] for i in range(0, len(sorted_releve), 30)],
            'supplier':supplier,

            'startdate':startdate,
            'enddate':enddate,

        })


def relevsupppfcprint(request):
    supplierid=request.GET.get('supplierid')
    supplier=Supplier.objects.get(pk=supplierid)
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    avoirs=Avoirsupplier.objects.filter(supplier_id=supplierid, avoirfacture=True, date__range=[startdate, enddate])
    reglementsbl = PaymentSupplier.objects.filter(
        supplier_id=supplierid,
        date__range=[startdate, enddate],
        bons__isfacture=True
    ).distinct()
    bons=Itemsbysupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate], isfacture=True)
    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'Bonlivraison') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return render(request, 'relevesuppprint.html', {
            'releve':[sorted_releve[i:i+30] for i in range(0, len(sorted_releve), 30)],
            'supplier':supplier,

            'startdate':startdate,
            'enddate':enddate,

        })


def relevfcprint(request):
    clientid=request.GET.get('clientid')
    client=Client.objects.get(pk=clientid)
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=True, date__range=[startdate, enddate])
    reglementsfc=PaymentClientfc.objects.filter(client_id=clientid, date__range=[startdate, enddate])

    bons=Facture.objects.filter(client_id=clientid, date__range=[startdate, enddate])

    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'Facture') for bon in bons),
    ((avoir, 'Avoirclient') for avoir in avoirs),
    ((reglementfc, 'PaymentClientfc') for reglementfc in reglementsfc),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return render(request, 'releveclfcprint.html', {
            'releve':[sorted_releve[i:i+36] for i in range(0, len(sorted_releve), 36)],
            'client':client,
            'startdate':startdate,
            'enddate':enddate,
        })

def cancelpdctnew(request):
    pdctid=request.GET.get('pdctid')
    product=Produit.objects.get(pk=pdctid)
    product.isnew = False
    product.save()
    return JsonResponse({
        'success':True
        })

def makepdctnew(request):
    pdctid=request.GET.get('pdctid')
    product=Produit.objects.get(pk=pdctid)
    product.isnew = True
    product.save()
    return JsonResponse({
        'success':True
        })

def filterclients(request):
    rep=request.GET.get('rep')
    region=request.GET.get('region')
    print(rep, region)
    if region=="":
        clients=Client.objects.filter(represent_id=rep).order_by('-soldtotal')
    else:
        clients=Client.objects.filter(represent_id=rep, region__icontains=region).order_by('-soldtotal')
    print(clients.count())
    return JsonResponse({
        'success':True,
        'html':render(request, 'clientslist.html', {'clients':clients}).content.decode('utf-8')
        })

def pdctrepport(request):
    ctx={
    'title':'Rapport des produits',
    'categories':Category.objects.all()
    }
    return render(request, 'pdctrepport.html', ctx)

# this is used in pdct rapports, it gives the outs of the product
def getpdctins(request):

    ref=request.GET.get('ref').lower().strip()
    product=Produit.objects.filter(ref=ref).first()
    # avoirs
    products1 = Stockin.objects.filter(product=product)

    producthistory = Achathistory.objects.filter(ref=ref)
    products = chain(*[
        ((bon, 0) for bon in products1),
        ((b, 1) for b in producthistory)
    ])

    pdctins=sorted(products, key=lambda item: item[0].date)
    pdctins_serializable = [
        {
            'date': item[0].date.strftime('%d/%m/%Y'),
            'total': item[0].total,
            'supplier': item[0].supplier.name if item[1] == 0 else '--',
            'quantity': item[0].quantity,
            'price': f"{item[0].price:.2f}" if item[1] == 0 else item[0].prixunitaire,
            'devise': item[0].devise,
            'type': 'Stockin' if item[1] == 0 else 'Achathistory'
        }
        for item in pdctins
    ]
    total=round(products1.aggregate(Sum('total'))['total__sum'] or 0, 2)+round(producthistory.aggregate(Sum('total'))['total__sum'] or 0, 2)
    totalqty=(producthistory.aggregate(Sum('quantity'))['quantity__sum'] or 0)+(products1.aggregate(Sum('quantity'))['quantity__sum'] or 0)
    return JsonResponse({
        'success':True,
        'pdctins':pdctins_serializable,
        'totalqtyin':totalqty,
        'total':total
    })

# this is used in pdct rapports, it gives the outs of the product
def getpdctouts(request):
    ref=request.GET.get('ref').lower().strip()
    product=Produit.objects.filter(ref=ref).first()
    stockout=Livraisonitem.objects.filter(product=product, isfacture=False)
    stockoutfc=Outfacture.objects.filter(product=product).exclude(facture__bon__isnull=True)
    avoirs=Returned.objects.filter(product=product)

    totalqtyavoirs=avoirs.aggregate(Sum('qty'))['qty__sum'] or 0
    totalavoirs=avoirs.aggregate(Sum('total'))['total__sum'] or 0
    releve = chain(*[
    ((outbl, 0) for outbl in stockout),
    ((outfc, 1) for outfc in stockoutfc),
    ])
    fortable = chain(*[
    ((outbl, 0) for outbl in stockout),
    ((outfc, 1) for outfc in stockoutfc),
    ])
    pdctins=sorted(releve, key=lambda item: item[0].date)
    print('>>pdctins', pdctins)
    listfortable=sorted(fortable, key=lambda item: item[0].date, reverse=True)
    grouped_by_month = groupby(pdctins, key=lambda item: item[0].date.strftime('%m/%Y'))
    # Prepare data for frontend
    by_month = []
    for month, items in grouped_by_month:
        items=[i[0].qty for i in items]
        count = sum(items)  # Counting items in each group
        by_month.append({'month': month, 'count': count})
    pdctouts_serializable = [
        {
            'date': item[0].date.strftime('%d/%m/%Y'),
            'total': item[0].total,
            'price': f"{item[0].price:.2f}",
            'remise': item[0].remise,
            'client': item[0].bon.client.name if item[1]==0 else item[0].facture.client.name,
            'quantity': item[0].qty,
            'type': 'Livraisonitem' if item[1] == 0 else 'Outfacture',
            'no': item[0].bon.bon_no if item[1] == 0 else item[0].facture.facture_no,
        }
        for item in listfortable
    ]
    totalqty=(stockout.aggregate(Sum('qty'))['qty__sum'] or 0)+(stockoutfc.aggregate(Sum('qty'))['qty__sum'] or 0)
    total=round(stockout.aggregate(Sum('total'))['total__sum'] or 0, 2)+round(stockoutfc.aggregate(Sum('total'))['total__sum'] or 0, 2)
    # Group by client

    client_quantities= defaultdict(int)
    client_avoirs= defaultdict(int)
    for i in avoirs:
        client_name = i.avoir.client.name
        client_avoirs[client_name] += i.qty
    for item in pdctins:
        client_name = item[0].bon.client.name if item[1] == 0 else item[0].facture.client.name
        client_id = item[0].bon.client.id if item[1] == 0 else item[0].facture.client.id
        client_quantities[client_name] += item[0].qty
        #client_quantities[client_name][1] = Returned.objects.filter(avoir__client_id=client_id, product=product).aggregate(Sum('qty'))['qty__sum'] or 0
        #client_data[client_name]['quantity'] += item[0].qty

    print('>>>>>> client_quantities', client_avoirs)
    clients_quantities_serializable = sorted([
    {'client': client, 'quantity': quantity}
    for client, quantity in client_quantities.items()
    ], key=lambda x: x['quantity'], reverse=True)[:10]
    clients_avoirs_serializable = sorted([
    {'client': client, 'quantity': quantity}
    for client, quantity in client_avoirs.items()
    ], key=lambda x: x['quantity'], reverse=True)
    return JsonResponse({
        'pdctstock':product.stocktotal if product.stocktotal else '--',
        'pdctimg':product.image.url if product.image else '--',
        'pdctname':product.name,
        'success':True,
        'totalavoirs':totalavoirs,
        'pdctouts':pdctouts_serializable,
        'totalqtyout':totalqty,
        'totalqtyavoirs':totalqtyavoirs,
        'totalout':total,
        'outbymonth':by_month,
        'clientsqty':clients_quantities_serializable,
        'clientsavoirs':clients_avoirs_serializable
    })

def pdctscategoryrepport(request):
    # category = Category.objects.get(pk=request.POST.get('category'))
    # products = category.product.filter(category=category)[:10]
    # get ten products from the category
    products = Produit.objects.filter(category__pk=request.GET.get('categoryid')).order_by('-stocktotal')
    print(products)
    pdctdata=[]
    print('start loop')
    for i in products:
        stockout=Livraisonitem.objects.filter(product=i, isfacture=False)
        stockoutfc=Outfacture.objects.filter(product=i).exclude(facture__bon__isnull=True)
        totalqtyout=(stockout.aggregate(Sum('qty'))['qty__sum'] or 0)+(stockoutfc.aggregate(Sum('qty'))['qty__sum'] or 0)
        products1 = Stockin.objects.filter(product=i)
        producthistory = Achathistory.objects.filter(ref=i.ref)
        totalqtyin=(products1.aggregate(Sum('quantity'))['quantity__sum'] or 0)+(producthistory.aggregate(Sum('quantity'))['quantity__sum'] or 0)
        pdctdata.append({
        'id':i.id,
        'ref':i.ref,
        'name':i.name,
        'devise':i.devise if i.devise else 0,
        'arrivage':i.minstock if i.minstock else 0,
        'newfob':i.newfob,
        'qtycommande':i.qtycommande,
        'stocktotal':i.stocktotal,
        'totalqtyout':totalqtyout,
        'totalqtyin':totalqtyin,
        })
    print('end loop')
    return JsonResponse({
        'pdctdata':pdctdata
    })

def commandpdct(request):
    pdct=Produit.objects.get(pk=request.GET.get('pdctid'))
    qty=request.GET.get('qty') or 0
    pdct.qtycommande=qty
    pdct.save()
    return JsonResponse({
    'success':True,
    })
# minstock is now the arrivage, here to determine the quantity in port or arriving soon
def minstockpdct(request):
    pdct=Produit.objects.get(pk=request.GET.get('pdctid'))
    qty=request.GET.get('qty') or 0
    pdct.minstock=qty
    pdct.save()
    return JsonResponse({
    'success':True,
    })
# new fob is the new devise of the product
def newfob(request):
    pdct=Produit.objects.get(pk=request.GET.get('pdctid'))
    qty=request.GET.get('qty') or 0
    print(request.GET.get('pdctid'), qty)
    pdct.newfob=qty
    pdct.save()
    return JsonResponse({
        'success':True,
    })



def etude(request):
    thisyear=timezone.now().year
    ctx={
    'title':'Etude',
    'etudes':Etude.objects.filter(created_at__year=thisyear).order_by('-created_at')
    }
    return render(request, 'etudes.html', ctx)
# the add etude html
def addetude(request):
    ctx={
    'title':'Etude',
    'suppliers':Supplier.objects.all()
    }
    return render(request, 'addetude.html', ctx)
# create etude view (this actually creates teh etude)
def createetude(request):
    # Retrieve the data from the GET request
    date = request.POST.get('datefacture')
    print('date', date)
    facture_no = request.POST.get('facture_no')
    print('facture_no', facture_no)
    supplierid = request.POST.get('supplier')
    print('supplierid', supplierid)
    facturedevise = request.POST.get('facturedevise')
    print('facturedevise', facturedevise)
    tauxChange = request.POST.get('tauxchange')
    print('tauxChange', tauxChange)
    facturedh = request.POST.get('facturedh')
    print('facturedh', facturedh)
    chargeandfacture = request.POST.get('totalchargesandfacture')
    print('chargeandfacture', chargeandfacture)
    cfr = request.POST.get('tauxcfr')
    print('cfr', cfr)


    # Charges
    transportInternational = request.POST.get('transport_international') or 0
    print('transportInternational', transportInternational)
    dounane = request.POST.get('douane') or 0
    print('dounane', dounane)
    magazinage = request.POST.get('magazinage') or 0
    print('magazinage', magazinage)
    surrestaries = request.POST.get('surrestaries') or 0
    print('surrestaries', surrestaries)
    transportCamion = request.POST.get('trsp_camion') or 0
    print('transportCamion', transportCamion)
    transitaire = request.POST.get('forfait_transitaire') or 0
    print('transitaire', transitaire)
    autre1 = request.POST.get('autre_1') or 0
    print('autre1', autre1)
    autre2 = request.POST.get('autre_2') or 0
    print('autre2', autre2)

    # Total charges and tauxcharge
    totalCharges = request.POST.get('total_charges') or 0
    print('totalCharges', totalCharges)
    tauxCharge = request.POST.get('tauxcharge') or 0
    print('tauxCharge', tauxCharge)
    pattcQty = request.POST.get('pattcQty') or 0
    print('pattcQty', pattcQty)
    tdt = request.POST.get('tdt') or 0
    print('tdt', tdt)
    tcharge = request.POST.get('tcharge') or 0
    print('tcharge', tcharge)
    # create etude
    etude=Etude.objects.create(
    date=date,
    supplier_id=supplierid,
    facture_no=facture_no,
    facturedevise=facturedevise,
    tauxChange=tauxChange,
    facturedh=facturedh,
    chargeandfacture=chargeandfacture,
    cfr=cfr,
    transportInternational=transportInternational,
    dounane=dounane,
    magazinage=magazinage,
    surrestaries=surrestaries,
    transportCamion=transportCamion,
    transitaire=transitaire,
    autre1=autre1,
    autre2=autre2,
    totalCharges=totalCharges,
    tauxCharge=tauxCharge,
    pattcQty=pattcQty,
    tdt=tdt,
    tcharge=tcharge,
    )
    # Extract tableData if present
    table_data = json.loads(request.POST.get('tableData'))
    for i in table_data:
            EtudeItem.objects.create(etude=etude,
            ref=i['ref'],
            name=i['name'],
            qty=i['qty'],
            devise=i['devise'],
            amount=i['amount'],
            dh=i['dh'],
            hs=i['hs'],
            dt=i['dt'],
            pattc=i['pattc'],
            paht=i['paht'],
            coeff=i['coeff'],
            pbrut=i['pbrut'],
            pnet=i['pnet'],
            marge=i['marge'],
            tdt=i['tdt'],
            tcharges=i['tcharges'])
    print('>> table data', len(table_data))
    return JsonResponse({
    'success':True
    })

def updateetude(request):
    # Retrieve the data from the GET request
    date = request.POST.get('datefacture')
    print('date', date)
    facture_no = request.POST.get('facture_no')
    etudeid = request.POST.get('etudeid')
    etude=Etude.objects.get(pk=etudeid)
    print('facture_no', facture_no)
    supplierid = request.POST.get('supplier')
    print('supplierid', supplierid)
    facturedevise = request.POST.get('facturedevise')
    print('facturedevise', facturedevise)
    tauxChange = request.POST.get('tauxchange')
    print('tauxChange', tauxChange)
    facturedh = request.POST.get('facturedh')
    print('facturedh', facturedh)
    chargeandfacture = request.POST.get('totalchargesandfacture')
    print('chargeandfacture', chargeandfacture)
    cfr = request.POST.get('tauxcfr')
    print('cfr', cfr)


    # Charges
    transportInternational = request.POST.get('transport_international') or 0
    print('transportInternational in update', transportInternational)
    dounane = request.POST.get('douane') or 0
    print('dounane in update', dounane)
    magazinage = request.POST.get('magazinage') or 0
    print('magazinage in update', magazinage)
    surrestaries = request.POST.get('surrestaries') or 0
    print('surrestaries in update', surrestaries)
    transportCamion = request.POST.get('trsp_camion') or 0
    print('transportCamion in update', transportCamion)
    transitaire = request.POST.get('forfait_transitaire') or 0
    print('transitaire in update', transitaire)
    autre1 = request.POST.get('autre_1') or 0
    print('autre1 in update', autre1)
    autre2 = request.POST.get('autre_2') or 0
    print('autre2 in update', autre2)

    # Total charges and tauxcharge
    totalCharges = request.POST.get('total_charges') or 0
    print('totalCharges in update', totalCharges)
    tauxCharge = request.POST.get('tauxcharge') or 0
    print('tauxCharge in update', tauxCharge)
    pattcQty = request.POST.get('pattcQty') or 0
    print('pattcQty in update', pattcQty)
    tdt = request.POST.get('tdt') or 0
    print('tdt in update', tdt)
    tcharge = request.POST.get('tcharge') or 0
    print('tcharge in update', tcharge)
    # update etude
    etude.date=date
    etude.supplier_id=supplierid
    etude.facture_no=facture_no
    etude.facturedevise=facturedevise
    etude.tauxChange=tauxChange
    etude.facturedh=facturedh
    etude.chargeandfacture=chargeandfacture
    etude.cfr=cfr
    etude.transportInternational=transportInternational
    etude.dounane=dounane
    etude.magazinage=magazinage
    etude.surrestaries=surrestaries
    etude.transportCamion=transportCamion
    etude.transitaire=transitaire
    etude.autre1=autre1
    etude.autre2=autre2
    etude.totalCharges=totalCharges
    etude.tauxCharge=tauxCharge
    etude.pattcQty=pattcQty
    etude.tdt=tdt
    etude.tcharge=tcharge
    etude.save()
    EtudeItem.objects.filter(etude_id=etudeid).delete()
    # Extract tableData if present
    table_data = json.loads(request.POST.get('tableData'))
    for i in table_data:
            EtudeItem.objects.create(etude_id=etudeid,
            ref=i['ref'],
            name=i['name'],
            qty=i['qty'],
            devise=i['devise'],
            amount=i['amount'],
            dh=i['dh'],
            hs=i['hs'],
            dt=i['dt'],
            pattc=i['pattc'],
            paht=i['paht'],
            coeff=i['coeff'],
            pbrut=i['pbrut'],
            pnet=i['pnet'],
            marge=i['marge'],
            tdt=i['tdt'],
            tcharges=i['tcharges'])
    print('>> table data', len(table_data))
    return JsonResponse({
    'success':True
    })


# the update etude html
def updateetudepage(request):
    etudeid=request.GET.get('etudeid')
    etude=Etude.objects.get(pk=etudeid)
    etudeitems=EtudeItem.objects.filter(etude=etude)
    ctx={
    'title':'Modifier Etude',
    'etude':etude,
    'etudeitems':etudeitems,
    'suppliers':Supplier.objects.all()
    }
    return render(request, 'updateetudepage.html', ctx)



def boncomparer(request):
    #bons=[f'BL240{i}' for i in range(5000, 6000)]
    bons=Bonlivraison.objects.all()
    bb=[]
    for i in bons:
        items=Livraisonitem.objects.filter(bon=i, isfacture=False)
        totals=[b.total for b in items]
        if not i.total==round(sum(totals), 2):
            bb.append((i.bon_no, i.total, round(sum(totals), 2)))
            print('>>>', i.bon_no, i.total, round(sum(totals), 2))
    return JsonResponse({
    'success':True,
    'data':bb
    })

def getachatfacture(request):
    thisyear=timezone.now().year
    year=request.GET.get('year') or thisyear
    print('year >>', year)
    bons=Itemsbysupplier.objects.filter(date__year=year, isfacture=True).order_by('-date')
    print(bons)
    total=round(bons.aggregate(Sum('total'))['total__sum'] or 0, 2)
    totaltva=round(bons.aggregate(Sum('tva'))['tva__sum'] or 0, 2)
    trs=''
    for order in bons:
        trs+=f'''
            <tr class="" orderid="{order.id}" ondblclick="createtab('bonachat{order.id}', 'Bon achat {order.nbon}', '/products/bonachatdetails/{order.id}')">
            <td>{ order.nbon }</td>
            <td>{ order.date.strftime("%d/%m/%Y") }</td>
            <td>{ order.supplier.name }</td>
            <td>{ order.total}</td>
            <td>{ order.tva}</td>
            <td>{"Facture"if order.isfacture else "Bl"}</td>
            <td>{ round(order.supplier.rest, 2) }</td>
            <td class="d-flex">
                <div>{"R0"if order.ispaid else "N1"}</div>
                <div style="width:15px; height:15px; border-radius:50%; background:{"green"if order.ispaid else "red"};" ></div>
            </td>
            <td>
              <button class="btn bi bi-download" onclick="printbonachat('{order.id}')"></button>
            </td>

          </tr>
            '''


    return JsonResponse({
    'success':True,
    'trs':trs,
    'total':total
    })
def etatblfc(request):
    current_year = datetime.now().year
    # Create a date object for the first day of the current year
    first_day_of_year = date(current_year, 1, 1)
    now = datetime.now()
    # Get the last day of the current month
    last_day_of_month = calendar.monthrange(now.year, now.month)[1]
    # Create a date object for the last day of the current month
    last_day_of_month = date(now.year, now.month, last_day_of_month)
    print('>>>>>ddd', first_day_of_year, last_day_of_month)
    start_date_str = request.GET.get('monthtostart',first_day_of_year.strftime('%Y-%m-%d'))
    end_date_str = request.GET.get('monthtoend', last_day_of_month.strftime('%Y-%m-%d'))
    rep=request.GET.get('rep')
    region=request.GET.get('region', '-').lower()
    # Parse dates
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    # Generate list of months between start_date and end_date
    months = []
    current = start_date

    while current <= end_date:
        print('>> current', current)
        months.append(current.strftime('%m/%y'))  # e.g., "01/24" for Jan 2024
        # Move to the next month
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)
    print('>>>>>>>>>>>>>>>>>>>>>>>> months', months)
    return render(request, 'etatblfc.html', {'title': 'Etat BL/FC client', 'reps':Represent.objects.all(), 'months': months, 'monthtostart': start_date_str, 'monthtoend': end_date_str})

def getetatblfc(request):
    current_year = datetime.now().year
    # Create a date object for the first day of the current year
    first_day_of_year = date(current_year, 1, 1)
    now = datetime.now()
    # Get the last day of the current month
    last_day_of_month = calendar.monthrange(now.year, now.month)[1]
    # Create a date object for the last day of the current month
    last_day_of_month = date(now.year, now.month, last_day_of_month)
    print('>>>>>', first_day_of_year, last_day_of_month)
    start_date_str = request.GET.get('monthtostart',first_day_of_year.strftime('%Y-%m-%d'))
    end_date_str = request.GET.get('monthtoend', last_day_of_month.strftime('%Y-%m-%d'))
    rep=request.GET.get('rep')
    region=request.GET.get('region', '-').lower()
    # Parse dates
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    # Generate list of months between start_date and end_date
    months = []
    current = start_date
    while current <= end_date:
        months.append(current.strftime('%m/%y'))  # e.g., "January 2024"
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)
    print('>> region: ', region, months)
    if region == '':
        clients = Client.objects.filter(represent_id=rep).order_by('code').exclude(diver=True).exclude(name__istartswith='.').exclude(name__istartswith='test')
    else:
        clients = Client.objects.filter(represent_id=rep, region=region).order_by('code').exclude(diver=True).exclude(name__istartswith='.').exclude(name__istartswith='test')
    print('>> clients', clients)
    serialized_data = []
    #client=Client.objects.get(pk=3758)
    totalsoldbl=0
    totalsoldfc=0
    for clientindex, client in enumerate(clients):
        sitdata=0
        client_data = {'client_name': client.name, 'client_id': client.id, 'client_code': client.code, 'client_city': client.city, 'client_region': client.region, 'client_moderegl': client.moderegl, 'client_represent': client.represent.name, 'monthly_data': [], 'totalsituation': 0, 'soldfc':client.soldfacture, 'soldbl':client.soldbl}

        monthly_data = {month: {'bons': 0, 'avoirs': 0, 'regls': 0, 'situation':0, 'factures':0, 'avoirfc':0, 'reglfc':0, 'situationfc':0} for month in months}
        # Filter data for the specified date range
        bons = Bonlivraison.objects.filter(
            client=client,
            date__range=[start_date, end_date],
            total__gt=0
        )
        avoirs = Avoirclient.objects.exclude(avoirfacture=True).filter(
            client=client,
            date__range=[start_date, end_date]
        )
        regls = PaymentClientbl.objects.filter(
            client=client,
            date__range=[start_date, end_date]
        )

        factures = Facture.objects.filter(
            client=client,
            date__range=[start_date, end_date],
            total__gt=0
        )
        avoirfc = Avoirclient.objects.filter(
            client=client,
            avoirfacture=True,
            date__range=[start_date, end_date]
        )
        reglfc = PaymentClientfc.objects.filter(
            client=client,
            date__range=[start_date, end_date]
        )

        # Initialize monthly data

        # Populate monthly data with factures, avoirs, and regls
        for bon in factures:
            month = bon.date.strftime('%m/%y')
            monthly_data[month]['factures'] += bon.total
            sitdata+=bon.total

        for avoir in avoirfc:
            month = avoir.date.strftime('%m/%y')
            monthly_data[month]['avoirfc'] += avoir.total
            sitdata-=bon.total

        for regl in reglfc:
            month = regl.date.strftime('%m/%y')
            monthly_data[month]['reglfc'] += regl.amount
            sitdata-=bon.total
        # Initialize monthly data

        # Populate monthly data with bons, avoirs, and regls
        for bon in bons:
            month = bon.date.strftime('%m/%y')
            monthly_data[month]['bons'] += bon.total
            # don't round the total, because if it's rounded the situations will be different in sit.CL
        for avoir in avoirs:
            month = avoir.date.strftime('%m/%y')
            monthly_data[month]['avoirs'] += avoir.total

        for regl in regls:
            month = regl.date.strftime('%m/%y')
            monthly_data[month]['regls'] += regl.amount

        # Calculate the client situation for each month and aggregate totals
        total_bons = total_avoirs = total_regls = 0
        total_factures = total_avoirfc = total_reglfc = 0
        # this for testing specific clients
        # if client.id==3785:
        #     print('>>>>>>>>> bons', bons, 'avoirs', avoirs, 'regls', regls, )
        for monthindex, month in enumerate(months):
            # BL
            total_bons += monthly_data[month]['bons']
            total_avoirs += monthly_data[month]['avoirs']
            total_regls += monthly_data[month]['regls']
            #monthly_data[month]['situation'] = round(monthly_data[month]['bons'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
            thismonthsit=round(monthly_data[month]['bons'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
            thisreg=monthly_data[month]['regls']
            if thisreg:
                for b in range(monthindex):
                    thisreg-=client_data['monthly_data'][b]['situation']
                    client_data['monthly_data'][b]['situation']=0
            thismonthsit=round(monthly_data[month]['bons'] - monthly_data[month]['avoirs'] - thisreg, 2)


            # client_data['monthly_data'].append({
            #     'month': month,
            #     'bons': monthly_data[month]['bons'],
            #     'avoirs': monthly_data[month]['avoirs'],
            #     'regls': monthly_data[month]['regls'],
            #     'situation': thismonthsit
            # })
            # end bl
            total_factures += monthly_data[month]['factures']
            total_avoirfc += monthly_data[month]['avoirfc']
            total_reglfc += monthly_data[month]['reglfc']
            #monthly_data[month]['situation'] = round(monthly_data[month]['factures'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
            thismonthsitfc=round(monthly_data[month]['factures'] - monthly_data[month]['avoirfc'] - monthly_data[month]['reglfc'], 2)
            thisregfc=monthly_data[month]['reglfc']
            if thisregfc:
                for b in range(monthindex):
                    thisregfc-=client_data['monthly_data'][b]['situationfc']
                    client_data['monthly_data'][b]['situationfc']=0
            thismonthsitfc=round(monthly_data[month]['factures'] - monthly_data[month]['avoirfc'] - thisregfc, 2)



            client_data['monthly_data'].append({
                'month': month,
                'factures': monthly_data[month]['factures'],
                'avoirfc': monthly_data[month]['avoirfc'],
                'reglfc': monthly_data[month]['reglfc'],
                'situationfc': thismonthsitfc,
                'bons': monthly_data[month]['bons'],
                'avoirs': monthly_data[month]['avoirs'],
                'regls': monthly_data[month]['regls'],
                'situation': thismonthsit
            })

            # Calculate total situation for the client
        client_data['totalsituationfc'] = round(total_factures - total_avoirfc - total_reglfc, 2)
        totalsoldfc+=round(total_factures - total_avoirfc - total_reglfc, 2)
        # Calculate total situation for the client
        client_data['totalsituation'] = round(total_bons - total_avoirs - total_regls, 2)
        serialized_data.append(client_data)
        totalsoldbl+=round(total_bons - total_avoirs - total_regls, 2)

        # Define start and end months for the date range
        # sitdata=0
        # client_data = {'client_name': client.name, 'client_code': client.code, 'client_city': client.city, 'client_region': client.region, 'client_represent': client.represent, 'monthly_data': [], 'totalsituation': 0}
        #
        # # Filter data for the specified date range
        # factures = Facture.objects.filter(
        #     client=client,
        #     date__range=[start_date, end_date],
        #     total__gt=0
        # )
        #
        # avoirs = Avoirclient.objects.filter(
        #     client=client,
        #     avoirfacture=True,
        #     date__range=[start_date, end_date]
        # )
        #
        # regls = PaymentClientfc.objects.filter(
        #     client=client,
        #     date__range=[start_date, end_date]
        # )
        #
        # # Initialize monthly data
        # monthly_data = {month: {'factures': 0, 'avoirs': 0, 'regls': 0, 'situation':0} for month in months}
        #
        # # Populate monthly data with factures, avoirs, and regls
        # for bon in factures:
        #     month = bon.date.strftime('%m/%y')
        #     monthly_data[month]['factures'] += bon.total
        #     sitdata+=bon.total
        #
        # for avoir in avoirs:
        #     month = avoir.date.strftime('%m/%y')
        #     monthly_data[month]['avoirs'] += avoir.total
        #     sitdata-=bon.total
        #
        # for regl in regls:
        #     month = regl.date.strftime('%m/%y')
        #     monthly_data[month]['regls'] += regl.amount
        #     sitdata-=bon.total
        # # Calculate the client situation for each month and aggregate totals
        # total_factures = total_avoirs = total_regls = 0
        #
        # for monthindex, month in enumerate(months):
        #     total_factures += monthly_data[month]['factures']
        #     total_avoirs += monthly_data[month]['avoirs']
        #     total_regls += monthly_data[month]['regls']
        #     #monthly_data[month]['situation'] = round(monthly_data[month]['factures'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
        #     thismonthsit=round(monthly_data[month]['factures'] - monthly_data[month]['avoirs'] - monthly_data[month]['regls'], 2)
        #     thisreg=monthly_data[month]['regls']
        #     if thisreg:
        #         for b in range(monthindex):
        #             thisreg-=client_data['monthly_data'][b]['situation']
        #             client_data['monthly_data'][b]['situation']=0
        #     thismonthsit=round(monthly_data[month]['factures'] - monthly_data[month]['avoirs'] - thisreg, 2)
        #
        #
        #
        #     client_data['monthly_data'].append({
        #         'month': month,
        #         'factures': monthly_data[month]['factures'],
        #         'avoirs': monthly_data[month]['avoirs'],
        #         'regls': monthly_data[month]['regls'],
        #         'situation': thismonthsit
        #     })
        #
        # # Calculate total situation for the client
        # client_data['totalsituation'] = round(total_factures - total_avoirs - total_regls, 2)
        # serialized_data.append(client_data)


    print('>>> soldbl', totalsoldbl)
    print('>>> soldfc', totalsoldfc)
    return JsonResponse({
        'trs':render(request, 'etatblfctrs.html', {'data': serialized_data, 'months': months, 'monthtostart': start_date_str, 'monthtoend': end_date_str, 'totalsoldbl':totalsoldbl, 'totalsoldfc':totalsoldfc}).content.decode('utf-8'),
        'totalsoldbl':totalsoldbl,
        'totalsoldfc':totalsoldfc,
    })


def excelecheaces(request):
    # create new record to get the id
    thisyear=timezone.now().year
    tvas=Tva.objects.filter(month__icontains=thisyear)
    try:
        lastcode = Client.objects.order_by('code').last()
        print('lastcode', lastcode.code)
        if lastcode:

            codecl = f"{int(lastcode.code) + 1:06}"
        else:
            codecl = f"000001"
    except:
        codecl="000001"
    ctx={
        #'lastcode':codecl,
        'today':timezone.now().date(),
        'title':'Tva calculations',
        #'tvatrs':render(request, 'tvatrs.html', {'tvas':tvas}),
    }
    return render(request, 'excelecheances.html', ctx)


def getnpiecedata(request):
    npiece=request.GET.get('npiece').lower()
    print('>>> npiece', npiece)
    data={
    'success':True,
    'echance':'',
    'client':'',
    'codeclient':'',
    'mode':'',
    'factures':'.', #this is related to CDDF124.fmkg
    'amount':'',
    'tva':'',
    'facturealreadyexist':False #exist the same facture in the past months
    }
    try:
        reg = PaymentClientfc.objects.filter(npiece=npiece).first()
        data['echance']=reg.echance.strftime('%d/%m/%Y')
        data['client']=reg.client.name
        data['codeclient']=reg.client.code
        data['mode']=reg.mode
        data['amount']=reg.amount
        data['factures'] = ', '.join([facture_no.replace('FC', '', 1) for facture_no in reg.factures.values_list('facture_no', flat=True)])
        data['facturealreadyexist']=Excelecheances.objects.filter(factures=', '.join([facture_no.replace('FC', '', 1) for facture_no in reg.factures.values_list('facture_no', flat=True)])).exists()

        data['tva']=round(reg.amount-(reg.amount/1.2), 2)
    except Exception as e:
        print('>>> Except of getting reg fc', e)
        try:
            reg = PaymentClientbl.objects.filter(npiece=npiece).first()
            print('>>> reg in bls', reg)
            data['echance']=reg.echance.strftime('%d/%m/%Y')
            data['client']=reg.client.name
            data['codeclient']=reg.client.code
            data['mode']=reg.mode
            data['amount']=reg.amount
            data['tva']=round(reg.amount-(reg.amount/1.2), 2)
        except:
            pass
    return JsonResponse(data)

def saverowech(request):
    facturesval=request.GET.get('facturesval').strip()
    id=request.GET.get('id')
    iscontable=True if request.GET.get('iscontable')=="true" else False
    # check if ther is a facture alread printrd, if so (2)
    #factureprintedalready=Excelecheances.objects.filter(factures=facturesval, isprinted=True).exists()
    # we don't need the . in factures
    factureprintedalready=Excelecheances.objects.filter(factures=facturesval).exists() and len(facturesval)>3

    #print('>> factureprintedalready', factureprintedalready)
    impye=True if request.GET.get('impye')=='true' else False
    nomde=request.GET.get('nomde')
    regle=True if request.GET.get('regle')=='true' else False
    isempty=False if request.GET.get('isempty')=='False' else True
    ispointage=True if request.GET.get('pointage')=='true' else False
    npieceval=request.GET.get('npieceval')
    mode=request.GET.get('mode')
    tva=request.GET.get('tva') or 0.00
    amount=request.GET.get('amount') or 0.00
    # used to get code from ajax
    code=request.GET.get('code')
    #code=uuid.uuid4().hex
    client=request.GET.get('client')
    codeclient=request.GET.get('codeclient')
    echeance=request.GET.get('echeance')
    monthyear=request.GET.get('monthyear')
    exists=Excelecheances.objects.filter(npiece=npieceval).exclude(pk=id).exists()
    if exists:
        print("hereX>>>>>>>>>>> already exist")
        return JsonResponse({
            'success':False,
            'message':'Deja exist'
        })

    # try to get this record
    try:
        data=Excelecheances.objects.get(pk=code)
        print('>>>>>> this', data)
        data.month=monthyear # Assuming 'monthyear' is in the format '09/2024'
        data.npiece=npieceval
        data.mode=mode
        data.tva=tva
        data.amount=amount
        data.echeance=echeance
        data.note=nomde
        data.client=client
        data.clientcode=codeclient
        data.factures=facturesval
        data.ispaid=regle   # Example default values
        data.isimpye=impye   # Example default values
        data.isempty=isempty
        data.ispointage=ispointage

        data.save()
    # creatte if not found
    except Exception as e:
        print('>>>>>>>', e)
        data=Excelecheances.objects.create(
            month=monthyear,  # Assuming 'monthyear' is in the format '09/2024'
            npiece=npieceval,
            mode=mode,
            echeance=echeance,
            note=nomde,
            client=client,
            amount=amount,
            tva=tva,
            clientcode=codeclient,
            factures=facturesval,
            ispaid=regle,   # Example default values
            isimpye=impye,   # Example default values
            isempty=isempty,  # Example default value
            ispointage=ispointage,  # Example default value
            isprinted=factureprintedalready,
            iscontable=factureprintedalready
        )
    return JsonResponse({
        'success':True,
        # code will be id
        'thisid':data.id,
        'nextid':Excelecheances.objects.last().id+1,
        'factureprintedalready':factureprintedalready
    })
def getmonthecheances(request):
    month=request.GET.get('month')
    data=Excelecheances.objects.filter(month=month).order_by('client')
    # check if total bank is the same as in this month
    bankrelve=data.filter(ispaid=True)
    print('>> relve', bankrelve.aggregate(Sum('amount'))['amount__sum'] or 0)
    unroundedtotal=data.aggregate(Sum('amount'))['amount__sum'] or 0
    total=round(unroundedtotal, 2)
    #totaltva=round(data.aggregate(Sum('tva'))['tva__sum'] or 0, 2)
    totaltva=round(total-(total/1.2), 2)
    idsource=Excelecheances.objects.filter(npiece=None).first() if Excelecheances.objects.filter(npiece=None) else Excelecheances.objects.create()
    grouped_data = {}
    for item in data:
        grandtotal = item.grandtotal
        if grandtotal not in grouped_data:
            grouped_data[grandtotal] = []
        grouped_data[grandtotal].append(item)
    print('>>>>> month', month)
    return JsonResponse({
        'trs':render(request, 'echeancestrs.html', {'grouped_data': grouped_data, 'data':data, 'code':idsource.id}).content.decode('utf-8'),
        'total':total,
        'totaltva':totaltva,
        'month':str(month)
    })
# gather muliple
def grouper(request):
    ids=json.loads(request.GET.get('ids'))
    print('>>>>>>', ids)
    echeances=Excelecheances.objects.filter(pk__in=ids)
    grandtotal=round(echeances.aggregate(Sum('amount'))['amount__sum'] or 0, 2)
    code=uuid.uuid4().hex
    echeances.update(code=code, grandtotal=grandtotal)

    return JsonResponse({
    'success':True
    })

def relevclientglobal(request):
    clientid=request.POST.get('clientid')
    client=Client.objects.get(pk=clientid)
    startdate=request.POST.get('datefrom')
    enddate=request.POST.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate])
    reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate])
    bons=Bonlivraison.objects.filter(client_id=clientid, date__range=[startdate, enddate], total__gt=0)
    norangeavoirsfc=round(Avoirclient.objects.filter(client_id=clientid, avoirfacture=True).aggregate(Sum('total'))['total__sum'] or 0, 2)

    norangereglementsfc=round(PaymentClientfc.objects.filter(client_id=clientid).aggregate(Sum('amount'))['amount__sum'] or 0, 2)

    norangefactures=round(Facture.objects.filter(client_id=clientid).aggregate(Sum('total'))['total__sum'] or 0, 2)



    norangeavoirsbl=round(Avoirclient.objects.filter(client_id=clientid, avoirfacture=False).aggregate(Sum('total'))['total__sum'] or 0, 2)
    norangereglementsbl=round(PaymentClientbl.objects.filter(client_id=clientid).aggregate(Sum('amount'))['amount__sum'] or 0, 2)
    norangebons=round(Bonlivraison.objects.filter(client_id=clientid, total__gt=0).aggregate(Sum('total'))['total__sum'] or 0, 2)

    soldfcnorange=round(norangefactures-norangereglementsfc-norangeavoirsfc, 2)

    soldblnorange=round(norangebons-norangereglementsbl-norangeavoirsbl, 2)
    print('>> soldbl', soldblnorange)
    print('>> soldfc', soldfcnorange)
    client.soldfacture=soldfcnorange
    client.soldbl=soldblnorange
    client.soldtotal=round(soldfcnorange+soldblnorange, 2)
    client.save()
    # totalcredit=round(avoirs.aggregate(Sum('total'))['total__sum'], 2)+round(reglementsbl.aggregate(Sum('amount'))['amount__sum'], 2)
    # totaldebit=round(bons.aggregate(Sum('total'))['total__sum'], 2)
    # sold=round(totaldebit-totalcredit, 2)

    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'bon') for bon in bons),
    ((avoir, 'Avoirclientbl') for avoir in avoirs),
    ((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    ## factures
    startdatefc=request.POST.get('datefromfc')
    enddatefc=request.POST.get('datetofc')
    startdatefc = datetime.strptime(startdatefc, '%Y-%m-%d')
    enddatefc = datetime.strptime(enddatefc, '%Y-%m-%d')
    avoirsfc=Avoirclient.objects.filter(client_id=clientid, avoirfacture=True, date__range=[startdatefc, enddatefc])
    reglementsfc=PaymentClientfc.objects.filter(client_id=clientid, date__range=[startdatefc, enddatefc])

    factures=Facture.objects.filter(client_id=clientid, date__range=[startdatefc, enddatefc])

    norangeavoirsfc=round(Avoirclient.objects.filter(client_id=clientid, avoirfacture=True).aggregate(Sum('total'))['total__sum'] or 0, 2)
    norangereglementsfc=round(PaymentClientfc.objects.filter(client_id=clientid).aggregate(Sum('amount'))['amount__sum'] or 0, 2)
    norangefactures=round(Facture.objects.filter(client_id=clientid).aggregate(Sum('total'))['total__sum'] or 0, 2)
    soldfacture=round(float(norangefactures)-float(norangeavoirsfc)-float(norangereglementsfc), 2)
    client.soldfacture=soldfacture
    client.save()
    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'Facturefc') for bon in factures),
    ((avoir, 'Avoirclientfc') for avoir in avoirsfc),
    ((reglementfc, 'PaymentClientfc') for reglementfc in reglementsfc),
    ])

    # Sort the items by date
    sorted_relevefc = sorted(releve, key=lambda item: item[0].date)
    releveglobal=sorted_relevefc+sorted_releve
    totaldebit=round(round(bons.aggregate(Sum('total')).get('total__sum') or 0, 2)+round(factures.aggregate(Sum('total')).get('total__sum') or 0, 2), 2)
    totalcredit=round(round(avoirs.aggregate(Sum('total')).get('total__sum') or 0, 2)+round(avoirsfc.aggregate(Sum('total')).get('total__sum') or 0, 2)+round(reglementsbl.aggregate(Sum('amount')).get('amount__sum') or 0, 2)+round(reglementsfc.aggregate(Sum('amount')).get('amount__sum') or 0, 2), 2)
    sold=round(totaldebit-totalcredit, 2)
    return JsonResponse({
        'html':render(request, 'releveclglobal.html', {
            'releve':[releveglobal[i:i+35] for i in range(0, len(releveglobal), 35)],
            'client':client,
            'startdate':startdate,
            'enddate':enddate,
            'totaldebit':totaldebit,
            'totalcredit':totalcredit,
            'sold':sold
        }).content.decode('utf-8')
    })


def printrelevclientglobal(request):
    clientid=request.GET.get('clientid')
    client=Client.objects.get(pk=clientid)
    print('>> client', client)
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate])
    reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate])
    bons=Bonlivraison.objects.filter(client_id=clientid, date__range=[startdate, enddate], total__gt=0)
    norangeavoirsfc=round(Avoirclient.objects.filter(client_id=clientid, avoirfacture=True).aggregate(Sum('total'))['total__sum'] or 0, 2)

    norangereglementsfc=round(PaymentClientfc.objects.filter(client_id=clientid).aggregate(Sum('amount'))['amount__sum'] or 0, 2)

    norangefactures=round(Facture.objects.filter(client_id=clientid).aggregate(Sum('total'))['total__sum'] or 0, 2)



    norangeavoirsbl=round(Avoirclient.objects.filter(client_id=clientid, avoirfacture=False).aggregate(Sum('total'))['total__sum'] or 0, 2)
    norangereglementsbl=round(PaymentClientbl.objects.filter(client_id=clientid).aggregate(Sum('amount'))['amount__sum'] or 0, 2)
    norangebons=round(Bonlivraison.objects.filter(client_id=clientid, total__gt=0).aggregate(Sum('total'))['total__sum'] or 0, 2)

    soldfcnorange=round(norangefactures-norangereglementsfc-norangeavoirsfc, 2)

    soldblnorange=round(norangebons-norangereglementsbl-norangeavoirsbl, 2)
    print('>> soldbl', soldblnorange)
    print('>> soldfc', soldfcnorange)
    client.soldfacture=soldfcnorange
    client.soldbl=soldblnorange
    client.soldtotal=round(soldfcnorange+soldblnorange, 2)
    client.save()
    # totalcredit=round(avoirs.aggregate(Sum('total'))['total__sum'], 2)+round(reglementsbl.aggregate(Sum('amount'))['amount__sum'], 2)
    # totaldebit=round(bons.aggregate(Sum('total'))['total__sum'], 2)
    # sold=round(totaldebit-totalcredit, 2)

    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'bon') for bon in bons),
    ((avoir, 'Avoirclientbl') for avoir in avoirs),
    ((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    ## factures
    startdatefc=request.GET.get('datefromfc')
    enddatefc=request.GET.get('datetofc')
    startdatefc = datetime.strptime(startdatefc, '%Y-%m-%d')
    enddatefc = datetime.strptime(enddatefc, '%Y-%m-%d')
    avoirsfc=Avoirclient.objects.filter(client_id=clientid, avoirfacture=True, date__range=[startdatefc, enddatefc])
    reglementsfc=PaymentClientfc.objects.filter(client_id=clientid, date__range=[startdatefc, enddatefc])

    factures=Facture.objects.filter(client_id=clientid, date__range=[startdatefc, enddatefc])

    norangeavoirsfc=round(Avoirclient.objects.filter(client_id=clientid, avoirfacture=True).aggregate(Sum('total'))['total__sum'] or 0, 2)
    norangereglementsfc=round(PaymentClientfc.objects.filter(client_id=clientid).aggregate(Sum('amount'))['amount__sum'] or 0, 2)
    norangefactures=round(Facture.objects.filter(client_id=clientid).aggregate(Sum('total'))['total__sum'] or 0, 2)
    soldfacture=round(float(norangefactures)-float(norangeavoirsfc)-float(norangereglementsfc), 2)
    client.soldfacture=soldfacture
    client.save()
    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'Facturefc') for bon in factures),
    ((avoir, 'Avoirclientfc') for avoir in avoirsfc),
    ((reglementfc, 'PaymentClientfc') for reglementfc in reglementsfc),
    ])

    # Sort the items by date
    sorted_relevefc = sorted(releve, key=lambda item: item[0].date)
    releveglobal=sorted_relevefc+sorted_releve
    totaldebit=round(round(bons.aggregate(Sum('total')).get('total__sum') or 0, 2)+round(factures.aggregate(Sum('total')).get('total__sum') or 0, 2), 2)
    totalcredit=round(round(avoirs.aggregate(Sum('total')).get('total__sum') or 0, 2)+round(avoirsfc.aggregate(Sum('total')).get('total__sum') or 0, 2)+round(reglementsbl.aggregate(Sum('amount')).get('amount__sum') or 0, 2)+round(reglementsfc.aggregate(Sum('amount')).get('amount__sum') or 0, 2), 2)
    sold=round(totaldebit-totalcredit, 2)
    return render(request, 'releveclglobal.html', {
            'releve':[releveglobal[i:i+35] for i in range(0, len(releveglobal), 35)],
            'client':client,
            'startdate':startdate,
            'enddate':enddate,
            'totaldebit':totaldebit,
            'totalcredit':totalcredit,
            'sold':sold
        })
def refactive(request):
    products=Produit.objects.filter(stockstocktotal__gt=0)
def updateiscontrebon(request):
    paymenttype=request.GET.get('iscontre')
    blid=request.GET.get('blid')
    bon=Bonlivraison.objects.get(pk=blid)
    bon.paymenttype=paymenttype
    bon.save()
    return JsonResponse({
        'success':True
        })
def getcontrenonpaid(request):
    bons = Bonlivraison.objects.filter(iscontre=True, ispaid=False, total__gt=0).order_by('-date')


    return JsonResponse({
        'html':render(request, 'bllist.html', {'bons':bons, 'notloading':True}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total')).get('total__sum') or 0, 2),

    })


def minidashboard(request):
    return render(request, 'minidashboard.html')

def configuration(request):
    setting=Setting.objects.first()
    ctx={
        'title':'Configuration'
    }
    if setting:
        ctx['setting']=setting
    return render(request, 'settings.html', ctx) 

def commandfromserver(request):
    serverip= Setting.objects.only('serverip').first()
    serverip=serverip.serverip if serverip else None
    if serverip:
        res=req.get(f'http://{serverip}/')
    items=request.GET.get('items')
    clientcode=request.GET.get('clientcode')
    total=request.GET.get('total')
    notesorder=request.GET.get('notesorder')
    #fix here, get rep from the
    userid=request.GET.get('userid')
    rep=request.GET.get('rep')
    client=Client.objects.get(code=clientcode)
    cmndfromclient=True if request.GET.get('cmndfromclient') == 'true' else False
    print('>>>>>>>>>>>>cc',items)
    print('>>>>>>>>>>>>client',cmndfromclient)
    print('>>>>>>>>>>>>total',total)
    print('>>>>>>>>>>>>represent', client.represent)
    if cmndfromclient:
        order=Order.objects.create(client=client, salseman=client.represent,  modpymnt='--', modlvrsn='--',total=total, isclientcommnd=True, note=notesorder)
        print('>> order from client')
    else:
        order=Order.objects.create(client=client, salseman_id=rep,  modpymnt='--', modlvrsn='--',total=total, note=notesorder)
        print('>> order from rep')

    #Ordersnotif.objects.create(user_id=1)

    # totalremise=request.POST.get('totalremise', 0)
    for i in json.loads(items):
        print(i['ref'])
        Orderitem.objects.create(order=order, ref=i['ref'], name=i['name'], qty=int(i['qty']), product_id=i['productid'], remise=i['remise'], price=i['price'], total=i['total'])
        print('>> create order item', i['name'])
    
    return JsonResponse({
        'success':True
    })

